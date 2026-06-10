"""
CAD图纸PDF识别系统 - 第二阶段：视觉识别（高速并发版）
================================================================
"""

import os
import re
import sys
import json
import math
import time
import base64
import argparse
from pathlib import Path
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor, as_completed

import fitz
import numpy as np

from cad_pdf_recognizer import PDFParser, CADInfoExtractor, ExcelExporter


# ═════════════════════════════════════════════
# 1. 分块渲染器（不变）
# ═════════════════════════════════════════════

class TileRenderer:
    """超大幅图纸 → 高清重叠分块（自动跳空白、限块数）"""

    def __init__(self, pdf_path, dpi=100, tile_px=1568, overlap=0.10,
                 max_tiles=36, ink_threshold=0.0015):
        self.pdf_path = pdf_path
        self.dpi = dpi
        self.tile_px = tile_px
        self.overlap = overlap
        self.max_tiles = max_tiles
        self.ink_threshold = ink_threshold
        # 每页渲染几何，供标注阶段复用： {page_no: {"scale","w_px","h_px"}}
        self.page_dims = {}

    def _plan(self, page):
        scale = self.dpi / 72
        step = self.tile_px * (1 - self.overlap)
        px_w = page.rect.width * scale
        px_h = page.rect.height * scale
        cols = max(1, math.ceil((px_w - self.tile_px) / step) + 1)
        rows = max(1, math.ceil((px_h - self.tile_px) / step) + 1)
        return scale, step, cols, rows

    def render(self):
        doc = fitz.open(self.pdf_path)
        tiles = []
        for pno, page in enumerate(doc):
            scale, step, cols, rows = self._plan(page)

            if cols * rows > self.max_tiles:
                shrink = math.sqrt(self.max_tiles / (cols * rows))
                self.dpi = max(72, int(self.dpi * shrink))
                scale, step, cols, rows = self._plan(page)
                print(f"  [自适应] 第{pno+1}页块数超限，DPI降至 {self.dpi} → {cols}x{rows}")

            gray = page.get_pixmap(
                matrix=fitz.Matrix(scale, scale), colorspace=fitz.csGRAY)
            arr = np.frombuffer(gray.samples, dtype=np.uint8).reshape(
                gray.height, gray.width)
            # 记录本页渲染几何（标注阶段按相同 scale 重渲染整页）
            self.page_dims[pno + 1] = {
                "scale": scale, "w_px": gray.width, "h_px": gray.height}

            for ri in range(rows):
                for ci in range(cols):
                    x0 = int(ci * step)
                    y0 = int(ri * step)
                    sub = arr[y0:y0 + self.tile_px, x0:x0 + self.tile_px]
                    if sub.size == 0 or (sub < 200).mean() < self.ink_threshold:
                        continue
                    # 实际块像素尺寸（边缘块会被页面边界裁短）
                    w_px = int(min(self.tile_px, gray.width - x0))
                    h_px = int(min(self.tile_px, gray.height - y0))
                    clip = fitz.Rect(
                        x0 / scale, y0 / scale,
                        (x0 + self.tile_px) / scale,
                        (y0 + self.tile_px) / scale)
                    pm = page.get_pixmap(
                        matrix=fitz.Matrix(scale, scale), clip=clip)
                    # 模型实际看到的图像尺寸（fitz 对超出页面的区域填充白色，
                    # pm.width/pm.height 才是模型收到的图像宽高，是 bbox 归一化的基准）
                    tiles.append({
                        "page": pno + 1,
                        "row": ri, "col": ci,
                        "png": pm.tobytes("png"),
                        # 标注用几何：本块在整页像素坐标中的位置与尺寸
                        "x0_px": x0, "y0_px": y0,
                        "w_px": pm.width,   # 用实际渲染宽（含白边填充）
                        "h_px": pm.height,  # 用实际渲染高（含白边填充）
                        # 真实内容边界（不超过页面），用于裁剪坐标防溢出
                        "content_w_px": w_px,
                        "content_h_px": h_px,
                        "scale": scale,
                    })
        doc.close()
        return tiles


# ═════════════════════════════════════════════
# 2. Vision 分析器（不变）
# ═════════════════════════════════════════════

VISION_PROMPT = """你是汽车工程图纸识别专家。下图是一张【大幅CAD图纸的局部裁剪块】，
图纸被切成多块分别识别，所以本块里看到的视图标题，通常就拥有本块内的尺寸。

【核心任务】识别本块所有工程元素，并尽最大努力判断每个尺寸/公差属于哪个视图或剖面。

【气泡编号识别】很多尺寸旁边有圆圈数字（气泡编号/Dim No.，如 ①②③ 或方框数字 1 2 3），
这是检验报告里的尺寸序号。若尺寸旁能看到圆圈/方框数字，填入 dim_no 字段；看不到则填 ""。

【角度识别】识别所有角度标注（°），包括度数、分、秒。type="angle"，value="30°"或"30°20'"。

【表面粗糙度识别】识别表面粗糙度符号（√/▽）及其标注值，包括 Ra、Rz、Rmax 等参数：
type="surface_roughness"，value="Ra 0.3"或"Rz20 MAX"或"Ra 2μm MIN ALL AROUND"。

【基准特征标签识别】识别单独的基准字母标签（如 J→K 箭头指向、R P Q 多基准），
这些不是 GD&T 控制框而是基准面标识。type="datum_feature"，value="J→K"或"R P Q"。

【GD&T公差→Upper】形位公差的值（如 ⌭0.10 中的 0.10、⊥0.07 中的 0.07）是最大允许偏差，
填入 gdt_tolerance 字段（数字），这个值将自动填入检验报告表的 Upper 列。

【GD&T符号严格区分，非常重要】以下符号容易混淆，必须仔细辨别：
  ⌒ (U+2312) 线轮廓度 —— 弧形开口向下，像一个拱形
  ⌓ (U+2313) 面轮廓度 —— 弧形下方有一横，像拱形下面加了底线
  ⊕ (U+2295) 位置度   —— 圆圈内有十字，圆形带加号
  ⊗ (U+2297) 圆跳动   —— 圆圈内有叉号，圆形带X
  ≡ (U+2261) 对称度   —— 三条平行横线
  ∥ (U+2225) 平行度   —— 两条平行竖线
  ⊥ (U+22A5) 垂直度   —— 上方横线下方垂直线，像倒T
  ⏥ (U+23E5) 平面度   —— 矩形框
  ⏤ (U+23E4) 直线度   —— 一条横线
识别时请仔细观察符号形状，不要把面轮廓度⌓误认为位置度⊕。

【视图归属判断规则，按优先级】
1. 顺着尺寸的引线/箭头，看它指向哪个视图的几何体，归到那个视图。
2. 若引线不清，归到该尺寸簇上方或最近的视图标题（如 DETAIL A、SECTION B-B、TOP VIEW、30W INTERFACE）。
3. 若本块只出现一个视图标题，则本块所有尺寸都归给它。
4. 只有当本块内【完全看不到任何视图标题】时，belongs_to_view 才填 ""。
   不要因为"不确定"就留空——请给出最可能的那个视图。
5. 同时用 view_confidence 标注把握程度：high(引线明确) / medium(就近推断) / low(仅一个候选)。

【严格只输出 JSON】不要解释、不要 markdown 代码块：
{
  "views": ["本块出现的视图/剖面标签，按图中原文"],
  "elements": [
    {
      "type": "dimension | gdt | thread | chamfer | radius | note | datum_label | angle | surface_roughness | datum_feature",
      "value": "标称值，如 21.1 / Φ1.6 / R0.3 / 30° / Ra 0.3 / J→K",
      "tolerance": "公差，如 ±0.1 / +0.021/0，无则空",
      "multiplier": "数量倍数，如 2X / 6X / 12X，无则空",
      "dim_no": "气泡编号(圆圈/方框数字)，看不到填空",
      "is_ref": "是否参考尺寸(带括号或REF标注)，true/false",
      "datums": ["GD&T基准字母，如 A,B,C；非GD&T留空数组"],
      "gdt_tolerance": "形位公差值(仅gdt类型填数字，如0.10)，无则空",
      "surface_param": "粗糙度参数类型(仅surface_roughness填)：Ra/Rz/Rmax",
      "belongs_to_view": "所属视图标签（尽量给出，按上面规则）",
      "view_confidence": "high | medium | low",
      "bbox": "该元素在本图中的边界框，格式 [x0,y0,x1,y1]，用 0~1000 的整数表示相对本块左上角的归一化坐标(x向右,y向下)；务必尽量给出",
      "raw": "原始标注文字"
    }
  ]
}
若本块是标题栏/BOM/纯文本表格，elements 可为空数组，views 注明 ["TITLE_BLOCK"]。"""


class VisionAnalyzer:
    def __init__(self, model="qwen-vl-max", mock=False, max_retries=2):
        self.model = model
        self.mock = mock
        self.max_retries = max_retries
        self.client = None
        if not mock:
            from openai import OpenAI
            self.client = OpenAI(
                api_key=os.environ.get("DASHSCOPE_API_KEY"),
                base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
            )

    def analyze_tile(self, tile):
        if self.mock:
            return self._mock_response(tile)

        b64 = base64.standard_b64encode(tile["png"]).decode()

        for attempt in range(self.max_retries + 1):
            try:
                resp = self.client.chat.completions.create(
                    model=self.model,
                    messages=[{
                        "role": "user",
                        "content": [
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/png;base64,{b64}"
                                },
                            },
                            {"type": "text", "text": VISION_PROMPT},
                        ],
                    }],
                    max_tokens=4096,
                )
                text = resp.choices[0].message.content or ""
                return self._parse_json(text)

            except Exception as e:
                if attempt < self.max_retries:
                    time.sleep(2 ** attempt)
                else:
                    print(f"    ! 区块({tile['row']},{tile['col']}) 识别失败: {e}")
                    return {"views": [], "elements": []}

    @staticmethod
    def _parse_json(text):
        text = re.sub(r"^```(?:json)?|```$", "", text.strip(),
                      flags=re.MULTILINE).strip()
        m = re.search(r"\{[\s\S]*\}", text)
        if not m:
            return {"views": [], "elements": []}
        try:
            data = json.loads(m.group(0))
            data.setdefault("views", [])
            data.setdefault("elements", [])
            return data
        except json.JSONDecodeError:
            return {"views": [], "elements": []}

    @staticmethod
    def _mock_response(tile):
        r, c = tile["row"], tile["col"]
        if (r + c) % 3 == 0:
            view = f"SECTION {chr(66 + c % 5)}-{chr(66 + c % 5)}"
            elements = [
                {"type": "dimension", "value": f"{20 + c}.1",
                 "tolerance": "±0.1", "multiplier": "2X" if c % 2 else "",
                 "datums": [], "belongs_to_view": view,
                 "view_confidence": "high", "dim_no": str(c + 1),
                 "is_ref": (c % 4 == 0), "bbox": [120, 150, 360, 230],
                 "raw": f"2X {20 + c}.1 ±0.1"},
                {"type": "gdt", "value": "0.1", "tolerance": "",
                 "multiplier": "", "datums": ["E"],
                 "gdt_tolerance": "0.1", "bbox": [400, 300, 640, 380],
                 "belongs_to_view": view, "view_confidence": "medium",
                 "raw": "0.1 E"},
            ]
            if c % 3 == 1:
                elements.append({
                    "type": "angle", "value": "30°",
                    "tolerance": "", "multiplier": "",
                    "dim_no": str(c + 10), "bbox": [200, 500, 380, 560],
                    "belongs_to_view": view,
                    "view_confidence": "high",
                    "raw": "30°",
                })
            if c % 4 == 2:
                elements.append({
                    "type": "surface_roughness", "value": "Ra 0.3",
                    "surface_param": "Ra",
                    "tolerance": "", "multiplier": "",
                    "belongs_to_view": view, "bbox": [600, 600, 820, 670],
                    "view_confidence": "high",
                    "raw": "Ra ≤ 0.3 ALL AROUND",
                })
            if c % 5 == 3:
                elements.append({
                    "type": "datum_feature", "value": "J→K",
                    "tolerance": "", "multiplier": "",
                    "belongs_to_view": view, "bbox": [700, 100, 880, 170],
                    "view_confidence": "medium",
                    "raw": "J arrow K",
                })
            if c == 4:
                elements[0]["multiplier"] = "12X"
                elements[0]["dim_no"] = "5"
                elements[0]["raw"] = "12X 24.1 ±0.1"
            return {"views": [view], "elements": elements}
        return {"views": [], "elements": []}


# ═════════════════════════════════════════════
# 3. 合并去重 + 清洗（不变）
# ═════════════════════════════════════════════

def normalize_tolerance(s):
    if not s:
        return s
    s = s.replace("+/-", "±").replace("＋/－", "±")
    s = re.sub(r"\s*±\s*", "±", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def clean_view_name(v):
    if not v:
        return None
    v = re.sub(r"\s+", " ", v).strip().upper()
    v = re.sub(r"^SEE\s+", "", v)
    if v == "TITLE_BLOCK":
        return v

    v = re.sub(r"(SECTION|DETAIL)\s+([A-Z]{1,2})\s*-\s*([A-Z]{1,2})", r"\1 \2-\3", v)

    if v in ("SECTION", "DETAIL", "VIEW"):
        return None
    if re.search(r"[/\\]", v):
        return None
    if re.fullmatch(r"[\d.]+", v):
        return None
    if re.match(r"^\d{5,}", v):
        return None
    if re.match(r"^REF\.?\b", v):
        return None
    if re.search(r"PBT|GF\d|ABS|PA6|NYLON|RAL", v):
        return None
    if re.search(r"CODE|EXCEPT|SAME AS|±|&", v):
        return None
    if re.fullmatch(r"[A-Z]\d+", v):
        return None

    if re.fullmatch(r"SECTION [A-Z]{1,2}-[A-Z]{1,2}", v):
        return v
    if re.fullmatch(r"DETAIL [A-Z]{1,2}", v):
        return v
    if re.fullmatch(r"\d+W", v):
        return v + " INTERFACE"
    if re.search(r"\d+W INTERFACE|TOP VIEW|FRONT VIEW|BOTTOM VIEW|SIDE VIEW", v):
        return v
    if re.fullmatch(r"[A-Z]{1,2}", v):
        return None

    return None


def parse_dimension(value, tolerance="", raw=""):
    nominal = lower = upper = None
    val_str = str(value or raw or "")
    val_clean = re.sub(r"^\d+\s*[Xx×]\s*", "", val_str).strip()

    if "°" in val_clean:
        ang_m = re.search(r"(\d+(?:\.\d+)?)\s*°", val_clean)
        if ang_m:
            nominal = float(ang_m.group(1))
            tol_str = str(tolerance or "").replace("°", "")
            tol_m = re.search(r"±\s*(\d+(?:\.\d+)?)", tol_str)
            if tol_m:
                t = float(tol_m.group(1))
                lower, upper = nominal - t, nominal + t
                lower, upper = round(lower, 4), round(upper, 4)
        return {"nominal": nominal, "lower": lower, "upper": upper}

    nom_m = re.search(r"[ΦØR]?\s*(\d+(?:\.\d+)?)", val_clean)
    if not nom_m:
        return {"nominal": None, "lower": None, "upper": None}
    nominal = float(nom_m.group(1))

    tol = (tolerance or "").strip()
    if not tol and raw:
        tm = re.search(
            r"(±\s*\d*\.?\d+"
            r"|\+\s*\d*\.?\d+\s*/?\s*[-−]\s*\d*\.?\d+"   # +0.3/-0.2 或 +0.3 -0.2
            r"|0\s*/?\s*[-−]\s*\d*\.?\d+)",
            raw)
        if tm:
            tol = tm.group(1)

    m = re.search(r"±\s*(\d*\.?\d+)", tol)
    if m:
        t = float(m.group(1))
        lower, upper = nominal - t, nominal + t
    else:
        # 非对称：+up /或空格 -lo  （支持 +0.02/-0.35、+0.02 -0.35、+.02/-.35）
        m = re.search(r"\+\s*(\d*\.?\d+)\s*/?\s*[-−]\s*(\d*\.?\d+)", tol)
        if m:
            up, lo = float(m.group(1)), float(m.group(2))
            lower, upper = nominal - lo, nominal + up
        else:
            # 单边：0/-x 或 0 -x  → 上限=nominal
            m = re.search(r"0\s*/?\s*[-−]\s*(\d*\.?\d+)", tol)
            if m:
                lower, upper = nominal - float(m.group(1)), nominal
            else:
                # 单边：+x/0 或 +x 0 → 下限=nominal
                m = re.search(r"\+\s*(\d*\.?\d+)\s*/?\s*0\b", tol)
                if m:
                    lower, upper = nominal, nominal + float(m.group(1))

    if lower is not None:
        lower, upper = round(lower, 4), round(upper, 4)
    return {"nominal": nominal, "lower": lower, "upper": upper}


class VisionMerger:
    def __init__(self):
        self.elements = []
        self.views = set()

    def add(self, tile, result):
        tile_id = f"P{tile['page']}-R{tile['row']}C{tile['col']}"
        tile_views = []
        for v in result.get("views", []):
            cv = clean_view_name(v)
            if cv:
                self.views.add(cv)
                if cv != "TITLE_BLOCK":
                    tile_views.append(cv)
        sole_view = tile_views[0] if len(tile_views) == 1 else None

        # 水印/签名行特征模式：匹配 PDF 页面底部的审核水印文字
        WATERMARK_PATTERNS = (
            re.compile(r"CUS0[12]", re.IGNORECASE),
            re.compile(r"pass\.nor", re.IGNORECASE),
            re.compile(r"\b\d{6,8}\b.*\d{2}:\d{2}"),   # 日期时间戳如 "13:23:36"
            re.compile(r"\bDate\s*:", re.IGNORECASE),
            re.compile(r"\bTime\s*:", re.IGNORECASE),
            re.compile(r"[a-z]{5,8},[0-9]{8},"),             # "xpzrro,12162024,..."
        )

        for el in result.get("elements", []):
            el = dict(el)
            etype = el.get("type", "")
            raw_val = str(el.get("value", "") or "")
            raw_txt = str(el.get("raw", "") or "")

            # ── 过滤水印行（问题B）──────────────────────────────────────
            # 水印文字出现在 value 或 raw 中则跳过
            combined = raw_val + " " + raw_txt
            if any(p.search(combined) for p in WATERMARK_PATTERNS):
                continue

            el["tolerance"] = normalize_tolerance(el.get("tolerance", ""))
            el["raw"] = normalize_tolerance(el.get("raw", ""))
            view = clean_view_name(el.get("belongs_to_view", "")) or ""
            conf = (el.get("view_confidence", "") or "").lower()

            # ── deviation_override 跨块兜底（问题A）────────────────────
            # 若 AI 漏打标签，但 belongs_to_view 明确是 DEVIATION_TABLE，强制修正类型
            if view == "DEVIATION_TABLE" and etype == "dimension":
                etype = "deviation_override"
                el["type"] = "deviation_override"

            if not view and sole_view:
                view = sole_view
                conf = conf or "low"
            el["belongs_to_view"] = view
            el["view_confidence"] = conf if conf in ("high", "medium", "low") else ""
            el["_tile"] = tile_id
            el["_page"] = tile.get("page", 1)
            # ── bbox：归一化(0~1000，相对本块) → 整页像素坐标 ──
            bb = el.get("bbox")
            el["_bbox_px"] = None
            if isinstance(bb, (list, tuple)) and len(bb) == 4:
                try:
                    x0n, y0n, x1n, y1n = [float(v) for v in bb]
                    # tw/th：模型实际看到的图像尺寸（含白边）—— bbox 归一化基准
                    tw = tile.get("w_px", 0) or 0
                    th = tile.get("h_px", 0) or 0
                    tx = tile.get("x0_px", 0)
                    ty = tile.get("y0_px", 0)
                    # 页面实际内容边界（不超过页面），用于防止坐标溢出
                    cw = tile.get("content_w_px", tw) or tw
                    ch = tile.get("content_h_px", th) or th
                    if tw and th:
                        X0 = tx + x0n / 1000.0 * tw
                        Y0 = ty + y0n / 1000.0 * th
                        X1 = tx + x1n / 1000.0 * tw
                        Y1 = ty + y1n / 1000.0 * th
                        # 限制在本块实际内容范围内，避免落在页面外白边区域
                        X0 = max(float(tx), min(X0, tx + cw))
                        X1 = max(float(tx), min(X1, tx + cw))
                        Y0 = max(float(ty), min(Y0, ty + ch))
                        Y1 = max(float(ty), min(Y1, ty + ch))
                        el["_bbox_px"] = (round(min(X0, X1), 1), round(min(Y0, Y1), 1),
                                          round(max(X0, X1), 1), round(max(Y0, Y1), 1))
                except (ValueError, TypeError):
                    el["_bbox_px"] = None
            el["dim_no"] = str(el.get("dim_no", "") or "").strip()
            el["is_ref"] = bool(el.get("is_ref", False)) or \
                ("REF" in el["raw"].upper() or el["raw"].strip().startswith("("))
            el["gdt_tolerance"] = el.get("gdt_tolerance", "")
            el["surface_param"] = el.get("surface_param", "")
            if etype == "dimension":
                calc = parse_dimension(el.get("value", ""),
                                       el.get("tolerance", ""), el.get("raw", ""))
                el["lower"] = calc["lower"]
                el["upper"] = calc["upper"]
            elif etype == "angle":
                el["lower"] = None
                el["upper"] = None
            elif etype == "deviation_override":
                el["belongs_to_view"] = "DEVIATION_TABLE"
                # Fix7: 解析偏差后规格的非对称公差，而非一律置 None
                calc = parse_dimension(el.get("value", ""),
                                       el.get("tolerance", ""), el.get("raw", ""))
                el["lower"] = calc["lower"]
                el["upper"] = calc["upper"]
            self.elements.append(el)

    def dedup(self):
        seen = {}
        for el in self.elements:
            key = (el.get("type", ""),
                   el.get("raw", "") or el.get("value", ""),
                   el.get("belongs_to_view", ""))
            if key not in seen:
                seen[key] = el
            else:
                # 优先保留有 bbox 坐标的版本（同一元素可能被多个重叠块识别到）
                existing = seen[key]
                has_new = bool(el.get("_bbox_px"))
                has_old = bool(existing.get("_bbox_px"))
                if has_new and not has_old:
                    seen[key] = el
                elif has_new and has_old:
                    # 两者都有坐标时，保留有 dim_no 的版本
                    if el.get("dim_no") and not existing.get("dim_no"):
                        seen[key] = el
        self.elements = list(seen.values())
        return self

    def by_type(self, t):
        return [e for e in self.elements if e.get("type") == t]


# ═════════════════════════════════════════════
# 3b. 元素框选标注（在整页图上把每个识别元素框出来）
# ═════════════════════════════════════════════

# 类型 → 颜色（RGB）
ELEMENT_COLORS = {
    "dimension": (220, 38, 38),         # 红
    "gdt": (37, 99, 235),               # 蓝
    "angle": (22, 163, 74),             # 绿
    "surface_roughness": (234, 88, 12), # 橙
    "datum_feature": (147, 51, 234),    # 紫
    "datum_label": (147, 51, 234),
    "thread": (202, 138, 4),            # 暗黄
    "chamfer": (8, 145, 178),           # 青
    "radius": (190, 24, 93),            # 玫红
    "note": (100, 116, 139),            # 灰
    "deviation_override": (192, 38, 211),  # 品红
}
DEFAULT_COLOR = (15, 23, 42)


class Annotator:
    """把识别到的每个元素在整页图纸上框出来，按类型上色并标注编号/类型。
    输出每页一张 PNG。需要元素带 _page 与 _bbox_px（整页像素坐标）。
    """

    def __init__(self, pdf_path, page_dims):
        self.pdf_path = pdf_path
        self.page_dims = page_dims or {}

    def annotate(self, elements, output_dir, stem):
        from PIL import Image, ImageDraw, ImageFont
        try:
            font = ImageFont.truetype("DejaVuSans.ttf", 14)
            font_sm = ImageFont.truetype("DejaVuSans.ttf", 11)
        except Exception:
            font = ImageFont.load_default()
            font_sm = font

        # 按页分组（仅保留有 bbox 的元素）
        by_page = {}
        for el in elements:
            if not el.get("_bbox_px"):
                continue
            by_page.setdefault(el.get("_page", 1), []).append(el)

        doc = fitz.open(self.pdf_path)
        outputs = []
        # 确保每页都输出，即使该页元素全无 bbox（便于用户确认问题）
        for pno, page in enumerate(doc, start=1):
            geo = self.page_dims.get(pno)
            if geo is None:
                # 该页在渲染阶段被跳过（空白页），仍然生成基本标注图
                scale = 1.0
            else:
                scale = geo["scale"]
            pm = page.get_pixmap(matrix=fitz.Matrix(scale, scale))
            img = Image.frombytes("RGB", (pm.width, pm.height), pm.samples)
            # 半透明叠加层，避免边框遮挡图纸
            overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
            draw = ImageDraw.Draw(overlay, "RGBA")

            # 整页像素边界
            page_w, page_h = img.size
            els = by_page.get(pno, [])
            counts = {}
            for el in els:
                x0, y0, x1, y1 = el["_bbox_px"]
                # 裁剪到页面范围内
                x0 = max(0.0, min(x0, page_w - 1))
                y0 = max(0.0, min(y0, page_h - 1))
                x1 = max(0.0, min(x1, page_w - 1))
                y1 = max(0.0, min(y1, page_h - 1))
                # 保证最小框尺寸（避免退化为点/线）
                MIN_SIZE = 8
                if x1 - x0 < MIN_SIZE:
                    cx = (x0 + x1) / 2
                    x0, x1 = cx - MIN_SIZE / 2, cx + MIN_SIZE / 2
                if y1 - y0 < MIN_SIZE:
                    cy = (y0 + y1) / 2
                    y0, y1 = cy - MIN_SIZE / 2, cy + MIN_SIZE / 2
                etype = el.get("type", "")
                color = ELEMENT_COLORS.get(etype, DEFAULT_COLOR)
                counts[etype] = counts.get(etype, 0) + 1
                # 框（线宽随图纸分辨率自适应，最小2px）
                lw = max(2, int(page_w / 3000))
                draw.rectangle([x0, y0, x1, y1], outline=color + (255,), width=lw)
                # 标签：气泡编号优先，否则类型缩写
                label = str(el.get("dim_no") or "").strip()
                if not label:
                    label = {"dimension": "D", "gdt": "G", "angle": "∠",
                             "surface_roughness": "Ra", "datum_feature": "DF",
                             "deviation_override": "Δ", "note": "N"}.get(etype, "·")
                # 标签底色块
                tb = draw.textbbox((0, 0), label, font=font_sm)
                tw, thh = tb[2] - tb[0], tb[3] - tb[1]
                lx, ly = x0, max(0, y0 - thh - 4)
                draw.rectangle([lx, ly, lx + tw + 6, ly + thh + 4],
                               fill=color + (235,))
                draw.text((lx + 3, ly + 2), label, fill=(255, 255, 255, 255),
                          font=font_sm)

            # 图例
            self._draw_legend(draw, counts, font, img.size)

            out_img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
            suffix = f"_标注_P{pno}.png" if doc.page_count > 1 else "_标注.png"
            out_path = str(Path(output_dir) / f"{stem}{suffix}")
            out_img.save(out_path)
            outputs.append(out_path)
        doc.close()
        return outputs

    @staticmethod
    def _draw_legend(draw, counts, font, size):
        if not counts:
            return
        items = [(t, n) for t, n in counts.items()]
        pad = 10
        line_h = 22
        box_w = 240
        box_h = pad * 2 + line_h * (len(items) + 1)
        x0 = 10
        y0 = 10
        draw.rectangle([x0, y0, x0 + box_w, y0 + box_h],
                       fill=(255, 255, 255, 230), outline=(0, 0, 0, 255))
        draw.text((x0 + pad, y0 + pad), "Detected elements / count",
                  fill=(0, 0, 0, 255), font=font)
        yy = y0 + pad + line_h
        for t, n in items:
            color = ELEMENT_COLORS.get(t, DEFAULT_COLOR)
            draw.rectangle([x0 + pad, yy + 3, x0 + pad + 14, yy + 17],
                           fill=color + (255,), outline=(0, 0, 0, 255))
            draw.text((x0 + pad + 22, yy + 2), f"{t}: {n}", fill=(0, 0, 0, 255),
                      font=font)
            yy += line_h


# ═════════════════════════════════════════════
# 4. 增强 Excel 导出（不变）
# ═════════════════════════════════════════════

class VisionExcelExporter(ExcelExporter):

    def __init__(self, extractor, merger, output_path, model_name):
        super().__init__(extractor, output_path)
        self.merger = merger
        self.model_name = model_name

    def _parse_multiplier(self, s):
        if not s:
            return 1
        m = re.search(r"(\d+)\s*[Xx×]", str(s))
        return int(m.group(1)) if m else 1

    def export(self):
        super().export()
        import openpyxl
        from openpyxl.styles import Font
        self.wb = openpyxl.load_workbook(self.output_path)
        m = self.merger

        dims = m.by_type("dimension")
        gdts = m.by_type("gdt")
        angles = m.by_type("angle")
        roughness = m.by_type("surface_roughness")
        datum_features = m.by_type("datum_feature")

        existing_nums = []
        for d in dims:
            dn = d.get("dim_no", "")
            try:
                existing_nums.append(int(dn))
            except (ValueError, TypeError):
                pass
        next_num = max(existing_nums) + 1 if existing_nums else 1
        for d in dims:
            if not d.get("dim_no", "").strip() and not d.get("is_ref"):
                raw = d.get("raw", "")
                d["dim_no"] = str(next_num)
                d["_auto_numbered"] = True
                next_num += 1

        ws = self.wb.create_sheet("🎯 尺寸-视图关联(AI)")
        self._table(
            ws,
            ["#", "Dim No.", "Ref", "所属视图", "归属置信", "数量倍数",
             "标称值", "公差", "下限", "上限", "原始标注", "来源块"],
            [(i+1, d.get("dim_no",""), "✓" if d.get("is_ref") else "",
              d.get("belongs_to_view",""), d.get("view_confidence",""),
              d.get("multiplier",""), d.get("value",""), d.get("tolerance",""),
              d.get("lower",""), d.get("upper",""),
              d.get("raw",""), d.get("_tile",""))
             for i, d in enumerate(dims)],
            [6, 9, 6, 20, 9, 9, 12, 14, 10, 10, 26, 12],
            empty_msg="（Vision未返回尺寸数据）")

        ws = self.wb.create_sheet("⊕ GD&T关联(AI)")
        self._table(
            ws,
            ["#", "所属视图", "公差值", "基准", "GD&T公差(Max)", "原始", "来源块"],
            [(i+1, g.get("belongs_to_view",""), g.get("value",""),
              ", ".join(g.get("datums",[]) or []),
              g.get("gdt_tolerance",""),
              g.get("raw",""), g.get("_tile",""))
             for i, g in enumerate(gdts)],
            [6, 22, 12, 14, 14, 24, 14],
            empty_msg="（Vision未返回GD&T）")

        ws = self.wb.create_sheet("📋 检验报告表")
        report_rows = []

        for d in dims:
            dim_no = d.get("dim_no", "")
            multiplier = self._parse_multiplier(d.get("multiplier", ""))
            lower, upper = d.get("lower"), d.get("upper")
            expand = max(1, multiplier)

            for inst in range(expand):
                # 第一行保留 dim_no，后续展开行 dim_no 置为 None（v3格式）
                instance_no = dim_no if inst == 0 else None

                if instance_no is not None or (lower is not None):
                    report_rows.append((
                        instance_no,
                        "Ref" if d.get("is_ref") else "",
                        d.get("value", ""),
                        lower if lower is not None else "",
                        upper if upper is not None else "",
                        d.get("belongs_to_view", ""),
                        d.get("raw", ""),
                    ))

        # GD&T 行：写入 "SEE GD&T"，Lower/Upper 均填 "OK"（v3格式）
        for g in gdts:
            gt = g.get("gdt_tolerance", "")
            if gt:
                report_rows.append((
                    g.get("dim_no", ""),
                    "",
                    "SEE GD&T",
                    "OK",
                    "OK",
                    g.get("belongs_to_view", ""),
                    g.get("raw", ""),
                ))

        for s in roughness:
            val = s.get("value", "")
            raw = s.get("raw", "")
            upper_val = ""
            max_m = re.search(r"(\d+(?:\.\d+)?)\s*MAX", str(val) + str(raw), re.IGNORECASE)
            if max_m:
                upper_val = max_m.group(1)
            report_rows.append((
                s.get("dim_no", ""),
                "",
                f"粗糙度: {val}",
                "",
                upper_val,
                s.get("belongs_to_view", ""),
                raw,
            ))

        def sort_key(r):
            try:
                base = re.split(r"[-_]", str(r[0]))[0]
                return (0, int(base))
            except (ValueError, TypeError):
                return (1, 0)
        report_rows.sort(key=sort_key)
        self._table(
            ws,
            ["Dim No.", "Ref Only", "Drawing Dimension", "Lower", "Upper", "所属视图", "原始标注"],
            report_rows,
            [10, 10, 18, 12, 12, 20, 26],
            empty_msg="（无可生成检验项的尺寸）")

        ws = self.wb.create_sheet("📐 角度标注(AI)")
        self._table(
            ws,
            ["#", "Dim No.", "所属视图", "角度值", "原始标注", "来源块"],
            [(i+1, a.get("dim_no",""), a.get("belongs_to_view",""),
              a.get("value",""), a.get("raw",""), a.get("_tile",""))
             for i, a in enumerate(angles)],
            [6, 10, 20, 14, 26, 12],
            empty_msg="（未识别到角度标注）")

        ws = self.wb.create_sheet("🔲 表面粗糙度(AI)")
        self._table(
            ws,
            ["#", "Dim No.", "所属视图", "参数类型", "粗糙度值", "原始标注", "来源块"],
            [(i+1, s.get("dim_no",""), s.get("belongs_to_view",""),
              s.get("surface_param",""), s.get("value",""),
              s.get("raw",""), s.get("_tile",""))
             for i, s in enumerate(roughness)],
            [6, 10, 20, 12, 20, 26, 12],
            empty_msg="（未识别到表面粗糙度）")

        ws = self.wb.create_sheet("🔤 基准特征(AI)")
        self._table(
            ws,
            ["#", "所属视图", "基准标签", "原始标注", "来源块"],
            [(i+1, df.get("belongs_to_view",""), df.get("value",""),
              df.get("raw",""), df.get("_tile",""))
             for i, df in enumerate(datum_features)],
            [6, 22, 20, 26, 12],
            empty_msg="（未识别到基准特征标签）")

        ws = self.wb.create_sheet("🗺 视图清单(AI)")
        self._table(
            ws, ["#", "视图/剖面标签"],
            [(i+1, v) for i, v in enumerate(sorted(m.views))],
            [6, 40], empty_msg="（未识别到视图）")

        ws = self.wb.create_sheet("🔍 全部视觉元素(AI)")
        self._table(
            ws,
            ["#", "类型", "标称值", "公差", "倍数", "基准", "所属视图", "归属置信", "原始", "来源块"],
            [(i+1, e.get("type",""), e.get("value",""), e.get("tolerance",""),
              e.get("multiplier",""), ", ".join(e.get("datums",[]) or []),
              e.get("belongs_to_view",""), e.get("view_confidence",""),
              e.get("raw",""), e.get("_tile",""))
             for i, e in enumerate(m.elements)],
            [6, 12, 14, 14, 8, 12, 22, 10, 26, 12])

        ov = self.wb["📋 图纸信息"]
        last = ov.max_row + 2
        ov.cell(row=last, column=1,
                value="【第二阶段 Vision 识别（百炼）】").font = Font(
            bold=True, color=self.C_SUB, name="Arial", size=11)
        with_view = sum(1 for d in dims if d.get("belongs_to_view"))
        rate = f"{100*with_view//len(dims)}%" if dims else "—"
        auto_num = sum(1 for d in dims if d.get("_auto_numbered"))
        expanded = sum(1 for d in dims if self._parse_multiplier(d.get("multiplier","")) > 1)
        stats = [
            ("使用模型", self.model_name),
            ("识别视图数", len(m.views)),
            ("视觉元素总数", len(m.elements)),
            ("其中尺寸", len(dims)),
            ("其中GD&T", len(gdts)),
            ("其中角度", len(angles)),
            ("其中表面粗糙度", len(roughness)),
            ("其中基准特征", len(datum_features)),
            ("NX倍数展开项", expanded),
            ("自动编号项", auto_num),
            ("尺寸视图归属率", f"{with_view}/{len(dims)} ({rate})"),
        ]
        for i, (k, v) in enumerate(stats, 1):
            ov.cell(row=last+i, column=1, value=k).font = Font(
                bold=True, name="Arial", size=9)
            ov.cell(row=last+i, column=2, value=str(v)).font = Font(
                name="Arial", size=9)

        self.wb.save(self.output_path)
        return self.output_path


# ═════════════════════════════════════════════
# 5. 主流程（高速并发版）
# ═════════════════════════════════════════════

def process_with_vision(pdf_path, output_dir=".", model="qwen-vl-max",
                        dpi=100, max_tiles=36, mock=False, annotate=True):
    pdf_path = str(pdf_path)
    stem = Path(pdf_path).stem
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    out = str(Path(output_dir) / f"{stem}_识别结果_V2.xlsx")

    print("[1/5] 第一阶段文本提取...")
    parser = PDFParser(pdf_path).parse()
    extractor = CADInfoExtractor(parser).extract_all()

    print(f"[2/5] 分块高清渲染 (DPI={dpi}, 上限{max_tiles}块)...")
    renderer = TileRenderer(pdf_path, dpi=dpi, max_tiles=max_tiles)
    tiles = renderer.render()
    if len(tiles) > max_tiles:
        print(f"      有效块 {len(tiles)} 超上限，截断至 {max_tiles}")
        tiles = tiles[:max_tiles]
    print(f"      → 有效区块 {len(tiles)} 个" + ("  [MOCK模式]" if mock else ""))

    print(f"[3/5] Vision 并发识别 (model={model}, workers=6)...")
    analyzer = VisionAnalyzer(model=model, mock=mock)
    merger = VisionMerger()

    # ═══ 核心改动：6 线程并发识别 ═══
    completed = 0
    total = len(tiles)

    def _worker(tile):
        """单个线程任务：识别一个区块"""
        result = analyzer.analyze_tile(tile)
        return tile, result

    with ThreadPoolExecutor(max_workers=6) as executor:
        # 提交所有任务
        future_to_tile = {
            executor.submit(_worker, tile): tile for tile in tiles
        }

        # 按完成顺序处理结果（不阻塞）
        for future in as_completed(future_to_tile):
            tile, result = future.result()
            merger.add(tile, result)
            completed += 1
            if completed % 5 == 0 or completed == total:
                print(f"      进度 {completed}/{total}，累计元素 {len(merger.elements)}")

    merger.dedup()
    print(f"      → 去重后 {len(merger.elements)} 个元素，{len(merger.views)} 个视图")

    print("[4/5] 融合文本+视觉结果...")
    print("[5/5] 生成增强 Excel...")
    VisionExcelExporter(extractor, merger, out, model).export()
    print(f"完成 → {out}")

    annotated = []
    if annotate:
        print("[+] 生成元素框选标注图...")
        try:
            annotated = Annotator(pdf_path, renderer.page_dims).annotate(
                merger.elements, output_dir, stem)
            n_box = sum(1 for e in merger.elements if e.get("_bbox_px"))
            print(f"      → 标注 {n_box}/{len(merger.elements)} 个带坐标的元素，"
                  f"输出 {len(annotated)} 张图")
            for p in annotated:
                print(f"        {p}")
            if n_box == 0:
                print("      ⚠ 模型未返回任何 bbox 坐标，标注图为空框。"
                      "请确认模型支持视觉定位(grounding)。")
        except Exception as e:
            print(f"      ! 标注生成失败: {e}")

    return {
        "output": out,
        "tiles": len(tiles),
        "elements": len(merger.elements),
        "views": sorted(merger.views),
        "annotated": annotated,
    }


def main():
    ap = argparse.ArgumentParser(description="CAD图纸 Vision 识别（百炼版·高速并发）")
    ap.add_argument("pdf", help="输入PDF路径")
    ap.add_argument("output_dir", nargs="?", default=".", help="输出目录")
    ap.add_argument("--model", default="qwen-vl-max",
                    help="模型 (默认 qwen-vl-max)")
    ap.add_argument("--dpi", type=int, default=100, help="渲染DPI (默认100)")
    ap.add_argument("--max-tiles", type=int, default=36,
                    help="区块数上限 (默认36)")
    ap.add_argument("--workers", type=int, default=6,
                    help="并发线程数 (默认6)")
    ap.add_argument("--mock", action="store_true",
                    help="模拟模式，无需密钥，验证流程")
    ap.add_argument("--no-annotate", action="store_true",
                    help="不生成元素框选标注图")
    args = ap.parse_args()

    if not args.mock and not os.environ.get("DASHSCOPE_API_KEY"):
        print("⚠  未检测到 DASHSCOPE_API_KEY 环境变量。")
        print("   设置方式：")
        print("   Windows:   set DASHSCOPE_API_KEY=sk-你的密钥")
        print("   Mac/Linux: export DASHSCOPE_API_KEY=sk-你的密钥")
        print("   或加 --mock 先验证流程。")
        sys.exit(1)

    result = process_with_vision(
        args.pdf, args.output_dir,
        model=args.model, dpi=args.dpi,
        max_tiles=args.max_tiles, mock=args.mock,
        annotate=not args.no_annotate,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()