"""
Smart CAD Diff Engine — 结构化实体匹配 & 噪声过滤
==================================================
解决原 DrawingComparator 的文本级逐行比对的误报问题：
  - 按工程内容指纹匹配（非检验序号）
  - OCR 大小写/空格/视图名称 差异过滤
  - 检验编号/气球编号重排检测
  - 偏差清单与主表交叉去重

用法（命令行）:
  python cad_smart_diff.py old.xlsx new.xlsx --csv report.csv
  python cad_smart_diff.py old.xlsx new.xlsx --json report.json

用法（Python API）:
  engine = SmartDiffEngine("old.xlsx", "new.xlsx")
  result = engine.compare_all()
  csv_str = engine.to_csv(result)
"""

import re
import csv
import json
import io
import argparse
from pathlib import Path
from collections import defaultdict

import openpyxl

# ══════════════════════════════════════════════════════════════════════
# 0. 常量
# ══════════════════════════════════════════════════════════════════════

MATCH_TOLERANCE = 0.005       # 标称值浮点匹配容差 (mm)
TOL_CMP_DECIMALS = 3          # 公差比较精度（小数位）
RENUMBER_THRESHOLD = 0.30     # 内容指纹匹配但编号不同的占比阈值 → 判定整体重编号
VIEW_NAME_LETTER_DIFF_PATTERN = re.compile(
    r"^(DETAIL|SECTION)\s+[A-Z]{1,2}(?:-[A-Z]{1,2})?$"
)

# ── 偏差表（DEVIATION TABLE）区域识别关键词 ──────────────
# 命中其中任意一个（出现在 view 或 raw 文本中）即判定该条目来源于偏差表区域，
# 这类条目不应被当成正式尺寸/GD&T 进行增删比对。
DEVIATION_REGION_KEYWORDS = (
    "DEVIATION TABLE",
    "DEVIATION ACCEPTABLE",
    "DEVIATION LIST",
    "DIM NO.",
    "DIM NO",
    "FOR TOOL NO.",
    "FOR TOOL NO",
    "DEVIATION",
)

# ── 标题栏标准公差范围表识别 ──────────────────────────
# 命中这些模式且 view==TITLE_BLOCK 的条目，视为"标准公差表"，做归一化豁免比对。
GENERAL_TOL_VALUE_PATTERNS = (
    re.compile(r"^\s*>\s*\d"),                       # >0  >20  >120
    re.compile(r"ANGULAR\s+TOLERANCE", re.IGNORECASE),
    re.compile(r"GENERAL\s+TOLERANCE", re.IGNORECASE),
    re.compile(r"^\s*\d+(\.\d+)?\s*/\s*\d+(\.\d+)?\s*$"),  # 19.8/20.2 形式的上下限
    re.compile(r"^\s*±\s*\d"),                       # ±0.15
)

# ── 修订历史 "WAS xxx" 提取 ───────────────────────────
# 匹配 "REVISED DIM #33 WAS Rz15-20" / "DIM #5 WAS 10.2" / "WAS 0.30"
REVISION_WAS_PATTERN = re.compile(
    r"(?:DIM\s*#?\s*(\d+)\s+)?WAS\s+([A-Za-z0-9.\-/±+]+)",
    re.IGNORECASE,
)

SHEET_CONFIG = {
    "检验报告表": {
        "cols": ["dim_no", "ref_only", "drawing", "lower", "upper", "view", "raw"],
        "type": "dimension",
    },
    "📋 检验报告表": {
        "cols": ["dim_no", "ref_only", "drawing", "lower", "upper", "view", "raw"],
        "type": "dimension",
    },
    "⊕ GD&T关联(AI)": {
        "cols": ["seq", "view", "tolerance_val", "datums", "gdt_max", "raw", "tile"],
        "type": "gdt",
    },
    "📋 偏差清单": {
        "cols": ["dim_ref", "original_spec", "deviated_spec"],
        "type": "deviation",
    },
    "📝 技术要求": {
        "cols": ["number", "level", "content"],
        "type": "note",
    },
    "📦 物料清单BOM": {
        "cols": ["seq", "part_number", "color", "material_spec", "assembly"],
        "type": "bom",
    },
    "⭐ 关键特性": {
        "cols": ["seq", "type", "spec", "description"],
        "type": "key_char",
    },
    "📐 尺寸标注": {
        "cols": ["seq", "multiplier", "value", "controlled"],
        "type": "raw_dim",
    },
    "🔲 表面粗糙度(AI)": {
        "cols": ["seq", "dim_no", "view", "param_type", "value", "raw", "tile"],
        "type": "surface_roughness",
    },
    "📐 角度标注(AI)": {
        "cols": ["seq", "dim_no", "view", "value", "raw", "tile"],
        "type": "angle",
    },
    "📄 原始文本": {
        "cols": ["text"],
        "type": "raw_text",
    },
}

# 图纸信息 key → 展示名映射
INFO_KEY_MAP = {
    "图号 (DWG NO.)": "Part Number",
    "图名 (DWG NAME)": "Part Name",
    "最新版本 (REV)": "Revision",
    "比例 (SCALE)": "Scale",
    "图幅 (SIZE)": "Sheet Size",
    "张数 (SHEET)": "Sheet",
    "标注标准": "Dimensioning Standard",
    "投影方式": "Projection",
    "单位": "Unit",
    "产地": "Origin",
    "涉及公司": "Companies",
    "PDF标题": "PDF Title",
    "创建工具": "Creator Tool",
    "创建日期": "Creation Date",
    "页数": "Page Count",
    "页面尺寸(pt)": "Page Size (pt)",
}


# ══════════════════════════════════════════════════════════════════════
# 1. Entity Normalizer — 文本/数值/公差标准化
# ══════════════════════════════════════════════════════════════════════

def normalize_text(s):
    """统一大小写、压缩空白、去首尾空格"""
    if s is None:
        return ""
    s = str(s).upper().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def normalize_tolerance_text(s):
    """统一公差文本格式：+/- → ±，处理各种空白"""
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r"\s*\+/\-\s*", "±", s)
    s = re.sub(r"\s*±\s*", "±", s)
    s = re.sub(r"\s+", " ", s)
    return s


def parse_nominal(text):
    """从标注文本提取标称数值。
    例: 'D4.5' → 4.5, 'R0.3' → 0.3, '21.1' → 21.1, 'M6x1.0' → 6.0
    返回 float 或 None
    """
    if not text:
        return None
    t = str(text).strip().upper()
    # 去掉直径/半径前缀
    t = re.sub(r"^[DØΦ⌀]", "", t)
    t = re.sub(r"^R\s*", "", t)
    t = re.sub(r"^M", "", t)  # 螺纹
    # 去掉倍数前缀
    t = re.sub(r"^\d+\s*[X×]\s*", "", t)
    # 去掉 ± 及以后的公差
    t = re.split(r"[±]", t)[0].strip()
    # 去掉 X 后面的螺距等（如 M6x1.0 → 取 6）
    t = re.split(r"[X×]", t)[0].strip()
    # 提取第一个数字
    m = re.search(r"(\d+\.?\d*)", t)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


def normalize_drawing_value(text):
    """将 Drawing Dimension 规范化为统一格式用于匹配。
    例: '2X Ø4.5±0.1' → 'D4.5'
         '6X R 0.3' → 'R0.3'
         '21.1' → '21.1'
         'Ra 0.3' → 'RA0.3'
         'Rz20 MAX' → 'RZ20'
    """
    if not text:
        return ""
    s = str(text).upper().strip()
    # 去掉倍数前缀
    s = re.sub(r"^\d+\s*[X×]\s*", "", s)
    # 去掉 ± 及之后的公差
    s = re.split(r"[±]", s)[0].strip()
    # 去掉 MAX / MIN 后缀
    s = re.sub(r"\s*(MAX|MIN)\s*$", "", s)
    # 检查类型
    is_dia = bool(re.search(r"[ØΦ⌀]", s) or re.match(r"^D\s*[\d.]", s))
    is_radius = bool(re.match(r"^R\s*[\d.]", s))
    is_rough = bool(re.match(r"^(RA|RZ|RMAX)\b", s))
    is_thread = bool(re.match(r"^M[\d.]", s))

    # 提取数字
    num_m = re.search(r"(\d+\.?\d*)", s)
    num = num_m.group(1) if num_m else ""

    if is_dia:
        return "D" + num
    if is_radius:
        return "R" + num
    if is_rough:
        pre = re.match(r"^(RA|RZ|RMAX)", s)
        return (pre.group(1) + num) if (pre and num) else s
    if is_thread:
        return "M" + num
    if num:
        return num
    return re.sub(r"\s+", "", s)


def parse_tolerance_bounds(lower_raw, upper_raw):
    """将 Lower/Upper 列的值解析为浮点数对。
    处理 'OK'、空字符串、None 等特殊值。
    返回 (lower_float_or_None, upper_float_or_None)
    """
    def _to_float(v):
        if v is None:
            return None
        s = str(v).strip().upper()
        if s in ("", "OK", "N/A", "NONE", "-"):
            return None
        # 去除可能的 ± 前缀（lower 有时带符号）
        s = s.replace("±", "").strip()
        try:
            return round(float(s), TOL_CMP_DECIMALS)
        except (ValueError, TypeError):
            return None

    return _to_float(lower_raw), _to_float(upper_raw)


def fuzzy_val_match(v1, v2, tol=MATCH_TOLERANCE):
    """两个数值在容差内匹配"""
    if v1 is None or v2 is None:
        return v1 is None and v2 is None
    return abs(v1 - v2) < tol


def is_view_label_ocr_noise(view_a, view_b):
    """判断两个视图名差异是否仅为 OCR 噪声。
    覆盖三类情况：
      1. 归一化后完全相同
      2. 一方是另一方的前缀（如 DETAIL A vs DETAIL AN —— OCR 漏读尾字母）
      3. 同类型标签末尾字母相邻（如 DETAIL AH vs DETAIL AG）
    """
    if not view_a or not view_b:
        # 一方为空：视图归属识别遗漏，不计为真实变更
        return True
    a = normalize_text(view_a)
    b = normalize_text(view_b)
    if a == b:
        return True

    # 情况2：前缀截断 —— DETAIL A 是 DETAIL AN 的前缀（或反之）
    # 要求必须是同一类型（DETAIL/SECTION），且短的是长的前缀
    for short, long in ((a, b), (b, a)):
        if long.startswith(short) and len(long) > len(short):
            suffix = long[len(short):]
            # 后缀只能是字母（如 N、BC），不能是数字或其他（避免误判 DETAIL A vs DETAIL A2）
            if re.fullmatch(r"[A-Z]+", suffix):
                return True

    # 情况3：同类型标签末尾字母 ASCII 相差 ≤ 1（H/G、N/M、I/J 等 OCR 近似）
    ma = re.fullmatch(r"(DETAIL|SECTION)\s+([A-Z]{1,3}(?:-[A-Z]{1,3})?)", a)
    mb = re.fullmatch(r"(DETAIL|SECTION)\s+([A-Z]{1,3}(?:-[A-Z]{1,3})?)", b)
    if ma and mb and ma.group(1) == mb.group(1):
        sa, sb = ma.group(2), mb.group(2)
        if len(sa) == len(sb):
            diffs = [(ca, cb) for ca, cb in zip(sa, sb) if ca != cb]
            if len(diffs) == 1 and abs(ord(diffs[0][0]) - ord(diffs[0][1])) <= 1:
                return True

    return False


def is_ocr_only_difference(text_a, text_b):
    """判断两个文本字符串是否仅在 OCR 噪声层面有差异：
    - 大小写不同
    - 尾部空格不同
    - 'MAX' vs 'Max' 之类
    """
    a = normalize_text(text_a)
    b = normalize_text(text_b)
    # 完全相同 → 毫无差异
    if a == b:
        return True
    # 去掉 MAX/MIN 后缀再比
    a_stripped = re.sub(r"\s*(MAX|MIN)\b", "", a).strip()
    b_stripped = re.sub(r"\s*(MAX|MIN)\b", "", b).strip()
    if a_stripped == b_stripped:
        return True
    # 同数字但格式不同（如 "0.5" vs ".5"）
    return False


# ── 问题1：标题栏标准公差表 归一化 & 识别 ──────────────────

def normalize_general_tol(text):
    """对标题栏标准公差表条目做强归一化，消除 OCR 格式抖动后再比对。
    例: '>  20   19.8 / 20.2' → '>20 19.8/20.2'
        'ANGULAR TOLERANCE  ±2 °' → 'ANGULARTOLERANCE±2'
        '+/- 0.15' → '±0.15'
    """
    if text is None:
        return ""
    s = str(text).upper()
    # 统一正负号写法
    s = re.sub(r"\+\s*/\s*-", "±", s)
    s = re.sub(r"\s*±\s*", "±", s)
    # 去掉度数符号 / 多余单位字符
    s = s.replace("°", "").replace("º", "")
    # 去掉所有空白
    s = re.sub(r"\s+", "", s)
    # 统一斜杠两侧
    s = re.sub(r"\s*/\s*", "/", s)
    return s.strip()


def is_general_tol_entry(view, value):
    """判断某条目是否为标题栏标准公差范围表条目（问题1豁免对象）。"""
    v = normalize_text(view)
    if v != "TITLE_BLOCK":
        return False
    val = str(value or "")
    return any(p.search(val) for p in GENERAL_TOL_VALUE_PATTERNS)


# ── 问题2：偏差表区域检测 ──────────────────────────────────

def is_deviation_region_text(*texts):
    """若任一文本中含偏差表关键词，则判定来源于偏差表区域。"""
    blob = " ".join(normalize_text(t) for t in texts if t)
    return any(kw in blob for kw in DEVIATION_REGION_KEYWORDS)


# ── 问题4：检验编号(DIM NO.)归一化 ─────────────────────────

def normalize_dim_no(s):
    """将各种检验编号写法归一为纯数字串，用于跨表关联。
    例: '39#' → '39', '#39' → '39', ' 7 ' → '7', 'D12' → '12'
    """
    if s is None:
        return ""
    t = str(s).strip().upper()
    m = re.search(r"(\d+)", t)
    return m.group(1) if m else t


# ── Fix4：气泡序号伪尺寸识别 ───────────────────────────────

def _is_bubble_sequence_artifact(dim) -> bool:
    """判断一个尺寸实体是否实际是被误读的"气泡序号"。
    特征：drawing 值是纯小整数(1~30)，等于自身 dim_no，且无任何公差。
    例: dim_no=94, drawing='94', 无公差 → 气泡序号，不是真实尺寸。
    """
    if dim.lower is not None or dim.upper is not None:
        return False
    raw = str(dim.drawing_raw or "").strip()
    # 去掉可能的倍数前缀
    raw = re.sub(r"^\d+\s*[X×]\s*", "", raw).strip()
    if not re.fullmatch(r"\d{1,2}", raw):
        return False
    try:
        v = int(raw)
    except ValueError:
        return False
    if not (1 <= v <= 999):
        return False
    return dim.dim_no and normalize_dim_no(dim.dim_no) == raw


# ── 问题3/Fix5：修订历史 WAS 解析（支持 FOR #N、跨行） ─────────

def parse_revision_was(text):
    """从修订说明文本中提取 (dim_no, was_value) 候选。
    支持：
      'REVISED DIM #33 WAS Rz15-20'              → [('33','Rz15-20')]
      '1.5+0.10/-0.10 WAS 2.75+0.08/-0.12 FOR 70#'→ [('70','2.75+0.08/-0.12')]
      'WAS 2.75+0.08/-0.12 FOR\\n70#'            → [('70','2.75+0.08/-0.12')]
      'WAS 0.30'                                  → [('','0.30')]
    """
    if not text:
        return []
    flat = re.sub(r"[\r\n]+", " ", str(text))
    out = []
    # 模式A：DIM #N WAS value
    for m in re.finditer(
            r"DIM\s*#?\s*(\d+)\s+WAS\s+([A-Za-z0-9.+\-/±]+)", flat, re.IGNORECASE):
        out.append((normalize_dim_no(m.group(1)), m.group(2).strip()))
    # 模式B：WAS value ... FOR N#  （value 可含 +x/-y，可跨原换行）
    for m in re.finditer(
            r"WAS\s+([0-9][A-Za-z0-9.+\-/±\s]*?)\s+FOR\s+#?(\d+)\s*#?",
            flat, re.IGNORECASE):
        out.append((normalize_dim_no(m.group(2)),
                    re.sub(r"\s+", "", m.group(1))))
    # 模式C：裸 WAS value（仅当前两种都没命中，避免重复）
    if not out:
        for m in REVISION_WAS_PATTERN.finditer(flat):
            dim_no = normalize_dim_no(m.group(1)) if m.group(1) else ""
            was_val = (m.group(2) or "").strip()
            if was_val:
                out.append((dim_no, was_val))
    return out


# ══════════════════════════════════════════════════════════════════════
# 2. Structured Entity Extractor — 从 Excel 提取类型化实体
# ══════════════════════════════════════════════════════════════════════

class Dimension:
    """尺寸/检验实体"""
    __slots__ = ("dim_no", "is_ref", "drawing_raw", "normalized_value",
                 "nominal", "lower", "upper", "multiplier", "is_gdt",
                 "view", "raw", "row_data", "region_type", "is_general_tol")

    def __init__(self, dim_no="", is_ref=False, drawing_raw="",
                 nominal=None, lower=None, upper=None, multiplier=1,
                 is_gdt=False, view="", raw="", row_data=None):
        self.dim_no = str(dim_no).strip() if dim_no else ""
        self.is_ref = bool(is_ref)
        self.drawing_raw = str(drawing_raw or "")
        self.normalized_value = normalize_drawing_value(self.drawing_raw)
        self.nominal = nominal
        self.lower = lower
        self.upper = upper
        self.multiplier = int(multiplier) if multiplier else 1
        self.is_gdt = bool(is_gdt)
        self.view = normalize_text(view)
        self.raw = str(raw or "")
        self.row_data = row_data or {}
        # 区域标记（问题1/2）：
        #   region_type ∈ {"", "deviation_table"}
        #   is_general_tol: 是否标题栏标准公差表条目
        self.region_type = (
            "deviation_table"
            if is_deviation_region_text(view, raw, self.drawing_raw)
            else ""
        )
        self.is_general_tol = is_general_tol_entry(view, self.drawing_raw)

    @property
    def fingerprint(self):
        """内容指纹：匹配用，不含编号/视图/位置"""
        return (self.normalized_value, self.lower, self.upper)

    @property
    def content_key(self):
        """宽匹配键：仅按标称值，用于检测公差变更"""
        return self.normalized_value

    @property
    def is_seegdt(self):
        return self.normalized_value == "SEE GD&T" or self.is_gdt


class GDT:
    """GD&T 几何公差实体"""
    __slots__ = ("tolerance_val", "datums", "gdt_type", "view", "raw",
                 "row_data", "region_type")

    def __init__(self, tolerance_val="", datums=None, gdt_type="",
                 view="", raw="", row_data=None):
        # Fix2: 公差值做浮点归一化，消除 0.1 / 0.10 / 0.100 等格式差异
        raw_tol = normalize_text(tolerance_val)
        m = re.search(r"-?\d+(?:\.\d+)?", raw_tol)
        if m:
            try:
                self.tolerance_val = str(round(float(m.group(0)), 4))
            except (ValueError, TypeError):
                self.tolerance_val = raw_tol
        else:
            self.tolerance_val = raw_tol
        self.datums = frozenset(d.strip() for d in (datums or []) if d.strip())
        self.gdt_type = normalize_text(gdt_type)
        self.view = normalize_text(view)
        self.raw = str(raw or "")
        self.row_data = row_data or {}
        self.region_type = (
            "deviation_table"
            if is_deviation_region_text(view, raw)
            else ""
        )

    @property
    def fingerprint(self):
        """GD&T 内容指纹"""
        return (self.tolerance_val, self.datums)

    @property
    def display(self):
        return f"Tol:{self.tolerance_val} Datums:{'|'.join(sorted(self.datums))}"


class DeviationItem:
    """偏差清单条目"""
    __slots__ = ("dim_ref", "original_spec", "deviated_spec",
                 "norm_orig", "norm_dev", "dim_key")

    def __init__(self, dim_ref="", original_spec="", deviated_spec=""):
        self.dim_ref = normalize_text(dim_ref)
        self.original_spec = normalize_text(original_spec)
        self.deviated_spec = normalize_text(deviated_spec)
        self.norm_orig = normalize_drawing_value(original_spec)
        self.norm_dev = normalize_drawing_value(deviated_spec)
        # 归一化检验编号，用于与正式尺寸条目按 DIM NO. 关联（问题4）
        self.dim_key = normalize_dim_no(dim_ref)

    @property
    def fingerprint(self):
        """偏差条目唯一指纹：按 编号 + 偏差后规格"""
        return (self.dim_key, self.deviated_spec)


class SurfaceRoughness:
    """表面粗糙度实体（仅用于修订历史 WAS 关联，问题3）"""
    __slots__ = ("dim_no", "view", "param_type", "value", "raw", "norm_value")

    def __init__(self, dim_no="", view="", param_type="", value="", raw=""):
        self.dim_no = normalize_dim_no(dim_no)
        self.view = normalize_text(view)
        self.param_type = normalize_text(param_type)
        self.value = str(value or "")
        self.raw = str(raw or "")
        self.norm_value = normalize_text(value)


class NoteItem:
    """技术要求条目"""
    __slots__ = ("number", "level", "content_norm", "content_raw")

    def __init__(self, number="", level=0, content=""):
        self.number = normalize_text(number)
        self.level = int(level) if level else 0
        self.content_raw = str(content or "")
        self.content_norm = normalize_text(content)


class BOMItem:
    """BOM 物料条目"""
    __slots__ = ("part_number", "color", "material_spec", "assembly")

    def __init__(self, part_number="", color="", material_spec="", assembly=""):
        self.part_number = normalize_text(part_number)
        self.color = normalize_text(color)
        self.material_spec = normalize_text(material_spec)
        self.assembly = normalize_text(assembly)

    @property
    def fingerprint(self):
        return (self.part_number, self.material_spec, self.assembly)


class KeyCharItem:
    """关键特性条目"""
    __slots__ = ("char_type", "spec", "description")

    def __init__(self, char_type="", spec="", description=""):
        self.char_type = normalize_text(char_type)
        self.spec = normalize_text(spec)
        self.description = normalize_text(description)

    @property
    def fingerprint(self):
        return (self.char_type, self.spec, self.description)


class StructuredEntityExtractor:
    """从识别结果 Excel 中提取结构化实体"""

    def __init__(self, xlsx_path):
        self.path = xlsx_path
        self.wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    # ── 尺寸提取（检验报告表） ────────────────────────────

    def extract_dimensions(self) -> list:
        """从检验报告表中提取所有尺寸实体。
        处理：
          - dim_no 为 None 的倍数展开行（只保留第一行）
          - SEE GD&T 行
          - 粗糙度行（带 '粗糙度:' 前缀）
        """
        dims = []
        sheet_name = None
        for sn in ("📋 检验报告表", "检验报告表"):
            if sn in self.wb.sheetnames:
                sheet_name = sn
                break
        if not sheet_name:
            return dims

        ws = self.wb[sheet_name]
        seen_fingerprint = set()
        last_drawing = None  # 跟踪倍数展开
        last_lower = None
        last_upper = None
        multiplier_count = 0

        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or len(r) < 5:
                continue
            dim_no = r[0]
            ref = r[1] if len(r) > 1 else ""
            drawing = r[2] if len(r) > 2 else ""
            lower = r[3] if len(r) > 3 else ""
            upper = r[4] if len(r) > 4 else ""
            view = r[5] if len(r) > 5 else ""
            raw = r[6] if len(r) > 6 else ""

            # 跳过空行
            drawing_str = str(drawing or "").strip()
            if not drawing_str:
                continue

            # 跳过粗糙度行（它们在单独的 surface_roughness 类型中处理）
            if drawing_str.lower().startswith("粗糙度"):
                continue

            is_seegdt = (drawing_str == "SEE GD&T")
            is_ref = str(ref or "").strip().upper() in ("REF", "✓", "TRUE", "1")

            if not is_seegdt:
                # 检测倍数展开行 (dim_no 为 None 的行)
                has_dim_no = (dim_no is not None and str(dim_no).strip() != "")
                if not has_dim_no:
                    # 倍数展开行：检查是否与前一行内容相同
                    cur_drawing = str(drawing or "").strip()
                    cur_lower = str(lower or "").strip()
                    cur_upper = str(upper or "").strip()
                    if (cur_drawing == last_drawing and
                        cur_lower == last_lower and
                        cur_upper == last_upper):
                        multiplier_count += 1
                        continue

                last_drawing = str(drawing or "").strip()
                last_lower = str(lower or "").strip()
                last_upper = str(upper or "").strip()
                multiplier_count = 0
            else:
                multiplier_count = 0

            # 解析
            nominal = parse_nominal(drawing_str)
            lo, hi = parse_tolerance_bounds(lower, upper)

            entity = Dimension(
                dim_no=dim_no,
                is_ref=is_ref,
                drawing_raw=drawing_str,
                nominal=nominal,
                lower=lo,
                upper=hi,
                multiplier=1,
                is_gdt=is_seegdt,
                view=view,
                raw=raw,
                row_data={
                    "dim_no_raw": dim_no,
                    "ref_raw": ref,
                    "drawing_raw": drawing_str,
                    "lower_raw": lower,
                    "upper_raw": upper,
                    "view_raw": view,
                    "raw_raw": raw,
                },
            )

            # 去重：同指纹只保留一个
            fp = entity.fingerprint
            if fp not in seen_fingerprint:
                seen_fingerprint.add(fp)
                dims.append(entity)

        return dims

    # ── GD&T 提取 ──────────────────────────────────────

    def extract_gdt(self) -> list:
        gdts = []
        for sn in ("⊕ GD&T关联(AI)",):
            if sn not in self.wb.sheetnames:
                continue
            ws = self.wb[sn]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r or len(r) < 5:
                    continue
                view = r[1] if len(r) > 1 else ""
                tol_val = r[2] if len(r) > 2 else ""
                datums_str = r[3] if len(r) > 3 else ""
                gdt_max = r[4] if len(r) > 4 else ""
                raw = r[5] if len(r) > 5 else ""

                # 跳过空行
                if not str(tol_val or "").strip():
                    continue

                # 基准可能是逗号分隔或空格分隔
                datums = []
                if datums_str:
                    d_str = str(datums_str).strip()
                    # 逗号分隔
                    if "," in d_str:
                        datums = [d.strip() for d in d_str.split(",") if d.strip()]
                    else:
                        # 空格分隔
                        datums = [d.strip() for d in d_str.split() if d.strip()]

                gdts.append(GDT(
                    tolerance_val=gdt_max if gdt_max else tol_val,
                    datums=datums,
                    gdt_type="",
                    view=view,
                    raw=raw,
                    row_data={"view": view, "tol": tol_val, "datums": datums_str,
                              "gdt_max": gdt_max, "raw": raw},
                ))
        return gdts

    # ── 偏差清单提取 ────────────────────────────────

    def extract_deviations(self) -> list:
        devs = []
        for sn in ("📋 偏差清单",):
            if sn not in self.wb.sheetnames:
                continue
            ws = self.wb[sn]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r or len(r) < 3:
                    continue
                dim_ref = r[0] if r[0] else ""
                orig = r[1] if len(r) > 1 else ""
                deviated = r[2] if len(r) > 2 else ""
                if not str(dim_ref or "").strip() and not str(orig or "").strip():
                    continue
                devs.append(DeviationItem(
                    dim_ref=str(dim_ref or ""),
                    original_spec=str(orig or ""),
                    deviated_spec=str(deviated or ""),
                ))
        return devs

    # ── 技术要求提取 ────────────────────────────────

    def extract_notes(self) -> list:
        notes = []
        for sn in ("📝 技术要求",):
            if sn not in self.wb.sheetnames:
                continue
            ws = self.wb[sn]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r or len(r) < 3:
                    continue
                num = r[0] if r[0] else ""
                level = r[1] if len(r) > 1 else 0
                content = r[2] if len(r) > 2 else ""
                if not str(content or "").strip():
                    continue
                notes.append(NoteItem(number=num, level=level, content=content))
        return notes

    # ── BOM 提取 ──────────────────────────────────

    def extract_bom(self) -> list:
        boms = []
        for sn in ("📦 物料清单BOM",):
            if sn not in self.wb.sheetnames:
                continue
            ws = self.wb[sn]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r or len(r) < 3:
                    continue
                part = r[1] if len(r) > 1 else ""
                color = r[2] if len(r) > 2 else ""
                mat = r[3] if len(r) > 3 else ""
                asm = r[4] if len(r) > 4 else ""
                if not str(part or "").strip():
                    continue
                boms.append(BOMItem(part_number=part, color=color,
                                    material_spec=mat, assembly=asm))
        return boms

    # ── 关键特性提取 ──────────────────────────────

    def extract_key_chars(self) -> list:
        chars = []
        for sn in ("⭐ 关键特性",):
            if sn not in self.wb.sheetnames:
                continue
            ws = self.wb[sn]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r or len(r) < 3:
                    continue
                ctype = r[1] if len(r) > 1 else ""
                spec = r[2] if len(r) > 2 else ""
                desc = r[3] if len(r) > 3 else ""
                if not str(ctype or "").strip():
                    continue
                chars.append(KeyCharItem(char_type=ctype, spec=spec, description=desc))
        return chars

    # ── 修订历史提取（问题3） ─────────────────────
    def extract_revisions(self) -> list:
        """从 '🔄 修订历史' sheet 提取修订记录。
        返回 [{"rev","date","desc","ecn"}, ...]
        """
        revs = []
        for sn in ("🔄 修订历史", "修订历史"):
            if sn not in self.wb.sheetnames:
                continue
            ws = self.wb[sn]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r:
                    continue
                rev = r[0] if len(r) > 0 else ""
                date = r[1] if len(r) > 1 else ""
                desc = r[2] if len(r) > 2 else ""
                ecn = r[3] if len(r) > 3 else ""
                if not str(desc or "").strip() and not str(rev or "").strip():
                    continue
                revs.append({
                    "rev": str(rev or "").strip(),
                    "date": str(date or "").strip(),
                    "desc": str(desc or "").strip(),
                    "ecn": str(ecn or "").strip(),
                })
        return revs

    # ── 表面粗糙度提取（问题3 WAS 关联用） ────────
    def extract_surface_roughness(self) -> list:
        items = []
        for sn in ("🔲 表面粗糙度(AI)",):
            if sn not in self.wb.sheetnames:
                continue
            ws = self.wb[sn]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r or len(r) < 5:
                    continue
                dim_no = r[1] if len(r) > 1 else ""
                view = r[2] if len(r) > 2 else ""
                param = r[3] if len(r) > 3 else ""
                value = r[4] if len(r) > 4 else ""
                raw = r[5] if len(r) > 5 else ""
                if not str(value or "").strip():
                    continue
                items.append(SurfaceRoughness(
                    dim_no=dim_no, view=view, param_type=param,
                    value=value, raw=raw))
        return items

    # ── 图纸信息提取 ──────────────────────────────

    def extract_info(self) -> dict:
        info = {}
        sn = "📋 图纸信息"
        if sn not in self.wb.sheetnames:
            return info
        ws = self.wb[sn]
        for r in ws.iter_rows(values_only=True):
            if r and len(r) >= 2 and r[0] is not None and r[1] is not None:
                k = str(r[0]).strip()
                v = str(r[1]).strip()
                if k and v:
                    info[k] = v
        return info


# ══════════════════════════════════════════════════════════════════════
# 3. Smart Matcher — 内容指纹匹配算法
# ══════════════════════════════════════════════════════════════════════

class SmartMatcher:
    """按内容指纹进行实体匹配，而不是按序号"""

    def __init__(self, entities_a: list, entities_b: list):
        self.a = entities_a
        self.b = entities_b

    def match(self) -> dict:
        """执行匹配，返回匹配结果字典。
        返回:
          {
            "matched": [(entity_a, entity_b, match_type), ...],
            "only_in_a": [entity, ...],    # 真实删除
            "only_in_b": [entity, ...],    # 真实新增
            "re_numbered": [(entity_a, entity_b), ...],  # 重编号
          }
        """
        if not self.a or not self.b:
            return {
                "matched": [],
                "only_in_a": list(self.a),
                "only_in_b": list(self.b),
                "re_numbered": [],
            }

        # 1. 构建指纹索引
        fp_a = defaultdict(list)  # fingerprint → [entities in A]
        fp_b = defaultdict(list)  # fingerprint → [entities in B]
        for e in self.a:
            fp_a[e.fingerprint].append(e)
        for e in self.b:
            fp_b[e.fingerprint].append(e)

        used_a = set()  # index of entity in self.a
        used_b = set()
        matched = []    # (entity_a, entity_b, "exact"|"fuzzy"|"value_only")
        re_numbered = []

        # 2. 精确指纹匹配
        fps_common = set(fp_a.keys()) & set(fp_b.keys())
        for fp in fps_common:
            list_a = fp_a[fp]
            list_b = fp_b[fp]
            for i in range(min(len(list_a), len(list_b))):
                idx_a = self.a.index(list_a[i])
                idx_b = self.b.index(list_b[i])
                if idx_a not in used_a and idx_b not in used_b:
                    used_a.add(idx_a)
                    used_b.add(idx_b)
                    matched.append((list_a[i], list_b[i], "exact"))

        # 3. 模糊匹配：同标称值但公差略有不同（浮点精度）
        # 仅对具有 normalized_value 属性的实体类型启用
        remaining_a = [self.a[i] for i in range(len(self.a)) if i not in used_a]
        remaining_b = [self.b[i] for i in range(len(self.b)) if i not in used_b]

        has_normalized_value = (remaining_a and
                                hasattr(remaining_a[0], "normalized_value"))
        if has_normalized_value:
            val_a = defaultdict(list)
            val_b = defaultdict(list)
            for e in remaining_a:
                val_a[e.normalized_value].append(e)
            for e in remaining_b:
                val_b[e.normalized_value].append(e)

            common_vals = set(val_a.keys()) & set(val_b.keys())
            for v in common_vals:
                list_a = val_a[v]
                list_b = val_b[v]
                for i in range(min(len(list_a), len(list_b))):
                    ea, eb = list_a[i], list_b[i]
                    idx_a = self.a.index(ea)
                    idx_b = self.b.index(eb)
                    if idx_a not in used_a and idx_b not in used_b:
                        used_a.add(idx_a)
                        used_b.add(idx_b)
                        # 检查是否仅公差变化（真正的修改）
                        has_tol = (hasattr(ea, "lower") and hasattr(eb, "lower"))
                        tol_same = True
                        if has_tol:
                            tol_same = (ea.lower == eb.lower and
                                       ea.upper == eb.upper)
                        match_type = "value_only" if not tol_same else "fuzzy"
                        matched.append((ea, eb, match_type))

        # 4. SEE GD&T 特殊匹配：按 GD&T 类型匹配
        # 仅对具有 is_seegdt 属性的实体类型启用
        if remaining_a and hasattr(remaining_a[0], "is_seegdt"):
            gdt_a = [(i, e) for i, e in enumerate(self.a)
                     if e.is_seegdt and i not in used_a]
            gdt_b = [(i, e) for i, e in enumerate(self.b)
                     if e.is_seegdt and i not in used_b]

            # SEE GD&T 按位置顺序简单匹配（它们没有标称值可以依靠）
            for pos_idx in range(min(len(gdt_a), len(gdt_b))):
                ia, ea = gdt_a[pos_idx]
                ib, eb = gdt_b[pos_idx]
                if ia not in used_a and ib not in used_b:
                    used_a.add(ia)
                    used_b.add(ib)
                    matched.append((ea, eb, "gdt_positional"))

        # 5. 收集剩余项
        only_in_a = [self.a[i] for i in range(len(self.a)) if i not in used_a]
        only_in_b = [self.b[i] for i in range(len(self.b)) if i not in used_b]

        # 6. 重编号检测（仅对具有 dim_no 属性的实体类型）
        has_dim_no = (self.a and hasattr(self.a[0], "dim_no"))
        if has_dim_no:
            total_a = len(self.a)
            total_b = len(self.b)
            if total_a > 0 and total_b > 0:
                matched_but_different_no = 0
                for ea, eb, _ in matched:
                    if ea.dim_no and eb.dim_no and ea.dim_no != eb.dim_no:
                        re_numbered.append((ea, eb))
                        matched_but_different_no += 1
                is_renumbered = (
                    matched_but_different_no >= max(3, total_a * RENUMBER_THRESHOLD)
                )

                if is_renumbered and only_in_a and only_in_b:
                    re_only_a = []
                    for e in only_in_a:
                        found = None
                        for j, eb in enumerate(only_in_b):
                            if e.fingerprint == eb.fingerprint:
                                found = (j, eb)
                                break
                        if found:
                            j, eb = found
                            matched.append((e, eb, "renumber_match"))
                            re_numbered.append((e, eb))
                            only_in_b.pop(j)
                        else:
                            re_only_a.append(e)
                    only_in_a = re_only_a
            else:
                is_renumbered = False
        else:
            is_renumbered = False

        return {
            "matched": matched,
            "only_in_a": only_in_a,
            "only_in_b": only_in_b,
            "re_numbered": re_numbered,
            "is_renumbered": is_renumbered,
        }


# ══════════════════════════════════════════════════════════════════════
# 4. Change Classifier — 置信度评分
# ══════════════════════════════════════════════════════════════════════

class ChangeResult:
    """单条变更记录"""
    __slots__ = ("change_type", "entity_type", "confidence",
                 "old_entity", "new_entity", "details", "reason")

    def __init__(self, change_type, entity_type, confidence,
                 old_entity=None, new_entity=None, details=None, reason=""):
        self.change_type = change_type    # added / removed / modified
        self.entity_type = entity_type    # dimension / gdt / note / bom / ...
        self.confidence = confidence       # high / medium / low
        self.old_entity = old_entity
        self.new_entity = new_entity
        self.details = details or {}
        self.reason = reason


class ChangeClassifier:
    """根据匹配结果对变更进行分类和置信度评分"""

    @staticmethod
    def classify_matched(ea, eb, match_type, entity_type="dimension"):
        """对匹配对进行分类：判断是真实修改还是无影响差异。
        如果没有实际变化，返回 None（不报告）。
        """
        details = {}
        reasons = []

        # 检查标称值变化
        nominal_changed = not fuzzy_val_match(ea.nominal, eb.nominal)
        tol_changed = not (fuzzy_val_match(ea.lower, eb.lower) and
                          fuzzy_val_match(ea.upper, eb.upper))
        view_changed = (ea.view != eb.view)
        dim_no_changed = (ea.dim_no != eb.dim_no)
        drawing_text_diff = (normalize_text(ea.drawing_raw) !=
                            normalize_text(eb.drawing_raw))

        # 如果完全无变化，不报告
        if not (nominal_changed or tol_changed or view_changed or
                dim_no_changed or drawing_text_diff):
            return None

        if nominal_changed:
            details["nominal"] = f"{ea.nominal} → {eb.nominal}"
            reasons.append(f"Nominal changed from {ea.nominal} to {eb.nominal}")
        if tol_changed:
            details["tolerance"] = f"({ea.lower},{ea.upper}) → ({eb.lower},{eb.upper})"
            reasons.append(f"Tolerance band changed")
        if view_changed:
            details["view"] = f"{ea.view} → {eb.view}"
            reasons.append(f"View label changed: {ea.view} → {eb.view}")
        if dim_no_changed:
            details["dim_no"] = f"{ea.dim_no} → {eb.dim_no}"
        if drawing_text_diff and not nominal_changed:
            details["drawing_text"] = f"{ea.drawing_raw} → {eb.drawing_raw}"

        # 置信度判定
        if nominal_changed or tol_changed:
            confidence = "high"
            if not reasons:
                reasons.append("Nominal or tolerance value changed")
        elif view_changed and not nominal_changed and not tol_changed:
            # 纯视图变化（标称/公差未变）
            if is_view_label_ocr_noise(ea.view, eb.view):
                confidence = "low"
                reasons = ["OCR noise: view label minor difference"]
            else:
                confidence = "medium"
                if not reasons:
                    reasons.append(f"View assignment changed: {ea.view} → {eb.view}")
        elif drawing_text_diff and not nominal_changed and not tol_changed:
            # 纯文本差异（标称/公差未变）
            if is_ocr_only_difference(ea.drawing_raw, eb.drawing_raw):
                confidence = "low"
                reasons = ["OCR noise: case/whitespace difference only"]
            else:
                confidence = "medium"
                reasons.append("Drawing text differs but nominal+tolerance match")
        elif dim_no_changed and not nominal_changed and not tol_changed:
            confidence = "low"
            reasons = ["Re-numbered: inspection number changed, content matches"]
        else:
            # 退化情况：同一实体在二次识别中可能有微小差异
            confidence = "low"
            reasons = ["Minor difference detected"]

        return ChangeResult(
            change_type="modified",
            entity_type=entity_type,
            confidence=confidence,
            old_entity=ea,
            new_entity=eb,
            details=details,
            reason="; ".join(reasons),
        )

    @staticmethod
    def classify_added(entity, entity_type="dimension",
                       is_renumbered=False) -> ChangeResult:
        """分类新增实体"""
        if is_renumbered:
            return ChangeResult(
                change_type="added", entity_type=entity_type,
                confidence="low", new_entity=entity,
                reason="Likely re-numbered; content same as existing in old revision",
            )
        return ChangeResult(
            change_type="added", entity_type=entity_type,
            confidence="high", new_entity=entity,
            reason="New entity with unique content fingerprint",
        )

    @staticmethod
    def classify_removed(entity, entity_type="dimension",
                         is_renumbered=False) -> ChangeResult:
        """分类删除实体"""
        if is_renumbered:
            return ChangeResult(
                change_type="removed", entity_type=entity_type,
                confidence="low", old_entity=entity,
                reason="Likely re-numbered; content same as existing in new revision",
            )
        return ChangeResult(
            change_type="removed", entity_type=entity_type,
            confidence="high", old_entity=entity,
            reason="Entity removed with no matching content in new revision",
        )


# ══════════════════════════════════════════════════════════════════════
# 5. Smart Diff Engine — 编排器
# ══════════════════════════════════════════════════════════════════════

class SmartDiffEngine:
    """智能差异比对引擎：结构实体匹配 → 噪声过滤 → 分类 → 报告"""

    def __init__(self, path_a: str, path_b: str):
        self.path_a = path_a
        self.path_b = path_b
        self.ext_a = StructuredEntityExtractor(path_a)
        self.ext_b = StructuredEntityExtractor(path_b)

    def compare_all(self) -> dict:
        """全量比对：所有实体类型 + 图纸信息"""
        changes = []
        renumbering_info = {}
        deviation_overrides = 0

        # ── 尺寸对比 ─────────────────────────────
        dims_a_all = self.ext_a.extract_dimensions()
        dims_b_all = self.ext_b.extract_dimensions()

        # 区域分流（问题1 & 问题2）：
        #   general_tol  : 标题栏标准公差表 → 归一化豁免比对
        #   dev_origin   : 偏差表区域来源 → 不当作正式尺寸增删，交由偏差覆盖逻辑处理
        #   main         : 正式尺寸标注 → 正常指纹匹配
        def _partition(dims):
            general, dev, main = [], [], []
            for d in dims:
                if d.is_general_tol:
                    general.append(d)
                elif d.region_type == "deviation_table":
                    dev.append(d)
                else:
                    main.append(d)
            return general, dev, main

        gtol_a, dev_origin_dims_a, dims_a = _partition(dims_a_all)
        gtol_b, dev_origin_dims_b, dims_b = _partition(dims_b_all)

        matcher = SmartMatcher(dims_a, dims_b)
        result = matcher.match()

        # 分类匹配项（跳过无实际变化的 None）
        for ea, eb, mt in result["matched"]:
            cr = ChangeClassifier.classify_matched(ea, eb, mt, "dimension")
            if cr is not None:
                changes.append(cr)

        is_renumbered = result.get("is_renumbered", False)
        renum_pairs = result.get("re_numbered", [])
        if renum_pairs:
            renumbering_info["dimensions"] = {
                "count": len(renum_pairs),
                "is_global_renumber": is_renumbered,
                "pairs": [
                    {"old_dim_no": ea.dim_no, "new_dim_no": eb.dim_no,
                     "value": ea.normalized_value}
                    for ea, eb in renum_pairs
                ],
            }

        # 主尺寸的真实增删；保留 only_in_b 引用以便后续被修订/偏差关联消费
        removed_dims = list(result["only_in_a"])
        added_dims = list(result["only_in_b"])

        # ── 问题1：标准公差表 归一化豁免比对 ───────────
        # 两版均含该表时差异来自 OCR 抖动 → 静默丢弃；
        # 只有一版有该表时才以 low 置信度提示。
        gtol_key_a = {normalize_general_tol(d.drawing_raw) for d in gtol_a}
        gtol_key_b = {normalize_general_tol(d.drawing_raw) for d in gtol_b}
        gtol_map_a = {normalize_general_tol(d.drawing_raw): d for d in gtol_a}
        gtol_map_b = {normalize_general_tol(d.drawing_raw): d for d in gtol_b}
        both_have_gtol = bool(gtol_a) and bool(gtol_b)
        if not both_have_gtol:
            for k in gtol_key_a - gtol_key_b:
                changes.append(ChangeResult(
                    "removed", "dimension", "low", old_entity=gtol_map_a[k],
                    details={"region": "title_block_general_tolerance"},
                    reason="Standard tolerance table entry only differs after OCR "
                           "normalization (exempted region)"))
            for k in gtol_key_b - gtol_key_a:
                changes.append(ChangeResult(
                    "added", "dimension", "low", new_entity=gtol_map_b[k],
                    details={"region": "title_block_general_tolerance"},
                    reason="Standard tolerance table entry only differs after OCR "
                           "normalization (exempted region)"))
        # else: 两版均含该表 → 纯 OCR 抖动，静默丢弃

        # ── GD&T 对比 ─────────────────────────────
        gdt_a_all = self.ext_a.extract_gdt()
        gdt_b_all = self.ext_b.extract_gdt()
        # 偏差表区域来源的 GD&T 不参与正式 GD&T 增删（问题2）
        gdt_a = [g for g in gdt_a_all if g.region_type != "deviation_table"]
        gdt_b = [g for g in gdt_b_all if g.region_type != "deviation_table"]
        dev_origin_gdt_b = [g for g in gdt_b_all if g.region_type == "deviation_table"]
        gdt_matcher = SmartMatcher(gdt_a, gdt_b)
        gdt_result = gdt_matcher.match()

        for ea, eb, mt in gdt_result["matched"]:
            tol_same = (ea.tolerance_val == eb.tolerance_val)
            datums_same = (ea.datums == eb.datums)
            if tol_same and datums_same:
                if ea.view != eb.view:
                    changes.append(ChangeResult(
                        "modified", "gdt", "low", ea, eb, {},
                        f"View label changed: {ea.view} → {eb.view} (likely OCR noise)"))
                else:
                    # 完全匹配，不报告
                    pass
            else:
                details = {}
                if not tol_same:
                    details["tolerance"] = f"{ea.tolerance_val} → {eb.tolerance_val}"
                if not datums_same:
                    details["datums"] = f"{'|'.join(sorted(ea.datums))} → {'|'.join(sorted(eb.datums))}"
                changes.append(ChangeResult(
                    "modified", "gdt", "high", ea, eb, details,
                    "GD&T tolerance or datums changed"))

        for e in gdt_result["only_in_a"]:
            changes.append(ChangeResult("removed", "gdt", "high", old_entity=e,
                                        reason="GD&T removed"))
        for e in gdt_result["only_in_b"]:
            changes.append(ChangeResult("added", "gdt", "high", new_entity=e,
                                        reason="GD&T added"))

        # ══════════════════════════════════════════════════
        # 问题3：修订历史 "WAS xxx" → 强制关联为 modified
        # ══════════════════════════════════════════════════
        # 解析新版修订说明中的 "REVISED DIM #N WAS xxx"，对候选尺寸做强制关联，
        # 并抑制由此衍生的孤立 added/removed/note。
        revs_b = self.ext_b.extract_revisions()
        rough_a = self.ext_a.extract_surface_roughness()
        rough_b = self.ext_b.extract_surface_roughness()

        # 候选搜索池：主尺寸 + 粗糙度（按归一化文本/编号检索）
        def _norm_raw(s):
            return normalize_text(s)

        was_candidates = []
        for rev in revs_b:
            was_candidates.extend(parse_revision_was(rev["desc"]))

        consumed_added = set()    # id() of added_dims consumed by revision/deviation
        consumed_removed = set()
        suppress_note_tokens = set()  # 抑制含这些 WAS 值的 added note

        for dim_no, was_val in was_candidates:
            was_norm = _norm_raw(was_val)
            suppress_note_tokens.add(was_norm)

            # 旧版一侧：值等于 WAS 值的实体（尺寸优先，其次粗糙度）
            old_match = None
            for e in removed_dims:
                if _norm_raw(e.drawing_raw) == was_norm and id(e) not in consumed_removed:
                    old_match = e
                    break
            if old_match is None:
                for e in rough_a:
                    if e.norm_value == was_norm:
                        old_match = e
                        break

            # 新版一侧：按 WAS 值的"类型"选择搜索池，避免误关联未变化的同号尺寸
            new_match = None
            rough_prefix = re.match(r"^(RA|RZ|RMAX)", was_norm)
            is_rough_was = bool(rough_prefix)

            if is_rough_was:
                # 粗糙度类：优先在粗糙度池中找（按编号，再按同前缀不同值）
                pfx = rough_prefix.group(1)
                if dim_no:
                    new_match = next(
                        (e for e in rough_b if e.dim_no == dim_no), None)
                if new_match is None:
                    new_match = next(
                        (e for e in rough_b
                         if e.norm_value.startswith(pfx)
                         and e.norm_value != was_norm), None)
            else:
                # 尺寸类：只在"真正新增"的尺寸中按编号关联，避免抓到未变化尺寸
                if dim_no:
                    new_match = next(
                        (e for e in added_dims
                         if normalize_dim_no(e.dim_no) == dim_no
                         and id(e) not in consumed_added), None)

            if old_match is not None or new_match is not None:
                old_disp = getattr(old_match, "drawing_raw", None) \
                    or getattr(old_match, "value", "") or was_val
                new_disp = getattr(new_match, "drawing_raw", None) \
                    or getattr(new_match, "value", "") or "(see revision)"
                # 标记消费，避免重复报告
                if old_match in removed_dims:
                    consumed_removed.add(id(old_match))
                if new_match in added_dims:
                    consumed_added.add(id(new_match))
                # 仅当 old_match/new_match 是 Dimension 时给 dim 字段
                is_dim = isinstance(new_match, Dimension) or isinstance(old_match, Dimension)
                changes.append(ChangeResult(
                    "modified",
                    "dimension" if is_dim else "surface_roughness",
                    "high",
                    old_entity=old_match if isinstance(old_match, Dimension) else None,
                    new_entity=new_match if isinstance(new_match, Dimension) else None,
                    details={
                        "revision": "revision_history_confirmed",
                        "change": f"{old_disp} → {new_disp}",
                        "dim_no": dim_no or "",
                    },
                    reason=f"Revision history: WAS {was_val}"
                           + (f" (DIM #{dim_no})" if dim_no else ""),
                ))

        # ══════════════════════════════════════════════════
        # 问题2 & 问题4：偏差清单(DEVIATION TABLE) 比对
        # ══════════════════════════════════════════════════
        devs_a = self.ext_a.extract_deviations()
        devs_b = self.ext_b.extract_deviations()
        dev_a_map = {d.fingerprint: d for d in devs_a}
        dev_b_map = {d.fingerprint: d for d in devs_b}

        # 当前版本所有尺寸（含偏差来源）按归一化编号建索引，用于 DIM NO. 关联
        def _build_dimno_index(dim_lists):
            idx = {}
            for lst in dim_lists:
                for d in lst:
                    k = normalize_dim_no(d.dim_no)
                    if k and k not in idx:
                        idx[k] = d
            return idx

        dimno_idx_b = _build_dimno_index([dims_b, dev_origin_dims_b])
        dimno_idx_a = _build_dimno_index([dims_a, dev_origin_dims_a])

        deviation_overrides = 0

        # 新增的偏差覆盖（B 有、A 无）
        for fp, dev in dev_b_map.items():
            if fp in dev_a_map:
                continue
            deviation_overrides += 1
            linked = dimno_idx_b.get(dev.dim_key) or dimno_idx_a.get(dev.dim_key)
            if linked is not None:
                # 问题4：关联到既有尺寸，作为该尺寸的偏差附加属性输出
                changes.append(ChangeResult(
                    "modified", "dimension", "high",
                    old_entity=linked if linked in dims_a or linked in dev_origin_dims_a else None,
                    new_entity=linked if linked in dims_b or linked in dev_origin_dims_b else None,
                    details={
                        "deviation_override_added": dev.deviated_spec,
                        "dim_no": dev.dim_ref,
                        "original_spec": dev.original_spec,
                    },
                    reason=f"Deviation override added for DIM {dev.dim_ref}: "
                           f"{dev.original_spec} → {dev.deviated_spec}",
                ))
            else:
                # 问题2：无法关联到既有尺寸 → 归类为 deviation_override，而非 dimension
                changes.append(ChangeResult(
                    "added", "deviation_override", "medium",
                    new_entity=dev,
                    details={
                        "dim_no": dev.dim_ref,
                        "original_spec": dev.original_spec,
                        "deviated_spec": dev.deviated_spec,
                    },
                    reason=f"New deviation table entry DIM {dev.dim_ref}",
                ))

        # 移除的偏差覆盖（A 有、B 无）
        for fp, dev in dev_a_map.items():
            if fp in dev_b_map:
                continue
            deviation_overrides += 1
            changes.append(ChangeResult(
                "removed", "deviation_override", "medium",
                old_entity=dev,
                details={
                    "dim_no": dev.dim_ref,
                    "original_spec": dev.original_spec,
                    "deviated_spec": dev.deviated_spec,
                },
                reason=f"Deviation table entry removed: DIM {dev.dim_ref}",
            ))

        # 偏差表来源的伪尺寸/伪 GD&T（问题2）：统一以 deviation_override 计入，
        # 不再作为 added,high,dimension / added,high,gdt 出现。
        # 已通过区域分流从主比对中剔除，这里仅在它们未被偏差清单覆盖时给出低噪声提示。
        accounted_dev_keys = {d.dim_key for d in devs_b}
        for d in dev_origin_dims_b:
            if normalize_dim_no(d.dim_no) not in accounted_dev_keys:
                changes.append(ChangeResult(
                    "added", "deviation_override", "low",
                    new_entity=d,
                    details={"region": "deviation_table",
                             "value": d.drawing_raw, "dim_no": d.dim_no},
                    reason="Entry inside deviation table region "
                           "(not a formal dimension change)",
                ))

        # ── Fix1: 偏差表回显启发式检测 ──────────────────
        # 偏差表会把已有尺寸连同"偏差后公差"再列一遍。这类回显条目的特征是：
        # 带公差、且其标称值在本版另一条尺寸上已出现。把它们从 added,high,dimension
        # 改判为 deviation_override，并计入 Deviation Overrides。
        from collections import Counter as _Counter
        val_count_b = _Counter(d.normalized_value for d in dims_b if d.normalized_value)
        echo_consumed = set()
        for e in added_dims:
            if id(e) in consumed_added:
                continue
            has_tol = (e.lower is not None or e.upper is not None)
            if has_tol and e.normalized_value and val_count_b[e.normalized_value] >= 2:
                echo_consumed.add(id(e))
                deviation_overrides += 1
                changes.append(ChangeResult(
                    "added", "deviation_override", "medium", new_entity=e,
                    details={"region": "deviation_table_echo",
                             "value": e.drawing_raw, "dim_no": e.dim_no,
                             "tolerance": f"({e.lower},{e.upper})"},
                    reason="Re-listed existing dimension with deviation tolerance "
                           "(deviation table echo)"))
        consumed_added |= echo_consumed

        # ── 输出残余主尺寸增删（扣除被修订/偏差消费的项） ──
        for e in removed_dims:
            if id(e) in consumed_removed:
                continue
            # Fix3: TITLE_BLOCK 区域带编号的条目极可能是标题栏内容被误标气泡号，
            # 不以 high 报删除
            if e.view == "TITLE_BLOCK" and e.dim_no:
                changes.append(ChangeResult(
                    "removed", "dimension", "low", old_entity=e,
                    details={"region": "title_block", "value": e.drawing_raw},
                    reason="Title block entry carrying a dim_no — likely OCR mis-label"))
                continue
            changes.append(
                ChangeClassifier.classify_removed(e, "dimension", is_renumbered))
        for e in added_dims:
            if id(e) in consumed_added:
                continue
            # Fix4: 气泡序号被误读成尺寸——value 就是编号本身(纯小整数)且无公差
            if _is_bubble_sequence_artifact(e):
                changes.append(ChangeResult(
                    "added", "dimension", "low", new_entity=e,
                    details={"artifact": "bubble_number_sequence",
                             "value": e.drawing_raw, "dim_no": e.dim_no},
                    reason="Value equals its own bubble number with no tolerance — "
                           "likely an inspection-bubble index misread as a dimension"))
                continue
            # Fix3: 同理处理 TITLE_BLOCK 新增噪声
            if e.view == "TITLE_BLOCK" and e.dim_no:
                changes.append(ChangeResult(
                    "added", "dimension", "low", new_entity=e,
                    details={"region": "title_block", "value": e.drawing_raw},
                    reason="Title block entry carrying a dim_no — likely OCR mis-label"))
                continue
            changes.append(
                ChangeClassifier.classify_added(e, "dimension", is_renumbered))

        # ── 技术要求对比 ──────────────────────────
        notes_a = self.ext_a.extract_notes()
        notes_b = self.ext_b.extract_notes()
        note_contents_a = {n.content_norm: n for n in notes_a}
        note_contents_b = {n.content_norm: n for n in notes_b}

        for content, nb in note_contents_b.items():
            if content not in note_contents_a:
                # 问题3：抑制由修订说明 "WAS xxx" 衍生的孤立新增 note
                if any(tok and tok in content for tok in suppress_note_tokens) \
                        and "WAS" in content:
                    continue
                changes.append(ChangeResult("added", "note", "high", new_entity=nb,
                                            reason="New note added"))
        for content, na in note_contents_a.items():
            if content not in note_contents_b:
                changes.append(ChangeResult("removed", "note", "high", old_entity=na,
                                            reason="Note removed"))
        for content in set(note_contents_a) & set(note_contents_b):
            na, nb = note_contents_a[content], note_contents_b[content]
            if na.number != nb.number:
                changes.append(ChangeResult(
                    "modified", "note", "low", na, nb,
                    {"number": f"{na.number} → {nb.number}"},
                    "Note re-numbered, content unchanged"))

        # ── BOM 对比 ──────────────────────────────
        bom_a = self.ext_a.extract_bom()
        bom_b = self.ext_b.extract_bom()
        bom_matcher = SmartMatcher(bom_a, bom_b)
        bom_result = bom_matcher.match()

        for e in bom_result["only_in_a"]:
            changes.append(ChangeResult("removed", "bom", "high", old_entity=e,
                                        reason="BOM item removed"))
        for e in bom_result["only_in_b"]:
            changes.append(ChangeResult("added", "bom", "high", new_entity=e,
                                        reason="BOM item added"))
        for ea, eb, mt in bom_result["matched"]:
            if (ea.color != eb.color or ea.material_spec != eb.material_spec):
                changes.append(ChangeResult(
                    "modified", "bom", "high", ea, eb,
                    {"changes": f"color:{ea.color}→{eb.color} mat:{ea.material_spec}→{eb.material_spec}"},
                    "BOM item material or color changed"))

        # ── 关键特性对比 ──────────────────────────
        kc_a = self.ext_a.extract_key_chars()
        kc_b = self.ext_b.extract_key_chars()
        kc_matcher = SmartMatcher(kc_a, kc_b)
        kc_result = kc_matcher.match()

        for e in kc_result["only_in_a"]:
            changes.append(ChangeResult("removed", "key_characteristic", "high",
                                        old_entity=e, reason="Key characteristic removed"))
        for e in kc_result["only_in_b"]:
            changes.append(ChangeResult("added", "key_characteristic", "high",
                                        new_entity=e, reason="Key characteristic added"))

        # ── 图纸信息对比 ──────────────────────────
        # 黑名单：纯统计/派生/元数据字段，不代表图纸工程内容变更
        DRAWING_INFO_BLACKLIST = frozenset({
            # PDF/文件元数据
            "文件名", "创建日期", "Creation Date",
            # 识别器统计指标
            "尺寸标注", "视觉元素总数", "NX倍数展开项", "尺寸视图归属率",
            "其中尺寸", "其中GD&T", "其中角度", "其中表面粗糙度",
            "其中基准特征", "自动编号项", "识别视图数", "使用模型",
            "偏差表条目(AI)",
            # 图纸信息 sheet 里的识别器计数行
            "GDT几何公差",
            # Fix6: 这些是条目计数，真正的增删已在各自类别单独上报，计数行属冗余噪声
            "技术要求", "修订历史", "物料清单", "关键特性", "页数", "页面尺寸(pt)",
        })
        info_a = self.ext_a.extract_info()
        info_b = self.ext_b.extract_info()
        for k in set(info_a) | set(info_b):
            if k in DRAWING_INFO_BLACKLIST:
                continue
            va = info_a.get(k, "")
            vb = info_b.get(k, "")
            va_norm = normalize_text(va)
            vb_norm = normalize_text(vb)
            if va_norm != vb_norm:
                t = "modified" if (va and vb) else ("removed" if va else "added")
                display_key = INFO_KEY_MAP.get(k, k)
                entity = type("InfoItem", (), {"key": k, "value": va or vb})()
                changes.append(ChangeResult(
                    t, "drawing_info", "high",
                    old_entity=entity if va else None,
                    new_entity=entity if vb else None,
                    details={"key": display_key, "old": va, "new": vb},
                    reason=f"Drawing info changed: {display_key}",
                ))

        # ── 构建摘要 ──────────────────────────────
        summary = self._build_summary(changes, renumbering_info, deviation_overrides)
        changes.sort(key=lambda c: (
            {"high": 0, "medium": 1, "low": 2}.get(c.confidence, 3),
            {"removed": 0, "added": 1, "modified": 2}.get(c.change_type, 3),
        ))

        return {
            "summary": summary,
            "changes": [self._change_to_dict(c) for c in changes],
            "renumbering": renumbering_info,
            "deviation_overrides": deviation_overrides,
            "meta": {
                "file_a": str(self.path_a),
                "file_b": str(self.path_b),
                "match_tolerance": MATCH_TOLERANCE,
            },
        }

    def _build_summary(self, changes, renumbering_info,
                       deviation_overrides) -> dict:
        real = {"added": 0, "removed": 0, "modified": 0}
        noise = {"added": 0, "removed": 0, "modified": 0}

        for c in changes:
            if c.confidence == "low":
                noise[c.change_type] = noise.get(c.change_type, 0) + 1
            else:
                real[c.change_type] = real.get(c.change_type, 0) + 1

        real["total"] = sum(real.values())
        noise["total"] = sum(noise.values())

        return {
            "real_changes": real,
            "likely_noise": noise,
            "re_numbered_count": renumbering_info.get("dimensions", {}).get("count", 0),
            "deviation_overrides": deviation_overrides,
        }

    def _change_to_dict(self, cr: ChangeResult) -> dict:
        """将 ChangeResult 序列化为字典"""
        d = {
            "type": cr.change_type,
            "entity_type": cr.entity_type,
            "confidence": cr.confidence,
            "reason": cr.reason,
            "details": cr.details,
        }

        if cr.entity_type == "dimension":
            if cr.old_entity:
                d["old"] = {
                    "dim_no": cr.old_entity.dim_no,
                    "value": cr.old_entity.drawing_raw,
                    "nominal": cr.old_entity.nominal,
                    "lower": cr.old_entity.lower,
                    "upper": cr.old_entity.upper,
                    "view": cr.old_entity.view,
                    "is_ref": cr.old_entity.is_ref,
                }
            if cr.new_entity:
                d["new"] = {
                    "dim_no": cr.new_entity.dim_no,
                    "value": cr.new_entity.drawing_raw,
                    "nominal": cr.new_entity.nominal,
                    "lower": cr.new_entity.lower,
                    "upper": cr.new_entity.upper,
                    "view": cr.new_entity.view,
                    "is_ref": cr.new_entity.is_ref,
                }
        elif cr.entity_type == "gdt":
            if cr.old_entity:
                d["old"] = {
                    "tolerance": cr.old_entity.tolerance_val,
                    "datums": "|".join(sorted(cr.old_entity.datums)),
                    "view": cr.old_entity.view,
                }
            if cr.new_entity:
                d["new"] = {
                    "tolerance": cr.new_entity.tolerance_val,
                    "datums": "|".join(sorted(cr.new_entity.datums)),
                    "view": cr.new_entity.view,
                }
        elif cr.entity_type == "note":
            if cr.old_entity:
                d["old"] = {"number": cr.old_entity.number,
                            "content": cr.old_entity.content_raw}
            if cr.new_entity:
                d["new"] = {"number": cr.new_entity.number,
                            "content": cr.new_entity.content_raw}
        elif cr.entity_type == "bom":
            if cr.old_entity:
                d["old"] = {"part_number": cr.old_entity.part_number,
                            "material": cr.old_entity.material_spec}
            if cr.new_entity:
                d["new"] = {"part_number": cr.new_entity.part_number,
                            "material": cr.new_entity.material_spec}
        elif cr.entity_type == "key_characteristic":
            if cr.old_entity:
                d["old"] = {"type": cr.old_entity.char_type,
                            "spec": cr.old_entity.spec}
            if cr.new_entity:
                d["new"] = {"type": cr.new_entity.char_type,
                            "spec": cr.new_entity.spec}
        elif cr.entity_type == "drawing_info":
            d["info_key"] = cr.details.get("key", "")
            d["old_value"] = cr.details.get("old", "")
            d["new_value"] = cr.details.get("new", "")
        elif cr.entity_type == "deviation_override":
            ent = cr.new_entity or cr.old_entity
            if isinstance(ent, DeviationItem):
                info = {
                    "dim_ref": ent.dim_ref,
                    "original_spec": ent.original_spec,
                    "deviated_spec": ent.deviated_spec,
                }
                if cr.new_entity:
                    d["new"] = info
                else:
                    d["old"] = info

        return d

    def to_csv(self, result: dict) -> str:
        """将比对结果导出为 CSV 字符串"""
        output = io.StringIO()
        writer = csv.writer(output)

        # 摘要区块
        summary = result["summary"]
        writer.writerow(["=== SMART DIFF SUMMARY ==="])
        writer.writerow(["Category", "Added", "Removed", "Modified", "Total"])
        writer.writerow([
            "Real Changes (high/medium confidence)",
            summary["real_changes"]["added"],
            summary["real_changes"]["removed"],
            summary["real_changes"]["modified"],
            summary["real_changes"]["total"],
        ])
        writer.writerow([
            "Likely Noise (low confidence)",
            summary["likely_noise"]["added"],
            summary["likely_noise"]["removed"],
            summary["likely_noise"]["modified"],
            summary["likely_noise"]["total"],
        ])
        writer.writerow([
            "Re-numbered Items", result["renumbering"].get("dimensions", {}).get("count", 0),
            "", "", ""
        ])
        writer.writerow([
            "Deviation Overrides", result.get("deviation_overrides", 0),
            "", "", ""
        ])
        writer.writerow([])

        # 重编号映射
        renum = result.get("renumbering", {}).get("dimensions", {})
        if renum.get("pairs"):
            writer.writerow(["=== RE-NUMBERING MAP ==="])
            if renum.get("is_global_renumber"):
                writer.writerow([f"GLOBAL RE-NUMBER DETECTED ({renum['count']} items)"])
            writer.writerow(["Old Dim No.", "New Dim No.", "Value"])
            for pair in renum["pairs"]:
                writer.writerow([
                    pair["old_dim_no"], pair["new_dim_no"], pair["value"]
                ])
            writer.writerow([])

        # 变更详情
        writer.writerow(["=== CHANGE DETAILS ==="])
        writer.writerow([
            "Type", "Confidence", "Entity Type",
            "Dim No. (Old)", "Dim No. (New)",
            "Value (Old)", "Value (New)",
            "Tolerance (Old)", "Tolerance (New)",
            "View (Old)", "View (New)",
            "Change Detail", "Reason",
        ])

        for c in result["changes"]:
            ct = c["type"]
            conf = c["confidence"]
            etype = c["entity_type"]

            if etype == "dimension":
                old = c.get("old", {}) or {}
                new = c.get("new", {}) or {}
                tol_old = ""
                tol_new = ""
                if old.get("lower") is not None or old.get("upper") is not None:
                    tol_old = f"{old.get('lower','')} / {old.get('upper','')}"
                if new.get("lower") is not None or new.get("upper") is not None:
                    tol_new = f"{new.get('lower','')} / {new.get('upper','')}"

                writer.writerow([
                    ct, conf, etype,
                    old.get("dim_no", ""), new.get("dim_no", ""),
                    old.get("value", ""), new.get("value", ""),
                    tol_old, tol_new,
                    old.get("view", ""), new.get("view", ""),
                    _format_details(c.get("details", {})),
                    c.get("reason", ""),
                ])
            elif etype == "gdt":
                old = c.get("old", {}) or {}
                new = c.get("new", {}) or {}
                writer.writerow([
                    ct, conf, etype, "", "",
                    f"Tol:{old.get('tolerance','')} Datums:{old.get('datums','')}",
                    f"Tol:{new.get('tolerance','')} Datums:{new.get('datums','')}",
                    "", "", old.get("view", ""), new.get("view", ""),
                    _format_details(c.get("details", {})),
                    c.get("reason", ""),
                ])
            elif etype == "drawing_info":
                writer.writerow([
                    ct, conf, etype, "", "",
                    c.get("old_value", ""), c.get("new_value", ""),
                    "", "", "", "",
                    f"{c.get('info_key','')}: {c.get('old_value','')} → {c.get('new_value','')}",
                    c.get("reason", ""),
                ])
            else:
                old = c.get("old", {}) or {}
                new = c.get("new", {}) or {}
                old_str = json.dumps(old, ensure_ascii=False) if old else ""
                new_str = json.dumps(new, ensure_ascii=False) if new else ""
                writer.writerow([
                    ct, conf, etype, "", "",
                    old_str[:80], new_str[:80],
                    "", "", "", "",
                    _format_details(c.get("details", {})),
                    c.get("reason", ""),
                ])

        return output.getvalue()

    def to_json(self, result: dict) -> str:
        """将比对结果导出为 JSON 字符串"""
        return json.dumps(result, ensure_ascii=False, indent=2, default=str)


def _format_details(details: dict) -> str:
    """格式化变更详情为单行字符串"""
    if not details:
        return ""
    parts = []
    for k, v in details.items():
        parts.append(f"{k}:{v}")
    return "; ".join(parts)


# ══════════════════════════════════════════════════════════════════════
# 6. CLI 入口
# ══════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="Smart CAD Diff — 结构化实体匹配 & 噪声过滤"
    )
    ap.add_argument("old_xlsx", help="旧版本识别结果 Excel")
    ap.add_argument("new_xlsx", help="新版本识别结果 Excel")
    ap.add_argument("--csv", help="输出 CSV 报告路径")
    ap.add_argument("--json", help="输出 JSON 报告路径")
    ap.add_argument("--print-summary", action="store_true",
                    help="在控制台打印摘要")
    args = ap.parse_args()

    print(f"[1/3] 加载旧版本: {args.old_xlsx}")
    print(f"[2/3] 加载新版本: {args.new_xlsx}")
    engine = SmartDiffEngine(args.old_xlsx, args.new_xlsx)
    print("[3/3] 执行智能比对...")
    result = engine.compare_all()

    summary = result["summary"]
    real = summary["real_changes"]
    noise = summary["likely_noise"]

    print(f"""
╔═══════════════════════════════════════╗
║        SMART DIFF 结果摘要           ║
╠═══════════════════════════════════════╣
║  真实工程变更 (high/medium)          ║
║    新增: {real['added']:<4d}  删除: {real['removed']:<4d}        ║
║    修改: {real['modified']:<4d}  合计: {real['total']:<4d}        ║
╠═══════════════════════════════════════╣
║  疑似噪声 (low confidence)           ║
║    新增: {noise['added']:<4d}  删除: {noise['removed']:<4d}        ║
║    修改: {noise['modified']:<4d}  合计: {noise['total']:<4d}        ║
╠═══════════════════════════════════════╣
║  重编号项: {summary['re_numbered_count']:<4d}                       ║
║  偏差覆盖: {summary['deviation_overrides']:<4d}                       ║
╚═══════════════════════════════════════╝
""")

    if args.csv:
        csv_str = engine.to_csv(result)
        with open(args.csv, "w", encoding="utf-8-sig", newline="") as f:
            f.write(csv_str)
        print(f"CSV 已保存到: {args.csv}")

    if args.json:
        json_str = engine.to_json(result)
        with open(args.json, "w", encoding="utf-8") as f:
            f.write(json_str)
        print(f"JSON 已保存到: {args.json}")


if __name__ == "__main__":
    main()