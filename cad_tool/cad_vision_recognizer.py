def normalize_bbox(b):
    # 四舍五入到 10 的倍数，减少浮动
    return [round(v / 10) * 10 for v in b]

"""
CAD图纸PDF识别系统 - 第二阶段：视觉识别（高速并发版）
================================================================
"""

import os
import re
import sys
import json
import math
import time
import base64
import hashlib
import argparse
from pathlib import Path
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor, as_completed

import fitz
import numpy as np

from cad_pdf_recognizer import PDFParser, CADInfoExtractor, ExcelExporter


# ═════════════════════════════════════════════
# 1. 分块渲染器（不变）
# ═════════════════════════════════════════════

class TileRenderer:
    """超大幅图纸 → 高清重叠分块（自动跳空白、限块数）"""

    def __init__(self, pdf_path, dpi=150, tile_px=1568, overlap=0.22,
                 max_tiles=40, ink_threshold=0.0015):
        self.pdf_path = pdf_path
        self.dpi = dpi
        self.tile_px = tile_px
        self.overlap = overlap
        self.max_tiles = max_tiles
        self.ink_threshold = ink_threshold
        # 每页渲染几何，供标注阶段复用： {page_no: {"scale","w_px","h_px"}}
        self.page_dims = {}

    def _plan(self, page):
        scale = self.dpi / 72
        step = self.tile_px * (1 - self.overlap)
        px_w = page.rect.width * scale
        px_h = page.rect.height * scale
        cols = max(1, math.ceil((px_w - self.tile_px) / step) + 1)
        rows = max(1, math.ceil((px_h - self.tile_px) / step) + 1)
        return scale, step, cols, rows

    def render(self):
        doc = fitz.open(self.pdf_path)
        tiles = []
        for pno, page in enumerate(doc):
            scale, step, cols, rows = self._plan(page)

            if cols * rows > self.max_tiles:
                shrink = math.sqrt(self.max_tiles / (cols * rows))
                self.dpi = max(72, int(self.dpi * shrink))
                scale, step, cols, rows = self._plan(page)
                print(f"  [自适应] 第{pno+1}页块数超限，DPI降至 {self.dpi} → {cols}x{rows}")

            gray = page.get_pixmap(
                matrix=fitz.Matrix(scale, scale), colorspace=fitz.csGRAY)
            arr = np.frombuffer(gray.samples, dtype=np.uint8).reshape(
                gray.height, gray.width)
            # 记录本页渲染几何（标注阶段按相同 scale 重渲染整页）
            self.page_dims[pno + 1] = {
                "scale": scale, "w_px": gray.width, "h_px": gray.height}

            for ri in range(rows):
                for ci in range(cols):
                    x0 = int(ci * step)
                    y0 = int(ri * step)
                    sub = arr[y0:y0 + self.tile_px, x0:x0 + self.tile_px]
                    if sub.size == 0 or (sub < 200).mean() < self.ink_threshold:
                        continue
                    # 实际块像素尺寸（边缘块会被页面边界裁短）
                    w_px = int(min(self.tile_px, gray.width - x0))
                    h_px = int(min(self.tile_px, gray.height - y0))
                    clip = fitz.Rect(
                        x0 / scale, y0 / scale,
                        (x0 + self.tile_px) / scale,
                        (y0 + self.tile_px) / scale)
                    pm = page.get_pixmap(
                        matrix=fitz.Matrix(scale, scale), clip=clip)
                    # 模型实际看到的图像尺寸（fitz 对超出页面的区域填充白色，
                    # pm.width/pm.height 才是模型收到的图像宽高，是 bbox 归一化的基准）
                    tiles.append({
                        "page": pno + 1,
                        "row": ri, "col": ci,
                        "png": pm.tobytes("png"),
                        # 标注用几何：本块在整页像素坐标中的位置与尺寸
                        "x0_px": x0, "y0_px": y0,
                        "w_px": pm.width,   # 用实际渲染宽（含白边填充）
                        "h_px": pm.height,  # 用实际渲染高（含白边填充）
                        # 真实内容边界（不超过页面），用于裁剪坐标防溢出
                        "content_w_px": w_px,
                        "content_h_px": h_px,
                        "scale": scale,
                    })
        doc.close()
        return tiles


# ═════════════════════════════════════════════
# 2. Vision 分析器（不变）
# ═════════════════════════════════════════════

VISION_PROMPT = """你是汽车工程图纸识别专家。下图是一张【大幅CAD图纸的局部裁剪块】，
图纸被切成多块分别识别，所以本块里看到的视图标题，通常就拥有本块内的尺寸。

【核心任务】识别本块所有工程元素，并尽最大努力判断每个尺寸/公差属于哪个视图或剖面。

【气泡编号识别】很多尺寸旁边有圆圈数字（气泡编号/Dim No.，如 ①②③ 或方框数字 1 2 3），
这是检验报告里的尺寸序号。若尺寸旁能看到圆圈/方框数字，填入 dim_no 字段；看不到则填 ""。

【角度识别】识别所有角度标注（°），包括度数、分、秒。type="angle"，value="30°"或"30°20'"。

【表面粗糙度识别】识别表面粗糙度符号（√/▽）及其标注值，包括 Ra、Rz、Rmax 等参数：
type="surface_roughness"，value="Ra 0.3"或"Rz20 MAX"或"Ra 2μm MIN ALL AROUND"。

【基准特征标签识别】识别单独的基准字母标签（如 J→K 箭头指向、R P Q 多基准），
这些不是 GD&T 控制框而是基准面标识。type="datum_feature"，value="J→K"或"R P Q"。

【GD&T公差→Upper】形位公差的值（如 ⌭0.10 中的 0.10、⊥0.07 中的 0.07）是最大允许偏差，
填入 gdt_tolerance 字段（数字），这个值将自动填入检验报告表的 Upper 列。

【GD&T符号严格区分，非常重要】以下符号容易混淆，必须仔细辨别：
  ⌒ (U+2312) 线轮廓度 —— 弧形开口向下，像一个拱形
  ⌓ (U+2313) 面轮廓度 —— 弧形下方有一横，像拱形下面加了底线
  ⊕ (U+2295) 位置度   —— 圆圈内有十字，圆形带加号
  ⊗ (U+2297) 圆跳动   —— 圆圈内有叉号，圆形带X
  ≡ (U+2261) 对称度   —— 三条平行横线
  ∥ (U+2225) 平行度   —— 两条平行竖线
  ⊥ (U+22A5) 垂直度   —— 上方横线下方垂直线，像倒T
  ⏥ (U+23E5) 平面度   —— 矩形框
  ⏤ (U+23E4) 直线度   —— 一条横线
识别时请仔细观察符号形状，不要把面轮廓度⌓误认为位置度⊕。

【视图归属判断规则，按优先级】
1. 顺着尺寸的引线/箭头，看它指向哪个视图的几何体，归到那个视图。
2. 若引线不清，归到该尺寸簇上方或最近的视图标题（如 DETAIL A、SECTION B-B、TOP VIEW、30W INTERFACE）。
3. 若本块只出现一个视图标题，则本块所有尺寸都归给它。
4. 只有当本块内【完全看不到任何视图标题】时，belongs_to_view 才填 ""。
   不要因为"不确定"就留空——请给出最可能的那个视图。
5. 同时用 view_confidence 标注把握程度：high(引线明确) / medium(就近推断) / low(仅一个候选)。

【严格只输出 JSON】不要解释、不要 markdown 代码块：
{
  "views": ["本块出现的视图/剖面标签，按图中原文"],
  "elements": [
    {
      "type": "dimension | gdt | thread | chamfer | radius | note | datum_label | angle | surface_roughness | datum_feature",
      "value": "标称值，如 21.1 / Φ1.6 / R0.3 / 30° / Ra 0.3 / J→K",
      "tolerance": "公差，如 ±0.1 / +0.021/0，无则空",
      "multiplier": "数量倍数，如 2X / 6X / 12X，无则空",
      "dim_no": "气泡编号(圆圈/方框数字)，看不到填空",
      "is_ref": "是否参考尺寸(带括号或REF标注)，true/false",
      "datums": ["GD&T基准字母，如 A,B,C；非GD&T留空数组"],
      "gdt_tolerance": "形位公差值(仅gdt类型填数字，如0.10)，无则空",
      "surface_param": "粗糙度参数类型(仅surface_roughness填)：Ra/Rz/Rmax",
      "belongs_to_view": "所属视图标签（尽量给出，按上面规则）",
      "view_confidence": "high | medium | low",
      "bbox": "该元素在本图中的边界框，格式 [x0,y0,x1,y1]，用 0~1000 的整数表示相对本块左上角的归一化坐标(x向右,y向下)；务必尽量给出",
      "raw": "原始标注文字"
    }
  ]
}
若本块是标题栏/BOM/纯文本表格，elements 可为空数组，views 注明 ["TITLE_BLOCK"]。

【公差字段填写规则，非常重要】
- tolerance 字段必须包含完整公差表达式：填 "±0.1" 而不是 "0.1"；填 "+0.02/-0.04" 表示非对称
- 单边公差："+0.1/0"（只有上偏差）或 "0/-0.2"（只有下偏差）
- 若无公差（参考尺寸/理论值），tolerance 填 ""，is_ref 填 true
- raw 字段：填图纸上原始标注全文，如 "4X Φ1.60±0.05"，不要简化

【粗糙度上限填写规则】
- value 格式：参数+数值，如 "Ra 0.8" / "Rz 6.3 MAX"
- tolerance 字段同时填粗糙度数值（如 "0.8"），方便系统提取上限
- surface_param 填：Ra / Rz / Rmax（只填参数类型）

【气泡编号识别要点】
- 气泡编号是圆圈或方框内的整数（1~99），紧靠尺寸线/引线端
- 倍数尺寸：dim_no 填第一个气泡号，multiplier 填 "4X"
- 不要把坐标、版本号、材料编号误认为气泡编号"""


class VisionAnalyzer:
    def __init__(self, model="qwen-vl-max", mock=False, max_retries=2):
        self.model = model
        self.mock = mock
        self.max_retries = max_retries
        self.client = None
        if not mock:
            from openai import OpenAI
            self.client = OpenAI(
                api_key=os.environ.get("DASHSCOPE_API_KEY"),
                base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
            )

    def analyze_tile(self, tile):
        if self.mock:
            return self._mock_response(tile)

        b64 = base64.standard_b64encode(tile["png"]).decode()

        for attempt in range(self.max_retries + 1):
            try:
                resp = self.client.chat.completions.create(
                    model=self.model,
                    messages=[{
                        "role": "user",
                        "content": [
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/png;base64,{b64}"
                                },
                            },
                            {"type": "text", "text": VISION_PROMPT},
                        ],
                    }],
                    max_tokens=8192,   # 防止密集块 JSON 被截断（原 4096）
                    temperature=0,     # 贪婪解码，消除随机采样波动
                    seed=42,           # 固定随机种子（部分模型版本支持）
                )
                text = resp.choices[0].message.content or ""
                return self._parse_json(text)

            except Exception as e:
                if attempt < self.max_retries:
                    time.sleep(2 ** attempt)
                else:
                    print(f"    ! 区块({tile['row']},{tile['col']}) 识别失败: {e}")
                    return {"views": [], "elements": []}

    @staticmethod
    def _parse_json(text):
        text = re.sub(r"^```(?:json)?|```$", "", text.strip(),
                      flags=re.MULTILINE).strip()
        m = re.search(r"\{[\s\S]*\}", text)
        if not m:
            return {"views": [], "elements": []}
        try:
            data = json.loads(m.group(0))
            data.setdefault("views", [])
            data.setdefault("elements", [])
            return data
        except json.JSONDecodeError:
            return {"views": [], "elements": []}

    @staticmethod
    def _mock_response(tile):
        r, c = tile["row"], tile["col"]
        if (r + c) % 3 == 0:
            view = f"SECTION {chr(66 + c % 5)}-{chr(66 + c % 5)}"
            elements = [
                {"type": "dimension", "value": f"{20 + c}.1",
                 "tolerance": "±0.1", "multiplier": "2X" if c % 2 else "",
                 "datums": [], "belongs_to_view": view,
                 "view_confidence": "high", "dim_no": str(c + 1),
                 "is_ref": (c % 4 == 0), "bbox": [120, 150, 360, 230],
                 "raw": f"2X {20 + c}.1 ±0.1"},
                {"type": "gdt", "value": "0.1", "tolerance": "",
                 "multiplier": "", "datums": ["E"],
                 "gdt_tolerance": "0.1", "bbox": [400, 300, 640, 380],
                 "belongs_to_view": view, "view_confidence": "medium",
                 "raw": "0.1 E"},
            ]
            if c % 3 == 1:
                elements.append({
                    "type": "angle", "value": "30°",
                    "tolerance": "", "multiplier": "",
                    "dim_no": str(c + 10), "bbox": [200, 500, 380, 560],
                    "belongs_to_view": view,
                    "view_confidence": "high",
                    "raw": "30°",
                })
            if c % 4 == 2:
                elements.append({
                    "type": "surface_roughness", "value": "Ra 0.3",
                    "surface_param": "Ra",
                    "tolerance": "", "multiplier": "",
                    "belongs_to_view": view, "bbox": [600, 600, 820, 670],
                    "view_confidence": "high",
                    "raw": "Ra ≤ 0.3 ALL AROUND",
                })
            if c % 5 == 3:
                elements.append({
                    "type": "datum_feature", "value": "J→K",
                    "tolerance": "", "multiplier": "",
                    "belongs_to_view": view, "bbox": [700, 100, 880, 170],
                    "view_confidence": "medium",
                    "raw": "J arrow K",
                })
            if c == 4:
                elements[0]["multiplier"] = "12X"
                elements[0]["dim_no"] = "5"
                elements[0]["raw"] = "12X 24.1 ±0.1"
            return {"views": [view], "elements": elements}
        return {"views": [], "elements": []}


# ═════════════════════════════════════════════
# 3. 合并去重 + 清洗（不变）
# ═════════════════════════════════════════════

def normalize_tolerance(s):
    if not s:
        return s
    s = s.replace("+/-", "±").replace("＋/－", "±")
    s = re.sub(r"\s*±\s*", "±", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def clean_view_name(v):
    if not v:
        return None
    v = re.sub(r"\s+", " ", v).strip().upper()
    v = re.sub(r"^SEE\s+", "", v)
    if v == "TITLE_BLOCK":
        return v

    v = re.sub(r"(SECTION|DETAIL)\s+([A-Z]{1,2})\s*-\s*([A-Z]{1,2})", r"\1 \2-\3", v)

    if v in ("SECTION", "DETAIL", "VIEW"):
        return None
    if re.search(r"[/\\]", v):
        return None
    if re.fullmatch(r"[\d.]+", v):
        return None
    if re.match(r"^\d{5,}", v):
        return None
    if re.match(r"^REF\.?\b", v):
        return None
    if re.search(r"PBT|GF\d|ABS|PA6|NYLON|RAL", v):
        return None
    if re.search(r"CODE|EXCEPT|SAME AS|±|&", v):
        return None
    if re.fullmatch(r"[A-Z]\d+", v):
        return None

    if re.fullmatch(r"SECTION [A-Z]{1,2}-[A-Z]{1,2}", v):
        return v
    if re.fullmatch(r"DETAIL [A-Z]{1,2}", v):
        return v
    if re.fullmatch(r"\d+W", v):
        return v + " INTERFACE"
    if re.search(r"\d+W INTERFACE|TOP VIEW|FRONT VIEW|BOTTOM VIEW|SIDE VIEW", v):
        return v
    if re.fullmatch(r"[A-Z]{1,2}", v):
        return None
    # 保留 'VIEW A' / 'VIEW J' 等单字母视图命名
    if re.fullmatch(r"VIEW [A-Z]", v):
        return v

    return None


def lookup_angle_general_tolerance(nominal_deg, standard="iso2768_m"):
    """根据角度标称值查询一般公差表，返回 (lower, upper) 或 (None, None)。

    默认使用 ISO 2768-1 中等精度 (m) 的角度一般公差：
      短边长度 ≤10mm:      ±1°
      10 < L ≤ 50mm:      ±0°30'  = ±0.5°
      50 < L ≤ 120mm:     ±0°20'  = ±0.333°
      120 < L ≤ 400mm:    ±0°10'  = ±0.167°
      L > 400mm:          ±0°5'   = ±0.083°

    本函数只按角度度数估算（无短边长度信息时退化到 ±1° 默认值）。
    调用方可提供 standard="custom" 并传自定义表格，或直接在图纸头部标注。
    返回的公差仅在 tolerance 字段为空时作为兜底值使用，并在 raw 里注明来源。
    """
    if nominal_deg is None:
        return None, None
    # 无短边长度信息：对小角度（≤10°）用 ±1°，其余统一用 ±1°（保守）
    # 如果图纸有明确一般公差块，识别器已经会填到 tolerance 字段，不会走到这里
    tol = 1.0
    lower = round(nominal_deg - tol, 4)
    upper = round(nominal_deg + tol, 4)
    return lower, upper


def build_general_tol_lookup(tol_table_rows):
    """将 📏 一般公差表 sheet 的数据构建为可查询的结构。

    tol_table_rows: list of (尺寸范围(mm), 公差(±mm)) 元组，来自 Excel sheet。
    返回:
        linear_table: [(lo, hi, tol), ...] 按尺寸范围，供 lookup_linear_general_tolerance 使用
        angular_tol:  float 或 None，角度公差值
    """
    linear_table = []
    angular_tol = None
    for row in tol_table_rows:
        if not row or len(row) < 2:
            continue
        range_str = str(row[0] or "").strip()
        tol_str   = str(row[1] or "").strip()
        if not tol_str or tol_str in ("", "（未识别）"):
            continue
        try:
            tol_val = float(tol_str)
        except ValueError:
            continue
        # 角度公差单独处理
        if "角度" in range_str or "ANGULAR" in range_str.upper():
            angular_tol = tol_val
            continue
        # 解析 "0 – 20" / "0 - 20" / "20 – 30" 等格式
        m = re.match(r"(\d+(?:\.\d+)?)\s*[–\-]\s*(\d+(?:\.\d+)?)", range_str)
        if m:
            linear_table.append((float(m.group(1)), float(m.group(2)), tol_val))
    # 按下限排序
    linear_table.sort(key=lambda x: x[0])
    return linear_table, angular_tol


def lookup_linear_general_tolerance(nominal, linear_table):
    """根据标称值在一般公差表中查找对应的 ±公差，返回 (lower, upper) 或 (None, None)。

    采用"大于下限且小于等于上限"的区间匹配，与图纸 CHART E1 的 FROM/TO 逻辑一致：
      > 0  TO 20  → 匹配 0 < nominal ≤ 20
      > 20 TO 30  → 匹配 20 < nominal ≤ 30
      ...以此类推

    边界特例：nominal == 0（如 "2X 0.00" 零位/对齐基准标注）按第一档处理，
    即落入下限最小的那一档（通常为 0–20，公差 ±0.15）。
    """
    if nominal is None or not linear_table:
        return None, None
    if nominal == 0:
        tol = linear_table[0][2]
        return round(-tol, 4), round(tol, 4)
    for (lo, hi, tol) in linear_table:
        if lo < nominal <= hi:
            lower = round(nominal - tol, 4)
            upper = round(nominal + tol, 4)
            return lower, upper
    # 超出表格范围：用最后一档（最大范围）兜底
    if linear_table and nominal > 0:
        tol = linear_table[-1][2]
        lower = round(nominal - tol, 4)
        upper = round(nominal + tol, 4)
        return lower, upper
    return None, None



def parse_dimension(value, tolerance="", raw=""):
    nominal = lower = upper = None
    val_str = str(value or raw or "")
    val_clean = re.sub(r"^\d+\s*[Xx×]\s*", "", val_str).strip()

    if "°" in val_clean:
        ang_m = re.search(r"(\d+(?:\.\d+)?)\s*°", val_clean)
        if ang_m:
            nominal = float(ang_m.group(1))
            # 1) 先从 tolerance 字段提取
            tol_str = str(tolerance or "").replace("°", "")
            tol_m = re.search(r"±\s*(\d+(?:\.\d+)?)", tol_str)
            if tol_m:
                t = float(tol_m.group(1))
                lower, upper = nominal - t, nominal + t
                lower, upper = round(lower, 4), round(upper, 4)
            else:
                # 2) 尝试从 raw 字段提取（如 "5.00° ±2°" 或 "70° +2°/-2°"）
                raw_no_deg = str(raw or "").replace("°", "")
                tol_m2 = re.search(r"±\s*(\d+(?:\.\d+)?)", raw_no_deg)
                if tol_m2:
                    t = float(tol_m2.group(1))
                    lower, upper = nominal - t, nominal + t
                    lower, upper = round(lower, 4), round(upper, 4)
                else:
                    # 3) 非对称格式：+x/-y
                    tol_m3 = re.search(
                        r"\+\s*(\d+(?:\.\d+)?)\s*/?\s*[-\u2212]\s*(\d+(?:\.\d+)?)",
                        raw_no_deg)
                    if tol_m3:
                        up, lo = float(tol_m3.group(1)), float(tol_m3.group(2))
                        lower, upper = round(nominal - lo, 4), round(nominal + up, 4)
                    # 4) 公差缺失——lower/upper 保持 None（export 层按需补一般公差）
        return {"nominal": nominal, "lower": lower, "upper": upper}

    nom_m = re.search(r"[ΦØR]?\s*(\d+(?:\.\d+)?)", val_clean)
    if not nom_m:
        return {"nominal": None, "lower": None, "upper": None}
    nominal = float(nom_m.group(1))

    tol = (tolerance or "").strip()
    if not tol and raw:
        tm = re.search(
            r"(±\s*\d*\.?\d+"
            r"|\+\s*\d*\.?\d+\s*/?\s*[-−]\s*\d*\.?\d+"   # +0.3/-0.2 或 +0.3 -0.2
            r"|\+\s*\d*\.?\d+\s*/\s*0"                         # +0.1/0 单边上偏
            r"|0\s*/?\s*[-−]\s*\d*\.?\d+"                       # 0/-0.2 单边下偏
            r"|\+\d*\.?\d+\s+[-−]\s*\d*\.?\d+)",             # +0.3 -0.2 空格分隔
            raw)
        if tm:
            tol = tm.group(1)

    m = re.search(r"±\s*(\d*\.?\d+)", tol)
    if m:
        t = float(m.group(1))
        lower, upper = nominal - t, nominal + t
    else:
        # 非对称：+up /或空格 -lo  （支持 +0.02/-0.35、+0.02 -0.35、+.02/-.35）
        m = re.search(r"\+\s*(\d*\.?\d+)\s*/?\s*[-−]\s*(\d*\.?\d+)", tol)
        if m:
            up, lo = float(m.group(1)), float(m.group(2))
            lower, upper = nominal - lo, nominal + up
        else:
            # 单边：0/-x 或 0 -x  → 上限=nominal
            m = re.search(r"0\s*/?\s*[-−]\s*(\d*\.?\d+)", tol)
            if m:
                lower, upper = nominal - float(m.group(1)), nominal
            else:
                # 单边：+x/0 或 +x 0 → 下限=nominal
                m = re.search(r"\+\s*(\d*\.?\d+)\s*/?\s*0\b", tol)
                if m:
                    lower, upper = nominal, nominal + float(m.group(1))

    if lower is not None:
        lower, upper = round(lower, 4), round(upper, 4)
    return {"nominal": nominal, "lower": lower, "upper": upper}


class VisionMerger:
    def __init__(self):
        self.elements = []
        self.views = set()

    def add(self, tile, result):
        tile_id = f"P{tile['page']}-R{tile['row']}C{tile['col']}"
        tile_views = []
        for v in result.get("views", []):
            cv = clean_view_name(v)
            if cv:
                self.views.add(cv)
                if cv != "TITLE_BLOCK":
                    tile_views.append(cv)
        sole_view = tile_views[0] if len(tile_views) == 1 else None

        # 水印/签名行特征模式：匹配 PDF 页面底部的审核水印文字
        WATERMARK_PATTERNS = (
            re.compile(r"CUS0[12]", re.IGNORECASE),
            re.compile(r"pass\.nor", re.IGNORECASE),
            re.compile(r"\b\d{6,8}\b.*\d{2}:\d{2}"),   # 日期时间戳如 "13:23:36"
            re.compile(r"\bDate\s*:", re.IGNORECASE),
            re.compile(r"\bTime\s*:", re.IGNORECASE),
            re.compile(r"[a-z]{5,8},[0-9]{8},"),             # "xpzrro,12162024,..."
        )

        for el in result.get("elements", []):
            el = dict(el)
            etype = el.get("type", "")
            raw_val = str(el.get("value", "") or "")
            raw_txt = str(el.get("raw", "") or "")

            # ── 过滤水印行（问题B）──────────────────────────────────────
            # 水印文字出现在 value 或 raw 中则跳过
            combined = raw_val + " " + raw_txt
            if any(p.search(combined) for p in WATERMARK_PATTERNS):
                continue

            el["tolerance"] = normalize_tolerance(el.get("tolerance", ""))
            el["raw"] = normalize_tolerance(el.get("raw", ""))
            view = clean_view_name(el.get("belongs_to_view", "")) or ""
            conf = (el.get("view_confidence", "") or "").lower()

            # ── deviation_override 跨块兜底（问题A）────────────────────
            # 若 AI 漏打标签，但 belongs_to_view 明确是 DEVIATION_TABLE，强制修正类型
            if view == "DEVIATION_TABLE" and etype == "dimension":
                etype = "deviation_override"
                el["type"] = "deviation_override"

            if not view and sole_view:
                view = sole_view
                conf = conf or "low"
            el["belongs_to_view"] = view
            el["view_confidence"] = conf if conf in ("high", "medium", "low") else ""
            el["_tile"] = tile_id
            el["_tile_row"] = tile.get("row", 0)
            el["_tile_col"] = tile.get("col", 0)
            el["_page"] = tile.get("page", 1)
            # ── bbox：归一化(0~1000，相对本块) → 整页像素坐标 ──
            bb = el.get("bbox")
            el["_bbox_px"] = None
            if isinstance(bb, (list, tuple)) and len(bb) == 4:
                try:
                    x0n, y0n, x1n, y1n = [float(v) for v in bb]
                    # tw/th：模型实际看到的图像尺寸（含白边）—— bbox 归一化基准
                    tw = tile.get("w_px", 0) or 0
                    th = tile.get("h_px", 0) or 0
                    tx = tile.get("x0_px", 0)
                    ty = tile.get("y0_px", 0)
                    # 页面实际内容边界（不超过页面），用于防止坐标溢出
                    cw = tile.get("content_w_px", tw) or tw
                    ch = tile.get("content_h_px", th) or th
                    if tw and th:
                        X0 = tx + x0n / 1000.0 * tw
                        Y0 = ty + y0n / 1000.0 * th
                        X1 = tx + x1n / 1000.0 * tw
                        Y1 = ty + y1n / 1000.0 * th
                        # 限制在本块实际内容范围内，避免落在页面外白边区域
                        X0 = max(float(tx), min(X0, tx + cw))
                        X1 = max(float(tx), min(X1, tx + cw))
                        Y0 = max(float(ty), min(Y0, ty + ch))
                        Y1 = max(float(ty), min(Y1, ty + ch))
                        bw, bh = abs(X1 - X0), abs(Y1 - Y0)
                        # ── 有效性校验 ──────────────────────────────────
                        # 1) 全0/退化为点：模型未给出有效坐标，丢弃 bbox（不画框，但保留数据）
                        if bw < 0.5 and bh < 0.5:
                            el["_bbox_px"] = None
                        # 2) 异常巨大：单边超过本块宽/高的 60%，视为模型给出的
                        #    整块/整表范围而非单个标注，丢弃 bbox 避免出现跨多列的巨型框
                        elif bw > tw * 0.6 or bh > th * 0.6:
                            el["_bbox_px"] = None
                        else:
                            el["_bbox_px"] = (round(min(X0, X1), 1), round(min(Y0, Y1), 1),
                                              round(max(X0, X1), 1), round(max(Y0, Y1), 1))
                except (ValueError, TypeError):
                    el["_bbox_px"] = None

            # ── note 类型噪声过滤（问题：整张表格逐行被误标为 note）──────
            # note 的 raw 文本过短（如纯分隔线/空单元格）或为纯数字/标点
            # 的情况大量出现时，往往是表格网格线被误识别，直接丢弃该元素
            if etype == "note":
                note_raw = (el.get("raw", "") or el.get("value", "") or "").strip()
                if len(note_raw) < 4 or re.fullmatch(r"[\d\s,.\-_/|]*", note_raw):
                    continue

            el["dim_no"] = str(el.get("dim_no", "") or "").strip()
            el["is_ref"] = bool(el.get("is_ref", False)) or \
                ("REF" in el["raw"].upper() or el["raw"].strip().startswith("("))
            el["gdt_tolerance"] = el.get("gdt_tolerance", "")
            el["surface_param"] = el.get("surface_param", "")
            if etype == "dimension":
                calc = parse_dimension(el.get("value", ""),
                                       el.get("tolerance", ""), el.get("raw", ""))
                el["lower"] = calc["lower"]
                el["upper"] = calc["upper"]
            elif etype == "angle":
                # 尝试从 tolerance/raw 解析角度公差；解析不出时保持 None
                # export 层可调用 lookup_angle_general_tolerance 补一般公差
                calc = parse_dimension(el.get("value", ""),
                                       el.get("tolerance", ""), el.get("raw", ""))
                el["lower"] = calc["lower"]
                el["upper"] = calc["upper"]
            elif etype == "deviation_override":
                el["belongs_to_view"] = "DEVIATION_TABLE"
                # Fix7: 解析偏差后规格的非对称公差，而非一律置 None
                calc = parse_dimension(el.get("value", ""),
                                       el.get("tolerance", ""), el.get("raw", ""))
                el["lower"] = calc["lower"]
                el["upper"] = calc["upper"]
            el["_elem_idx"] = len(self.elements)   # 全局插入序号，用于稳定排序
            self.elements.append(el)

    def dedup(self):
        # 先按空间位置排序，保证去重结果不依赖线程完成顺序
        def _sort_key(el):
            tile = el.get("_tile", "")
            pm = re.search(r"P(\d+)", tile)
            page = int(pm.group(1)) if pm else 0
            return (page, el.get("_tile_row", 0), el.get("_tile_col", 0),
                    el.get("_elem_idx", 0))   # 块内序号，确保完全确定性排序
        self.elements.sort(key=_sort_key)

        seen = {}
        for el in self.elements:
            # 去重键：只用内容，不含 belongs_to_view
            # 同一标注在两个重叠块里视图归属可能略不同，不应产生两条
            raw_key = re.sub(r"\s+", "",
                             (el.get("raw", "") or el.get("value", "")).upper())
            # 同一标注在不同视图出现时不应去重（如同尺寸出现在TOP VIEW和SECTION A-A）
            view_key = el.get("belongs_to_view", "") or ""
            key = (el.get("type", ""), raw_key, view_key)
            if key not in seen:
                seen[key] = el
            else:
                existing = seen[key]
                # 优先保留信息更完整的那条
                score_new = (bool(el.get("_bbox_px")),
                             bool(el.get("dim_no")),
                             bool(el.get("belongs_to_view")))
                score_old = (bool(existing.get("_bbox_px")),
                             bool(existing.get("dim_no")),
                             bool(existing.get("belongs_to_view")))
                if score_new > score_old:
                    seen[key] = el
                elif score_new == score_old:
                    # 分数相同时，保留空间位置更靠前的（_elem_idx更小）
                    if el.get("_elem_idx", 0) < existing.get("_elem_idx", 0):
                        seen[key] = el
        self.elements = list(seen.values())

        # ── 跨类型 bbox 重叠去重 ──────────────────────────────────────
        # 同一标注被模型重复输出为不同 type（如 dimension/gdt/datum_feature
        # 各给一份几乎重合的 bbox），按 (页, 中心点距离<20px, raw内容高度相似)
        # 只保留一个：优先级 dimension > gdt > angle > datum_feature/label
        #            > surface_roughness > deviation_override > note > 其他
        TYPE_PRIORITY = {
            "dimension": 0, "gdt": 1, "angle": 2,
            "datum_feature": 3, "datum_label": 3,
            "surface_roughness": 4, "deviation_override": 5,
            "thread": 6, "chamfer": 6, "radius": 6, "note": 9,
        }

        def _center(el):
            bb = el.get("_bbox_px")
            if not bb:
                return None
            return ((bb[0] + bb[2]) / 2, (bb[1] + bb[3]) / 2)

        def _norm_raw(el):
            return re.sub(r"\s+", "", (el.get("raw", "") or el.get("value", "") or "").upper())

        boxed = [el for el in self.elements if el.get("_bbox_px")]
        unboxed = [el for el in self.elements if not el.get("_bbox_px")]
        boxed.sort(key=_sort_key)

        drop_ids = set()
        for i, a in enumerate(boxed):
            if id(a) in drop_ids:
                continue
            ca = _center(a)
            ra = _norm_raw(a)
            for b in boxed[i + 1:]:
                if id(b) in drop_ids:
                    continue
                if a.get("_page") != b.get("_page"):
                    continue
                if a.get("type") == b.get("type"):
                    continue  # 同类型已在上面按 raw_key 去重过
                cb = _center(b)
                rb = _norm_raw(b)
                if not ca or not cb:
                    continue
                dist = ((ca[0] - cb[0]) ** 2 + (ca[1] - cb[1]) ** 2) ** 0.5
                if dist >= 20:
                    continue
                # 中心点很近：内容完全相同，或一个是另一个的前缀/子串（信息量较少的一份）
                same_content = (ra == rb) or (ra and rb and (ra in rb or rb in ra))
                if not same_content:
                    continue
                pa = TYPE_PRIORITY.get(a.get("type", ""), 5)
                pb = TYPE_PRIORITY.get(b.get("type", ""), 5)
                if pa <= pb:
                    drop_ids.add(id(b))
                else:
                    drop_ids.add(id(a))
                    break

        self.elements = unboxed + [el for el in boxed if id(el) not in drop_ids]

        # 最终确定性排序：保证无论线程完成顺序如何，输出完全相同
        self.elements.sort(key=_sort_key)
        return self

    def by_type(self, t):
        return [e for e in self.elements if e.get("type") == t]


# ═════════════════════════════════════════════
# 3b. 元素框选标注（在整页图上把每个识别元素框出来）
# ═════════════════════════════════════════════

# 类型 → 颜色（RGB）
ELEMENT_COLORS = {
    "dimension": (220, 38, 38),         # 红
    "gdt": (37, 99, 235),               # 蓝
    "angle": (22, 163, 74),             # 绿
    "surface_roughness": (234, 88, 12), # 橙
    "datum_feature": (147, 51, 234),    # 紫
    "datum_label": (147, 51, 234),
    "thread": (202, 138, 4),            # 暗黄
    "chamfer": (8, 145, 178),           # 青
    "radius": (190, 24, 93),            # 玫红
    "note": (100, 116, 139),            # 灰
    "deviation_override": (192, 38, 211),  # 品红
}
DEFAULT_COLOR = (15, 23, 42)


class Annotator:
    """把识别到的每个元素在整页图纸上框出来，按类型上色并标注编号/类型。
    输出每页一张 PNG。需要元素带 _page 与 _bbox_px（整页像素坐标）。
    """

    def __init__(self, pdf_path, page_dims):
        self.pdf_path = pdf_path
        self.page_dims = page_dims or {}

    def annotate(self, elements, output_dir, stem):
        from PIL import Image, ImageDraw, ImageFont
        try:
            font = ImageFont.truetype("DejaVuSans.ttf", 14)
            font_sm = ImageFont.truetype("DejaVuSans.ttf", 11)
        except Exception:
            font = ImageFont.load_default()
            font_sm = font

        # 按页分组（仅保留有 bbox 的元素）
        by_page = {}
        for el in elements:
            if not el.get("_bbox_px"):
                continue
            by_page.setdefault(el.get("_page", 1), []).append(el)

        doc = fitz.open(self.pdf_path)
        outputs = []
        # 确保每页都输出，即使该页元素全无 bbox（便于用户确认问题）
        for pno, page in enumerate(doc, start=1):
            geo = self.page_dims.get(pno)
            if geo is None:
                # 该页在渲染阶段被跳过（空白页），仍然生成基本标注图
                scale = 1.0
            else:
                scale = geo["scale"]
            pm = page.get_pixmap(matrix=fitz.Matrix(scale, scale))
            img = Image.frombytes("RGB", (pm.width, pm.height), pm.samples)
            # 半透明叠加层，避免边框遮挡图纸
            overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
            draw = ImageDraw.Draw(overlay, "RGBA")

            # 整页像素边界
            page_w, page_h = img.size
            els = by_page.get(pno, [])
            counts = {}
            for el in els:
                x0, y0, x1, y1 = el["_bbox_px"]
                # 裁剪到页面范围内
                x0 = max(0.0, min(x0, page_w - 1))
                y0 = max(0.0, min(y0, page_h - 1))
                x1 = max(0.0, min(x1, page_w - 1))
                y1 = max(0.0, min(y1, page_h - 1))
                # 保证最小框尺寸（避免退化为点/线）
                MIN_SIZE = 8
                if x1 - x0 < MIN_SIZE:
                    cx = (x0 + x1) / 2
                    x0, x1 = cx - MIN_SIZE / 2, cx + MIN_SIZE / 2
                if y1 - y0 < MIN_SIZE:
                    cy = (y0 + y1) / 2
                    y0, y1 = cy - MIN_SIZE / 2, cy + MIN_SIZE / 2
                etype = el.get("type", "")
                color = ELEMENT_COLORS.get(etype, DEFAULT_COLOR)
                counts[etype] = counts.get(etype, 0) + 1
                # 框（线宽随图纸分辨率自适应，最小2px）
                lw = max(2, int(page_w / 3000))
                draw.rectangle([x0, y0, x1, y1], outline=color + (255,), width=lw)
                # 标签：气泡编号优先，否则类型缩写
                label = str(el.get("dim_no") or "").strip()
                if not label:
                    label = {"dimension": "D", "gdt": "G", "angle": "∠",
                             "surface_roughness": "Ra", "datum_feature": "DF",
                             "deviation_override": "Δ", "note": "N"}.get(etype, "·")
                # 标签底色块
                tb = draw.textbbox((0, 0), label, font=font_sm)
                tw, thh = tb[2] - tb[0], tb[3] - tb[1]
                lx, ly = x0, max(0, y0 - thh - 4)
                draw.rectangle([lx, ly, lx + tw + 6, ly + thh + 4],
                               fill=color + (235,))
                draw.text((lx + 3, ly + 2), label, fill=(255, 255, 255, 255),
                          font=font_sm)

            # 图例
            self._draw_legend(draw, counts, font, img.size)

            out_img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
            suffix = f"_标注_P{pno}.png" if doc.page_count > 1 else "_标注.png"
            out_path = str(Path(output_dir) / f"{stem}{suffix}")
            out_img.save(out_path)
            outputs.append(out_path)
        doc.close()
        return outputs

    @staticmethod
    def _draw_legend(draw, counts, font, size):
        if not counts:
            return
        items = [(t, n) for t, n in counts.items()]
        pad = 10
        line_h = 22
        box_w = 240
        box_h = pad * 2 + line_h * (len(items) + 1)
        x0 = 10
        y0 = 10
        draw.rectangle([x0, y0, x0 + box_w, y0 + box_h],
                       fill=(255, 255, 255, 230), outline=(0, 0, 0, 255))
        draw.text((x0 + pad, y0 + pad), "Detected elements / count",
                  fill=(0, 0, 0, 255), font=font)
        yy = y0 + pad + line_h
        for t, n in items:
            color = ELEMENT_COLORS.get(t, DEFAULT_COLOR)
            draw.rectangle([x0 + pad, yy + 3, x0 + pad + 14, yy + 17],
                           fill=color + (255,), outline=(0, 0, 0, 255))
            draw.text((x0 + pad + 22, yy + 2), f"{t}: {n}", fill=(0, 0, 0, 255),
                      font=font)
            yy += line_h


# ═════════════════════════════════════════════
# 4. 增强 Excel 导出（不变）
# ═════════════════════════════════════════════

class VisionExcelExporter(ExcelExporter):

    def __init__(self, extractor, merger, output_path, model_name):
        super().__init__(extractor, output_path)
        self.merger = merger
        self.model_name = model_name

    def _parse_multiplier(self, s):
        if not s:
            return 1
        m = re.search(r"(\d+)\s*[Xx×]", str(s))
        return int(m.group(1)) if m else 1

    def export(self):
        super().export()
        import openpyxl
        from openpyxl.styles import Font
        self.wb = openpyxl.load_workbook(self.output_path)
        m = self.merger

        dims = m.by_type("dimension")
        gdts = m.by_type("gdt")
        angles = m.by_type("angle")
        roughness = m.by_type("surface_roughness")
        datum_features = m.by_type("datum_feature")

        # ── 读取图纸自带一般公差表，构建查询结构 ──────────────────────
        # 来源：cad_pdf_recognizer 已将 "GENERAL TOLERANCE" 块写入 📏 一般公差表 sheet
        _linear_tol_table = []
        _angular_tol_from_chart = None
        if "📏 一般公差表" in self.wb.sheetnames:
            tol_ws = self.wb["📏 一般公差表"]
            tol_rows = []
            for row in tol_ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] and str(row[0]).strip() not in ("", "（未识别）", "尺寸范围(mm)"):
                    tol_rows.append(row)
            _linear_tol_table, _angular_tol_from_chart = build_general_tol_lookup(tol_rows)

        # ── 为 lower/upper 缺失的 dimension 补一般公差 ─────────────────
        # 包括 Ref 参考尺寸：图纸上没有单独标公差的元素，按一般公差表统一补全，
        # 以便在 PPAP 检验报告/导出表中也能体现验收限（不再因"无验收限"被丢弃）。
        _fallback_count = 0
        _fallback_ref_count = 0
        for d in dims:
            if d.get("lower") is not None and d.get("upper") is not None:
                continue   # 已有公差，无需补
            # 提取标称值（数字部分）
            val_str = str(d.get("value", "") or d.get("raw", "") or "")
            val_str_clean = re.sub(r"^\d+\s*[Xx×]\s*", "", val_str).strip()
            num_m = re.search(r"[ΦØ⌀R]?\s*(\d+(?:\.\d+)?)", val_str_clean)
            if not num_m:
                continue
            nominal = float(num_m.group(1))
            lo, up = lookup_linear_general_tolerance(nominal, _linear_tol_table)
            if lo is not None:
                d["lower"] = lo
                d["upper"] = up
                # 在 raw 字段末尾注明来源，便于人工核查
                if d.get("is_ref"):
                    d["raw"] = (d.get("raw") or "") + "  [一般公差·Ref]"
                    _fallback_ref_count += 1
                else:
                    d["raw"] = (d.get("raw") or "") + "  [一般公差]"
                    _fallback_count += 1

        # 角度行：如果图纸里有明确的角度一般公差，优先用它覆盖默认的 ISO ±1°
        if _angular_tol_from_chart is not None:
            for a in angles:
                if a.get("lower") is not None:
                    continue
                calc = parse_dimension(a.get("value", ""),
                                       a.get("tolerance", ""), a.get("raw", ""))
                if calc["lower"] is None:
                    nom = calc.get("nominal")
                    if nom is not None:
                        a["lower"] = round(nom - _angular_tol_from_chart, 4)
                        a["upper"] = round(nom + _angular_tol_from_chart, 4)
                        a["raw"] = (a.get("raw") or "") + "  [一般公差±角度]"
                    else:
                        a["lower"] = calc["lower"]
                        a["upper"] = calc["upper"]
                else:
                    a["lower"] = calc["lower"]
                    a["upper"] = calc["upper"]

        if _fallback_count or _fallback_ref_count:
            print(f"      [一般公差兜底] 共补全 {_fallback_count} 个尺寸 + "
                  f"{_fallback_ref_count} 个Ref参考尺寸 的 lower/upper")


        # 按空间位置排序后再编号，确保编号分配完全确定
        dims.sort(key=lambda d: (d.get("_page", 0), d.get("_tile_row", 0),
                                  d.get("_tile_col", 0), d.get("_elem_idx", 0)))
        existing_nums = []
        for d in dims:
            dn = d.get("dim_no", "")
            try:
                existing_nums.append(int(dn))
            except (ValueError, TypeError):
                pass
        next_num = max(existing_nums) + 1 if existing_nums else 1
        for d in dims:
            if not d.get("dim_no", "").strip() and not d.get("is_ref"):
                d["dim_no"] = f"A-{next_num}"   # A- 前缀标识自动补号，区别于图纸气泡号
                d["_auto_numbered"] = True
                next_num += 1

        ws = self.wb.create_sheet("🎯 尺寸-视图关联(AI)")
        self._table(
            ws,
            ["#", "Dim No.", "Ref", "所属视图", "归属置信", "数量倍数",
             "标称值", "公差", "下限", "上限", "原始标注", "来源块"],
            [(i+1, d.get("dim_no",""), "✓" if d.get("is_ref") else "",
              d.get("belongs_to_view",""), d.get("view_confidence",""),
              d.get("multiplier",""), d.get("value",""), d.get("tolerance",""),
              d.get("lower",""), d.get("upper",""),
              d.get("raw",""), d.get("_tile",""))
             for i, d in enumerate(dims)],
            [6, 9, 6, 20, 9, 9, 12, 14, 10, 10, 26, 12],
            empty_msg="（Vision未返回尺寸数据）")

        ws = self.wb.create_sheet("⊕ GD&T关联(AI)")
        self._table(
            ws,
            ["#", "所属视图", "公差值", "基准", "GD&T公差(Max)", "原始", "来源块"],
            [(i+1, g.get("belongs_to_view",""), g.get("value",""),
              ", ".join(g.get("datums",[]) or []),
              g.get("gdt_tolerance",""),
              g.get("raw",""), g.get("_tile",""))
             for i, g in enumerate(gdts)],
            [6, 22, 12, 14, 14, 24, 14],
            empty_msg="（Vision未返回GD&T）")

        ws = self.wb.create_sheet("📋 检验报告表")
        report_rows = []

        for d in dims:
            dim_no = d.get("dim_no", "")
            multiplier = self._parse_multiplier(d.get("multiplier", ""))
            lower, upper = d.get("lower"), d.get("upper")
            expand = max(1, multiplier)

            for inst in range(expand):
                # 展开行：第一行用原始 dim_no，后续行用 "X-1","X-2" 标识子序号
                if expand == 1:
                    instance_no = dim_no
                elif inst == 0:
                    instance_no = dim_no
                else:
                    instance_no = f"{dim_no}-{inst+1}" if dim_no else None

                if instance_no is not None or (lower is not None):
                    report_rows.append((
                        instance_no,
                        "Ref" if d.get("is_ref") else "",
                        d.get("value", ""),
                        lower if lower is not None else "",
                        upper if upper is not None else "",
                        d.get("belongs_to_view", ""),
                        d.get("raw", ""),
                    ))

        # GD&T 行：写入 "SEE GD&T"，Lower/Upper 均填 "OK"（v3格式）
        for g in gdts:
            gt = g.get("gdt_tolerance", "")
            if gt:
                report_rows.append((
                    g.get("dim_no", ""),
                    "",
                    "SEE GD&T",
                    "OK",
                    "OK",
                    g.get("belongs_to_view", ""),
                    g.get("raw", ""),
                ))

        # 角度行：有公差时写入检验报告，无公差时跳过（export 层已在 cad_ppap_export 里处理）
        for a in angles:
            lower, upper = a.get("lower"), a.get("upper")
            if lower is None and upper is None:
                # 公差缺失时尝试查一般公差表作为兜底
                calc = parse_dimension(a.get("value", ""),
                                       a.get("tolerance", ""), a.get("raw", ""))
                lower, upper = calc["lower"], calc["upper"]
            if lower is not None or upper is not None:
                report_rows.append((
                    a.get("dim_no", ""),
                    "",
                    a.get("value", ""),
                    lower if lower is not None else "",
                    upper if upper is not None else "",
                    a.get("belongs_to_view", ""),
                    a.get("raw", ""),
                ))

        for s in roughness:
            val = s.get("value", "")
            raw = s.get("raw", "")
            upper_val = ""
            combined_rs = str(val) + " " + str(raw)
            # 优先: MAX 标注
            max_m = re.search(r"(\d+(?:\.\d+)?)\s*MAX", combined_rs, re.IGNORECASE)
            if max_m:
                upper_val = max_m.group(1)
            else:
                # Ra / Rz / Rmax 后跟数值 → 该值即为上限
                ra_m = re.search(r"(?i)\b(?:Ra|Rz|Rmax)\s*[≤<=]?\s*(\d+(?:\.\d+)?)", combined_rs)
                if ra_m:
                    upper_val = ra_m.group(1)
                else:
                    # 纯数值兜底
                    nums_r = re.findall(r"\d+(?:\.\d+)?", combined_rs)
                    nums_r = [float(x) for x in nums_r if float(x) <= 50]
                    if nums_r:
                        upper_val = str(max(nums_r))
            report_rows.append((
                s.get("dim_no", ""),
                "",
                f"粗糙度: {val}",
                "",
                upper_val,
                s.get("belongs_to_view", ""),
                raw,
            ))

        def sort_key(r):
            try:
                base = re.split(r"[-_]", str(r[0]))[0]
                return (0, int(base), str(r[2]))
            except (ValueError, TypeError):
                return (1, 0, str(r[2]))
        report_rows.sort(key=sort_key)
        self._table(
            ws,
            ["Dim No.", "Ref Only", "Drawing Dimension", "Lower", "Upper", "所属视图", "原始标注"],
            report_rows,
            [10, 10, 18, 12, 12, 20, 26],
            empty_msg="（无可生成检验项的尺寸）")

        ws = self.wb.create_sheet("📐 角度标注(AI)")
        self._table(
            ws,
            ["#", "Dim No.", "所属视图", "角度值", "原始标注", "来源块"],
            [(i+1, a.get("dim_no",""), a.get("belongs_to_view",""),
              a.get("value",""), a.get("raw",""), a.get("_tile",""))
             for i, a in enumerate(angles)],
            [6, 10, 20, 14, 26, 12],
            empty_msg="（未识别到角度标注）")

        ws = self.wb.create_sheet("🔲 表面粗糙度(AI)")
        self._table(
            ws,
            ["#", "Dim No.", "所属视图", "参数类型", "粗糙度值", "原始标注", "来源块"],
            [(i+1, s.get("dim_no",""), s.get("belongs_to_view",""),
              s.get("surface_param",""), s.get("value",""),
              s.get("raw",""), s.get("_tile",""))
             for i, s in enumerate(roughness)],
            [6, 10, 20, 12, 20, 26, 12],
            empty_msg="（未识别到表面粗糙度）")

        ws = self.wb.create_sheet("🔤 基准特征(AI)")
        self._table(
            ws,
            ["#", "所属视图", "基准标签", "原始标注", "来源块"],
            [(i+1, df.get("belongs_to_view",""), df.get("value",""),
              df.get("raw",""), df.get("_tile",""))
             for i, df in enumerate(datum_features)],
            [6, 22, 20, 26, 12],
            empty_msg="（未识别到基准特征标签）")

        ws = self.wb.create_sheet("🗺 视图清单(AI)")
        self._table(
            ws, ["#", "视图/剖面标签"],
            [(i+1, v) for i, v in enumerate(sorted(m.views))],
            [6, 40], empty_msg="（未识别到视图）")

        ws = self.wb.create_sheet("🔍 全部视觉元素(AI)")
        self._table(
            ws,
            ["#", "类型", "标称值", "公差", "倍数", "基准", "所属视图", "归属置信", "原始", "来源块"],
            [(i+1, e.get("type",""), e.get("value",""), e.get("tolerance",""),
              e.get("multiplier",""), ", ".join(e.get("datums",[]) or []),
              e.get("belongs_to_view",""), e.get("view_confidence",""),
              e.get("raw",""), e.get("_tile",""))
             for i, e in enumerate(m.elements)],
            [6, 12, 14, 14, 8, 12, 22, 10, 26, 12])

        ov = self.wb["📋 图纸信息"]
        last = ov.max_row + 2
        ov.cell(row=last, column=1,
                value="【第二阶段 Vision 识别（百炼）】").font = Font(
            bold=True, color=self.C_SUB, name="Arial", size=11)
        with_view = sum(1 for d in dims if d.get("belongs_to_view"))
        rate = f"{100*with_view//len(dims)}%" if dims else "—"
        auto_num = sum(1 for d in dims if d.get("_auto_numbered"))
        expanded = sum(1 for d in dims if self._parse_multiplier(d.get("multiplier","")) > 1)
        stats = [
            ("使用模型", self.model_name),
            ("识别视图数", len(m.views)),
            ("视觉元素总数", len(m.elements)),
            ("其中尺寸", len(dims)),
            ("其中GD&T", len(gdts)),
            ("其中角度", len(angles)),
            ("其中表面粗糙度", len(roughness)),
            ("其中基准特征", len(datum_features)),
            ("NX倍数展开项", expanded),
            ("自动编号项", auto_num),
            ("尺寸视图归属率", f"{with_view}/{len(dims)} ({rate})"),
        ]
        for i, (k, v) in enumerate(stats, 1):
            ov.cell(row=last+i, column=1, value=k).font = Font(
                bold=True, name="Arial", size=9)
            ov.cell(row=last+i, column=2, value=str(v)).font = Font(
                name="Arial", size=9)

        self.wb.save(self.output_path)
        return self.output_path


# ═════════════════════════════════════════════
# 5. 主流程（高速并发版）
# ═════════════════════════════════════════════

def process_with_vision(pdf_path, output_dir=".", model="qwen-vl-max",
                        dpi=150, max_tiles=40, mock=False, annotate=True,
                        use_cache=True):
    pdf_path = str(pdf_path)
    stem = Path(pdf_path).stem
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    out = str(Path(output_dir) / f"{stem}_识别结果_V2.xlsx")

    print("[1/5] 第一阶段文本提取...")
    parser = PDFParser(pdf_path).parse()
    extractor = CADInfoExtractor(parser).extract_all()

    print(f"[2/5] 分块高清渲染 (DPI={dpi}, 上限{max_tiles}块)...")
    renderer = TileRenderer(pdf_path, dpi=dpi, max_tiles=max_tiles)
    tiles = renderer.render()
    if len(tiles) > max_tiles:
        print(f"      有效块 {len(tiles)} 超上限，截断至 {max_tiles}")
        tiles = tiles[:max_tiles]
    print(f"      → 有效区块 {len(tiles)} 个" + ("  [MOCK模式]" if mock else ""))

    print(f"[3/5] Vision 并发识别 (model={model}, workers=6)...")
    analyzer = VisionAnalyzer(model=model, mock=mock)
    merger = VisionMerger()

    # ── 磁盘缓存：同一PDF的同一块不重复调API ──
    cache_path = Path(output_dir) / f"{stem}_tile_cache.json"
    tile_cache = {}
    if use_cache and cache_path.exists():
        try:
            tile_cache = json.loads(cache_path.read_text(encoding="utf-8"))
            print(f"      已加载缓存 {len(tile_cache)} 条")
        except Exception:
            tile_cache = {}

    def _tile_hash(tile):
        """基于图像内容生成稳定的缓存键"""
        h = hashlib.md5(tile["png"]).hexdigest()[:16]
        return f"P{tile['page']}-R{tile['row']}C{tile['col']}-{h}"

    def _worker(idx, tile):
        """单个线程任务：识别一个区块（有缓存则跳过API）"""
        cache_key = _tile_hash(tile)
        if cache_key in tile_cache:
            return idx, tile, tile_cache[cache_key], True
        result = analyzer.analyze_tile(tile)
        return idx, tile, result, False

    # ═══ 并发识别，但结果按原始顺序收集 ═══
    completed = 0
    total = len(tiles)
    results_bag = [None] * total   # 按索引占位，保证顺序

    with ThreadPoolExecutor(max_workers=6) as executor:
        future_to_idx = {
            executor.submit(_worker, i, tile): i
            for i, tile in enumerate(tiles)
        }
        for future in as_completed(future_to_idx):
            idx, tile, result, cached = future.result()
            results_bag[idx] = (tile, result)
            if not cached:
                tile_cache[_tile_hash(tile)] = result
            completed += 1
            tag = " (cached)" if cached else ""
            if completed % 5 == 0 or completed == total:
                print(f"      进度 {completed}/{total}{tag}，累计元素估算中")

    # 按原始空间顺序依次合并（消除线程完成顺序的随机性）
    for tile, result in results_bag:
        merger.add(tile, result)

    # 持久化缓存（下次运行同一PDF直接复用，结果100%一致）
    if use_cache:
        try:
            cache_path.write_text(
                json.dumps(tile_cache, ensure_ascii=False), encoding="utf-8")
        except Exception as e:
            print(f"      ⚠ 缓存写入失败: {e}")

    merger.dedup()
    print(f"      → 去重后 {len(merger.elements)} 个元素，{len(merger.views)} 个视图")

    print("[4/5] 融合文本+视觉结果...")
    print("[5/5] 生成增强 Excel...")
    VisionExcelExporter(extractor, merger, out, model).export()
    print(f"完成 → {out}")

    annotated = []
    if annotate:
        print("[+] 生成元素框选标注图...")
        try:
            annotated = Annotator(pdf_path, renderer.page_dims).annotate(
                merger.elements, output_dir, stem)
            n_box = sum(1 for e in merger.elements if e.get("_bbox_px"))
            print(f"      → 标注 {n_box}/{len(merger.elements)} 个带坐标的元素，"
                  f"输出 {len(annotated)} 张图")
            for p in annotated:
                print(f"        {p}")
            if n_box == 0:
                print("      ⚠ 模型未返回任何 bbox 坐标，标注图为空框。"
                      "请确认模型支持视觉定位(grounding)。")
        except Exception as e:
            print(f"      ! 标注生成失败: {e}")

    return {
        "output": out,
        "tiles": len(tiles),
        "elements": len(merger.elements),
        "views": sorted(merger.views),
        "annotated": annotated,
    }


def main():
    ap = argparse.ArgumentParser(description="CAD图纸 Vision 识别（百炼版·高速并发）")
    ap.add_argument("pdf", help="输入PDF路径")
    ap.add_argument("output_dir", nargs="?", default=".", help="输出目录")
    ap.add_argument("--model", default="qwen-vl-max",
                    help="模型 (默认 qwen-vl-max)")
    ap.add_argument("--dpi", type=int, default=150, help="渲染DPI (默认150)")
    ap.add_argument("--max-tiles", type=int, default=40,
                    help="区块数上限 (默认40)")
    ap.add_argument("--workers", type=int, default=6,
                    help="并发线程数 (默认6)")
    ap.add_argument("--mock", action="store_true",
                    help="模拟模式，无需密钥，验证流程")
    ap.add_argument("--no-annotate", action="store_true",
                    help="不生成元素框选标注图")
    ap.add_argument("--no-cache", action="store_true",
                    help="禁用区块缓存（强制重新调用API）")
    args = ap.parse_args()

    if not args.mock and not os.environ.get("DASHSCOPE_API_KEY"):
        print("⚠  未检测到 DASHSCOPE_API_KEY 环境变量。")
        print("   设置方式：")
        print("   Windows:   set DASHSCOPE_API_KEY=sk-你的密钥")
        print("   Mac/Linux: export DASHSCOPE_API_KEY=sk-你的密钥")
        print("   或加 --mock 先验证流程。")
        sys.exit(1)

    result = process_with_vision(
        args.pdf, args.output_dir,
        model=args.model, dpi=args.dpi,
        max_tiles=args.max_tiles, mock=args.mock,
        annotate=not args.no_annotate,
        use_cache=not args.no_cache,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()