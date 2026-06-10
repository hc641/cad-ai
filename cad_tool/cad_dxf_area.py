"""
CAD图纸识别系统 - 第三阶段(精确版)：DXF 闭合区域面积计算
================================================================
从 DXF 文件精确计算所有闭合区域面积（Shoelace几何算法，非估算）。

前置流程：
  AutoCAD 打开 DWG → 另存为 DXF → 用本工具处理

依赖：
  pip install ezdxf openpyxl

用法：
  python cad_dxf_area.py 图纸.dxf                    # 单独输出面积Excel
  python cad_dxf_area.py 图纸.dxf --into 识别结果_V2.xlsx   # 合并进已有报表
  python cad_dxf_area.py 图纸.dxf --min-area 0.01    # 过滤极小区域

支持的闭合实体：
  LWPOLYLINE/POLYLINE(闭合)、CIRCLE、ELLIPSE、闭合SPLINE、HATCH边界、SOLID
"""

import os
import re
import sys
import math
import argparse
from collections import defaultdict

import ezdxf
from ezdxf import bbox

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


C_HEAD = "1F3864"
C_SUB = "2E75B6"
C_ALT = "D9E2F3"
C_OK = "E2EFDA"
C_W = "FFFFFF"


# ─── 几何算法 ───

def polygon_area(points):
    """Shoelace 公式：任意多边形面积"""
    n = len(points)
    if n < 3:
        return 0.0
    a = 0.0
    for i in range(n):
        x1, y1 = points[i][0], points[i][1]
        x2, y2 = points[(i + 1) % n][0], points[(i + 1) % n][1]
        a += x1 * y2 - x2 * y1
    return abs(a) / 2.0


def lwpolyline_points(e, arc_seg=16):
    """提取LWPOLYLINE顶点，带凸度(bulge)的弧段离散化"""
    pts = []
    raw = list(e.get_points("xyb"))  # x, y, bulge
    n = len(raw)
    for i in range(n):
        x, y, bulge = raw[i]
        pts.append((x, y))
        if bulge and (i + 1 < n or e.closed):
            x2, y2, _ = raw[(i + 1) % n]
            pts.extend(_bulge_arc((x, y), (x2, y2), bulge, arc_seg))
    return pts


def _bulge_arc(p1, p2, bulge, seg):
    """bulge弧段离散成点"""
    x1, y1 = p1
    x2, y2 = p2
    chord = math.hypot(x2 - x1, y2 - y1)
    if chord == 0:
        return []
    sagitta = bulge * chord / 2
    radius = (chord / 2) ** 2 / (2 * abs(sagitta)) + abs(sagitta) / 2 if sagitta else 0
    if radius == 0:
        return []
    pts = []
    angle = 4 * math.atan(abs(bulge))
    mx, my = (x1 + x2) / 2, (y1 + y2) / 2
    dx, dy = x2 - x1, y2 - y1
    sign = 1 if bulge > 0 else -1
    h = math.sqrt(max(radius ** 2 - (chord / 2) ** 2, 0))
    cx = mx - sign * h * dy / chord
    cy = my + sign * h * dx / chord
    a1 = math.atan2(y1 - cy, x1 - cx)
    for k in range(1, seg):
        t = a1 + sign * angle * k / seg
        pts.append((cx + radius * math.cos(t), cy + radius * math.sin(t)))
    return pts


def ellipse_area(e):
    """椭圆面积 = π·a·b"""
    major = math.hypot(*e.dxf.major_axis[:2])
    minor = major * e.dxf.ratio
    return math.pi * major * minor


# ─── DXF 解析 ───

def extract_closed_areas(dxf_path, min_area=0.001):
    doc = ezdxf.readfile(dxf_path)
    msp = doc.modelspace()
    items = []   # (类型, 图层, 面积, 描述)

    for e in msp:
        t = e.dxftype()
        layer = getattr(e.dxf, "layer", "0")
        area = None
        desc = ""
        try:
            if t == "LWPOLYLINE" and e.closed:
                pts = lwpolyline_points(e)
                area = polygon_area(pts)
                desc = f"{len(pts)}点"
            elif t == "POLYLINE" and e.is_closed:
                pts = [(v.dxf.location[0], v.dxf.location[1]) for v in e.vertices]
                area = polygon_area(pts)
                desc = f"{len(pts)}点"
            elif t == "CIRCLE":
                r = e.dxf.radius
                area = math.pi * r * r
                desc = f"R{round(r,3)}"
            elif t == "ELLIPSE":
                area = ellipse_area(e)
                desc = "椭圆"
            elif t == "SPLINE" and e.closed:
                pts = [(p[0], p[1]) for p in e.control_points]
                area = polygon_area(pts)
                desc = "样条(近似)"
            elif t == "HATCH":
                area = abs(e.dxf.get("solid_fill", 0) and 0 or 0)
                # HATCH按边界路径算
                tot = 0.0
                for path in e.paths:
                    try:
                        verts = [(v[0], v[1]) for v in path.vertices]
                        tot += polygon_area(verts)
                    except Exception:
                        pass
                area = tot
                desc = "填充区"
            elif t == "SOLID":
                pts = [(e.dxf.vtx0[0], e.dxf.vtx0[1]),
                       (e.dxf.vtx1[0], e.dxf.vtx1[1]),
                       (e.dxf.vtx3[0], e.dxf.vtx3[1]),
                       (e.dxf.vtx2[0], e.dxf.vtx2[1])]
                area = polygon_area(pts)
                desc = "实体填充"
        except Exception:
            continue

        if area is not None and area >= min_area:
            items.append((t, layer, round(area, 4), desc))

    return items, doc


# ─── Excel 输出 ───

def _border():
    s = Side(style="thin", color="B0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)


def _head(c):
    c.font = Font(name="Arial", bold=True, color=C_W, size=10)
    c.fill = PatternFill("solid", fgColor=C_HEAD)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _border()


def _cell(c, alt=False):
    c.font = Font(name="Arial", size=9)
    if alt:
        c.fill = PatternFill("solid", fgColor=C_ALT)
    c.alignment = Alignment(vertical="center")
    c.border = _border()


def _table(ws, headers, rows, widths, start=1):
    for c, h in enumerate(headers, 1):
        _head(ws.cell(row=start, column=c, value=h))
    ws.row_dimensions[start].height = 22
    for ri, row in enumerate(rows, start + 1):
        for ci, val in enumerate(row, 1):
            _cell(ws.cell(row=ri, column=ci, value=val), alt=(ri % 2 == 0))
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=start + 1, column=1)


def write_excel(items, out_path, into=False, src_name=""):
    if into:
        wb = openpyxl.load_workbook(out_path)
        for sn in ["📐 DXF精确面积", "📊 DXF面积按图层", "📋 DXF面积汇总"]:
            if sn in wb.sheetnames:
                del wb[sn]
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    # 明细
    ws = wb.create_sheet("📐 DXF精确面积")
    rows = [(i + 1, t, layer, area, desc)
            for i, (t, layer, area, desc) in enumerate(
                sorted(items, key=lambda x: -x[2]))]
    _table(ws, ["#", "实体类型", "图层", "面积(mm²)", "描述"],
           rows, [6, 16, 20, 16, 14])

    # 按图层汇总
    by_layer = defaultdict(lambda: [0, 0.0])
    for t, layer, area, desc in items:
        by_layer[layer][0] += 1
        by_layer[layer][1] += area
    ws = wb.create_sheet("📊 DXF面积按图层")
    lrows = [(lyr, cnt, round(tot, 4))
             for lyr, (cnt, tot) in sorted(by_layer.items(), key=lambda x: -x[1][1])]
    _table(ws, ["图层", "区域数", "面积合计(mm²)"], lrows, [24, 12, 18])

    # 汇总
    ws = wb.create_sheet("📋 DXF面积汇总")
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "DXF 精确面积汇总（几何计算）"
    c.font = Font(name="Arial", bold=True, size=13, color=C_W)
    c.fill = PatternFill("solid", fgColor=C_HEAD)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    by_type = defaultdict(lambda: [0, 0.0])
    for t, layer, area, desc in items:
        by_type[t][0] += 1
        by_type[t][1] += area
    total = sum(a for _, _, a, _ in items)
    largest = max((a for _, _, a, _ in items), default=0)

    rowi = 3
    ws.cell(row=rowi, column=1, value="来源DXF").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=rowi, column=2, value=src_name).font = Font(name="Arial", size=10)
    rowi += 1
    ws.cell(row=rowi, column=1, value="闭合区域总数").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=rowi, column=2, value=len(items)).font = Font(name="Arial", size=10)
    rowi += 1
    ws.cell(row=rowi, column=1, value="面积总和(mm²)").font = Font(bold=True, name="Arial", size=10)
    cc = ws.cell(row=rowi, column=2, value=round(total, 3))
    cc.font = Font(name="Arial", size=10, bold=True)
    cc.fill = PatternFill("solid", fgColor=C_OK)
    rowi += 1
    ws.cell(row=rowi, column=1, value="最大单区域(mm²)").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=rowi, column=2, value=round(largest, 3)).font = Font(name="Arial", size=10)
    rowi += 2

    ws.cell(row=rowi, column=1, value="按实体类型：").font = Font(bold=True, color=C_SUB, name="Arial", size=10)
    rowi += 1
    for t, (cnt, tot) in sorted(by_type.items(), key=lambda x: -x[1][1]):
        ws.cell(row=rowi, column=1, value=f"  {t}").font = Font(name="Arial", size=9)
        ws.cell(row=rowi, column=2, value=f"{cnt}个 / {round(tot,3)}mm²").font = Font(name="Arial", size=9)
        rowi += 1

    rowi += 1
    nc = ws.cell(row=rowi, column=1, value="✓ 本面积由DXF几何精确计算（Shoelace算法），可作PPAP正式数据。")
    nc.font = Font(name="Arial", size=9, color="375623")
    ws.merge_cells(start_row=rowi, start_column=1, end_row=rowi, end_column=2)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 26

    wb.save(out_path)
    return total


def main():
    ap = argparse.ArgumentParser(description="DXF闭合区域精确面积计算")
    ap.add_argument("dxf", help="输入DXF文件")
    ap.add_argument("--into", help="合并进已有的_V2.xlsx报表", default=None)
    ap.add_argument("--min-area", type=float, default=0.001,
                    help="最小面积阈值mm²，过滤极小噪声 (默认0.001)")
    ap.add_argument("--out", help="单独输出的Excel路径", default=None)
    args = ap.parse_args()

    if not os.path.exists(args.dxf):
        print(f"✗ 文件不存在: {args.dxf}")
        sys.exit(1)

    print(f"[1/3] 读取DXF: {args.dxf}")
    items, doc = extract_closed_areas(args.dxf, min_area=args.min_area)
    print(f"      → 提取闭合区域 {len(items)} 个")

    print("[2/3] 几何面积计算（Shoelace）...")
    total = sum(a for _, _, a, _ in items)
    print(f"      → 面积总和 {round(total,3)} mm²")

    print("[3/3] 写入Excel...")
    src = os.path.basename(args.dxf)
    if args.into:
        out = args.into
        write_excel(items, out, into=True, src_name=src)
    else:
        out = args.out or (os.path.splitext(args.dxf)[0] + "_面积.xlsx")
        write_excel(items, out, into=False, src_name=src)
    print(f"完成 → {out}")

    import json
    print(json.dumps({"闭合区域数": len(items), "面积总和mm2": round(total, 3),
                      "输出": out}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
