from collections import Counter
"""
CAD图纸PDF识别系统 - 第一阶段
功能：PDF解析 → 结构化信息提取 → Excel报表输出
适用于：汽车行业工程图纸（连接器/钣金/总成等）
"""

import re
import sys
import json
from pathlib import Path
from datetime import datetime
from collections import OrderedDict

import fitz  # PyMuPDF (主文本引擎)
import pdfplumber  # 表格检测辅助
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ═════════════════════════════════════════════
# 1. PDF 解析层
# ═════════════════════════════════════════════

class PDFParser:
    """提取文本(PyMuPDF)、表格(pdfplumber)、元数据"""

    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.text = ""
        self.lines = []
        self.pages_tables = []
        self.metadata = {}
        self.page_count = 0
        self.page_dims = []

    def parse(self):
        doc = fitz.open(self.pdf_path)
        self.page_count = doc.page_count
        meta = doc.metadata or {}
        parts = []
        for page in doc:
            parts.append(page.get_text())
            self.page_dims.append((round(page.rect.width), round(page.rect.height)))
        self.text = "\n".join(parts)
        self.lines = [l.strip() for l in self.text.splitlines() if l.strip()]
        self.metadata = {
            "文件名": Path(self.pdf_path).name,
            "PDF标题": meta.get("title", ""),
            "创建工具": meta.get("creator", ""),
            "创建日期": meta.get("creationDate", ""),
            "页数": self.page_count,
            "页面尺寸(pt)": " / ".join(f"{w}x{h}" for w, h in self.page_dims),
        }
        doc.close()

        with pdfplumber.open(self.pdf_path) as pdf:
            for page in pdf.pages:
                try:
                    self.pages_tables.append(page.extract_tables() or [])
                except Exception:
                    self.pages_tables.append([])
        return self


# ═════════════════════════════════════════════
# 2. 信息提取层
# ═════════════════════════════════════════════

class CADInfoExtractor:

    def __init__(self, parser: PDFParser):
        self.p = parser
        self.t = parser.text
        self.lines = parser.lines
        self.r = OrderedDict([
            ("标题栏信息", OrderedDict()),
            ("技术要求",   []),
            ("修订历史",   []),
            ("物料清单BOM", []),
            ("关键特性",   []),
            ("偏差清单",   []),
            ("尺寸标注",   []),
            ("GDT几何公差", []),
            ("孔位螺纹",   []),
            ("材料信息",   []),
            ("零件接口矩阵", []),
            ("一般公差表", []),
        ])

    def extract_all(self):
        self._title_block()
        self._notes()
        self._revisions()
        self._bom()
        self._characteristics()
        self._deviations()
        self._dimensions()
        self._gdt()
        self._holes()
        self._materials()
        self._part_matrix()
        self._tolerance_chart()
        return self

    # ── 标题栏（锚点+频次法，规避乱序） ──────────
    def _title_block(self):
        tb = self.r["标题栏信息"]
        eight = re.findall(r"\b(\d{8})\b", self.t)
        # 图号：优先用文件名中的8位号（最可靠），否则取频次最高
        fname_num = re.search(r"(\d{8})", self.p.metadata.get("文件名", ""))
        if fname_num and fname_num.group(1) in eight:
            tb["图号 (DWG NO.)"] = fname_num.group(1)
        elif eight:
            tb["图号 (DWG NO.)"] = Counter(eight).most_common(1)[0][0]
        m = re.search(r"(TAXI CONN[^\n]*|[A-Z][A-Z0-9 .]*CONN[A-Z0-9 .]*HEADER)", self.t)
        tb["图名 (DWG NAME)"] = m.group(1).strip() if m else ""
        m = re.search(r"\b(\d{1,2}:\d{1,2})\b", self.t)
        tb["比例 (SCALE)"] = m.group(1) if m else ""
        m = re.search(r"\b(A\d+x\d+)\b", self.t)
        tb["图幅 (SIZE)"] = m.group(1) if m else ""
        m = re.search(r"(\d+)\s*OF\s*(\d+)", self.t)
        tb["张数 (SHEET)"] = f"{m.group(1)} OF {m.group(2)}" if m else ""
        tb["最新版本 (REV)"] = ""  # 占位，_revisions() 后回填
        designers = re.findall(r"\b([A-Z][A-Z]+(?:\s[A-Z]+)+)\s+(\d{2}[A-Z]{2}\d{2})\b", self.t)
        if designers:
            uniq = list(OrderedDict.fromkeys((n, d) for n, d in designers))
            tb["设计/审核人员"] = "; ".join(f"{n}({d})" for n, d in uniq[:6])
        tb["标注标准"] = "ASME Y14.5-2009" if "Y14.5-2009" in self.t else ""
        tb["投影方式"] = "第三角投影 THIRD ANGLE" if "THIRD ANGLE" in self.t else ""
        tb["单位"] = "MILLIMETERS (mm)" if "MILLIMETER" in self.t else ""
        m = re.search(r"\b([A-Z][a-z]+,\s*China)\b", self.t)
        tb["产地"] = m.group(1) if m else ""
        comps = [c for c in ["APTIV", "CONTINENTAL", "MOLEX"] if c in self.t]
        tb["涉及公司"] = ", ".join(comps)

    # ── 技术要求 NOTES（分级编号） ──────────────
    def _notes(self):
        idx = self.t.find("NOTES")
        if idx < 0:
            return
        block = self.t[idx:idx + 6000]
        # 编号识别：排除孤立的 2、3 等（可能是上标 mm² 断裂产生）
        # 用负向前瞻确保不是跟在单位后面的上标数字
        pattern = re.compile(
            r"\n\s*(\d+(?:\.\d+)*)\.?\s+([A-Z][^\n]{3,}(?:\n(?!\s*\d+(?:\.\d+)*\.?\s)[^\n]+)*)"
        )
        for m in pattern.finditer(block):
            num = m.group(1)
            content = re.sub(r"\s+", " ", m.group(2)).strip()
            if len(content) > 3:
                self.r["技术要求"].append({
                    "编号": num,
                    "层级": num.count(".") + 1,
                    "内容": content[:400],
                })

        # 后处理：修复上标字符断裂
        # 模式：数字/单位（mm/cm/m）后跟孤立的 2 或 3 → 合并为上标形式
        self._fix_superscripts()

    # ── 上标修复辅助（mm²/mm³ 断裂合并） ─────
    def _fix_superscripts(self):
        """修复 PDF 文本提取中上标字符被拆分为独立行的问题。
        识别 '数字+单位' 后紧跟孤立的 '2' 或 '3' 的模式，合并为 mm² / mm³。
        """
        superscript_map = {"2": "²", "3": "³"}
        for note in self.r["技术要求"]:
            content = note["内容"]
            # 模式: 数字 + 可选空格 + mm/cm/m + 可选空格 + 2/3 → 合并
            # 例如: "1mm 2" → "1mm²", "5 cm 3" → "5 cm³"
            # 注意 "2" 或 "3" 后面应该是空格、标点或字符串结尾
            content = re.sub(
                r"(\d+\s*(?:mm|cm|m|in|ft))\s+(2|3)\b",
                lambda m: m.group(1) + superscript_map.get(m.group(2), m.group(2)),
                content,
                flags=re.IGNORECASE
            )
            note["内容"] = content

    # ── 修订历史 ─────────────────────────────
    def _revisions(self):
        seen = {}
        # 锚定日期码，在其后窗口内独立抓取 版本/描述/ECN
        for m in re.finditer(r"\b(\d{2}[A-Z]{2}\d{2})\b", self.t):
            date = m.group(1)
            window = self.t[m.end(): m.end() + 160]
            rev_m = re.search(r"R\s*\n?\s*0?(\d{1,2})\b", window)
            desc_m = re.search(
                r"(INITIAL RELEASED|ADDED[^\n]*|UPDATED[^\n]*|REVISED[^\n]*|FOR CODE[^\n]*)",
                window)
            ecn_m = re.search(r"\b(\d{6})\b", window)
            if not rev_m:
                continue
            rev = int(rev_m.group(1))
            if not (1 <= rev <= 50):
                continue
            ver = f"R{rev:02d}"
            desc = re.sub(r"\s+", " ", desc_m.group(1)).strip() if desc_m else ""
            # 同一版本保留信息最全的一条
            if ver not in seen or (desc and not seen[ver]["变更说明"]):
                seen[ver] = {
                    "版本": ver,
                    "日期": date,
                    "变更说明": desc[:120],
                    "ECN": ecn_m.group(1) if ecn_m else "",
                }
        self.r["修订历史"] = sorted(seen.values(), key=lambda x: x["版本"])
        # 回填标题栏最新版本（取修订历史中的最大版本，可靠）
        if self.r["修订历史"]:
            self.r["标题栏信息"]["最新版本 (REV)"] = self.r["修订历史"][-1]["版本"]

    # ── BOM 物料清单 ─────────────────────────
    def _bom(self):
        seen = set()
        for m in re.finditer(r"(\d{8})[\s\S]{0,80}?USED IN ASSEMBLY:\s*(\d{6,8})", self.t):
            part, asm = m.group(1), m.group(2)
            if (part, asm) in seen:
                continue
            seen.add((part, asm))
            ctx = m.group(0)
            color = next((c for c in ["BLACK", "BLUE", "GREY", "GRAY", "YELLOW"] if c in ctx), "")
            mm = re.search(r"(PBT[ -]?GF\d+|S\d{7}/\d+)", ctx)
            self.r["物料清单BOM"].append({
                "零件号": part,
                "颜色": color,
                "材料规格": mm.group(1) if mm else "",
                "所属总成": asm,
            })

        # 后处理：颜色和材料规格向下继承（图纸中同一颜色/材料的连续零件仅在首条标注）
        last_color = ""
        last_material = ""
        for item in self.r["物料清单BOM"]:
            if item["颜色"]:
                last_color = item["颜色"]
            else:
                item["颜色"] = last_color
            if item["材料规格"]:
                last_material = item["材料规格"]
            else:
                item["材料规格"] = last_material

    # ── 关键特性 SC/CI/KPC ───────────────────
    def _characteristics(self):
        # 使用更宽松的空白匹配，避免因换行/空格分布导致漏提取
        # 说明字段扩展捕获范围：MATING / RETENTION / FEMALE / MODULE 等
        pattern = re.compile(
            r"\b(SC|CI|KPC)\s+([\d.]+(?:\+/-[\d.]+)?)\s+"
            r"((?:MATING|RETENTION|FEMALE|MODULE|WITH)[^\n]*)",
            re.IGNORECASE
        )
        seen = set()
        for m in pattern.finditer(self.t):
            spec = m.group(2).replace("+/-", "±")
            desc = re.sub(r"\s+", " ", m.group(3)).strip()[:120]
            key = (m.group(1), spec, desc)
            if key in seen:
                continue
            seen.add(key)
            self.r["关键特性"].append({
                "类型": m.group(1),
                "规格": spec,
                "说明": desc,
            })
        # 补充扫描：KPC表格区域常见特征——查找 "SC" 后跟数字的行
        # 以防表格换行过于复杂导致主正则漏掉
        kpc_section_start = self.t.find("CHARACTERISTIC")
        if kpc_section_start < 0:
            kpc_section_start = self.t.find("KPC")
        if kpc_section_start >= 0:
            kpc_block = self.t[kpc_section_start:kpc_section_start + 3000]
            fallback = re.compile(
                r"(SC|CI|KPC)\s+([\d.]+±[\d.]+)",
                re.IGNORECASE
            )
            for m in fallback.finditer(kpc_block):
                spec = m.group(2)
                key = (m.group(1), spec)
                if any(key[1] == e["规格"] and key[0] == e["类型"] for e in self.r["关键特性"]):
                    continue
                self.r["关键特性"].append({
                    "类型": m.group(1),
                    "规格": spec,
                    "说明": "(表格提取-说明文本在扫描窗口外)",
                })

    # ── 偏差清单 ─────────────────────────────
    def _deviations(self):
        """解析两种偏差表格式：
        格式1 (旧)：DEVIATION LIST  —  "39# 6.4+/-0.02  6.3+/-0.02"
        格式2 (新)：DEVIATION TABLE —  "DIM NO. DEVIATION ACCEPTABLE"
                    行格式 "39 6.30+0.02/-0.04" 或 "85 0.15 A B C" 或 "10 Rz15-30"
        """
        seen = set()

        # ── 格式1: DEVIATION LIST ──────────────────────────────────────
        idx1 = self.t.find("DEVIATION LIST")
        if idx1 >= 0:
            block = self.t[idx1:idx1 + 3000]
            pattern = re.compile(
                r"(\d+)#\s*[\n ]\s*([\d.]+(?:\+/-[\d.]+)?)\s*[\n ]\s*([\d.+/\-]+(?:\s*TBT)?)")
            for m in pattern.finditer(block):
                key = m.group(0)
                if key in seen:
                    continue
                seen.add(key)
                self.r["偏差清单"].append({
                    "尺寸编号": m.group(1) + "#",
                    "原规格": m.group(2).replace("+/-", "±"),
                    "偏差后": m.group(3).replace("+/-", "±").strip(),
                })

        # ── 格式2: DEVIATION TABLE (R03+ 样式) ────────────────────────
        for marker in ("DEVIATION TABLE", "DIM NO. DEVIATION", "DIM NO."):
            idx2 = self.t.find(marker)
            if idx2 >= 0:
                break
        else:
            idx2 = -1

        if idx2 >= 0:
            block2 = self.t[idx2:idx2 + 4000]
            # dim_no 必须是纯整数（排除 "5.5"、"453021" 等非编号数字）
            # 偏差规格：贪婪匹配到行尾（含基准字母、Rz、ALLOWED WARPAGE 等）
            pat2 = re.compile(
                r"^[ \t]*(\d{1,4})[ \t]+"            # dim_no: 1~4位整数
                r"(.+?)[ \t]*$",                         # spec: 贪婪到行尾
                re.MULTILINE,
            )
            for m in pat2.finditer(block2):
                dim_no = m.group(1)
                spec = m.group(2).strip().replace("+/-", "±")
                # 排除纯大写关键词行（FOR、TOOL、WARPAGE、NO、DIM 等）
                if re.fullmatch(r"[A-Z ]+", spec):
                    continue
                # 排除表头行（含 DEVIATION、ACCEPTABLE、TOOL）
                if any(kw in spec.upper() for kw in
                       ("DEVIATION", "ACCEPTABLE", "TOOL NO", "WARPAGE ALLOWED",
                        "ALLOWED WARPAGE")):
                    # 但保留 "0.35 ALLOWED WARPAGE -0.20" 这种带数值的行
                    if not re.search(r"\d", spec[:6]):
                        continue
                key = f"TABLE:{dim_no}:{spec}"
                if key in seen:
                    continue
                seen.add(key)
                self.r["偏差清单"].append({
                    "尺寸编号": dim_no + "#",
                    "原规格": "",
                    "偏差后": spec,
                })

    # ── 尺寸标注（合并标注值+公差，识别*号受控尺寸） ──
    def _dimensions(self):
        # 第一阶段：收集所有候选标注行（含索引）
        candidates = []  # (line_index, text, is_starred)
        for i, l in enumerate(self.lines):
            if len(l) > 40:
                continue
            has_num = bool(re.search(r"\d+\.\d+|\d+", l))
            has_tol = "±" in l or "+/-" in l
            has_prefix = bool(re.match(r"^[ΦØR]", l)) or bool(re.search(r"X45", l))
            has_mult = bool(re.match(r"^\d+X\s+", l))
            if has_num and (has_tol or has_prefix or has_mult or
                            re.fullmatch(r"[\d.]+", l) or
                            re.match(r"^[ΦØR]?\s*[\d.]+", l)):
                is_starred = l.rstrip().endswith("*") or bool(re.search(r"[\d.]\*", l))
                text = l.rstrip("*").strip()
                candidates.append((i, text, is_starred))

        # 第二阶段：合并相邻的标注值+公差行
        merged = []
        skip_next = set()
        for idx, (ci, ct, cs) in enumerate(candidates):
            if ci in skip_next:
                continue
            tol_text = ""
            cur_is_numeric = bool(re.fullmatch(r"[\d.]+", ct))

            if idx + 1 < len(candidates):
                ni, nt, ns = candidates[idx + 1]
                next_is_tolerance = (nt.startswith("±") or nt.startswith("+/-") or
                                     bool(re.fullmatch(r"[\d.]+", nt)))
                gap_ok = (ni - ci) <= 2
                # Case A: 纯数值 + 纯数值（公差值缺失±） → 合并为 val±tol
                if cur_is_numeric and next_is_tolerance and gap_ok:
                    tol_text = nt if (nt.startswith("±") or nt.startswith("+/-")) else "±" + nt
                    merged.append((ci, ct + tol_text, cs or ns))
                    skip_next.add(ni)
                    continue
                # Case B: 当前任意标注 + 下一行以±开头 → 拼接
                if (nt.startswith("±") or nt.startswith("+/-")) and gap_ok:
                    tol_text = nt.replace("+/-", "±")
                    merged.append((ci, ct + tol_text, cs or ns))
                    skip_next.add(ni)
                    continue

            merged.append((ci, ct.replace("+/-", "±"), cs))

        # 第三阶段：去重 + 提取数量倍数
        seen = set()
        for _, text, is_star in merged:
            norm = text.strip()
            if not norm or len(norm) < 2:
                continue
            if norm in seen:
                continue
            seen.add(norm)
            mult_m = re.match(r"^(\d+)X\s+", norm)
            self.r["尺寸标注"].append({
                "数量倍数": mult_m.group(1) + "X" if mult_m else "",
                "标注值": norm,
                "偏差受控": "是 ⚠" if is_star else "",
            })

    # ── GD&T 几何公差（带基准，类型细分） ───────
    def _gdt(self):
        # GD&T 特征名称 → 符号映射表（ASME Y14.5）
        gdt_type_map = [
            (r"TRUE\s*POSITION|POSITION", "⊕ 位置度 (Position)"),
            (r"FLATNESS", "▱ 平面度 (Flatness)"),
            (r"STRAIGHTNESS", "— 直线度 (Straightness)"),
            (r"CIRCULARITY|ROUNDNESS", "○ 圆度 (Circularity)"),
            (r"CYLINDRICITY", "⌭ 圆柱度 (Cylindricity)"),
            (r"PROFILE\s*OF\s*(?:A\s*)?LINE", "⌒ 线轮廓度 (Profile of a Line)"),
            (r"PROFILE\s*OF\s*(?:A\s*)?SURFACE|SURFACE\s*PROFILE", "⌓ 面轮廓度 (Profile of a Surface)"),
            (r"PARALLELISM", "∥ 平行度 (Parallelism)"),
            (r"PERPENDICULARITY|SQUARENESS", "⊥ 垂直度 (Perpendicularity)"),
            (r"ANGULARITY", "∠ 倾斜度 (Angularity)"),
            (r"CONCENTRICITY|COAXIALITY", "◎ 同轴度 (Concentricity)"),
            (r"SYMMETRY", "↔ 对称度 (Symmetry)"),
            (r"CIRCULAR\s*RUNOUT|RADIAL\s*RUNOUT", "↗ 圆跳动 (Circular Runout)"),
            (r"TOTAL\s*RUNOUT", "↗ 全跳动 (Total Runout)"),
            (r"POSITION|TRUE\s*POSITION", "⊕ 位置度 (Position)"),
        ]

        def classify_gdt(context_text):
            """根据周围文本推断GD&T类型"""
            ctx_upper = context_text.upper()
            for pattern, label in gdt_type_map:
                if re.search(pattern, ctx_upper):
                    return label
            return "位置/形位 (带基准)"

        seen = set()
        for l in self.lines:
            # 匹配公差值 + 基准字母（如 "0.1 A B", "0.25 A"）
            m = re.fullmatch(r"(0\.\d+)\s+([A-Z](?:\s+[A-Z]){0,2})\s*\*?", l)
            if not m:
                # 也匹配带直径符号的（如 "Ø0.1 A"）
                m = re.fullmatch(r"[ΦØ]?\s*(0\.\d+)\s+([A-Z](?:\s+[A-Z]){0,2})\s*\*?", l)
            if not m:
                continue
            key = (m.group(1), m.group(2))
            if key in seen:
                continue
            seen.add(key)

            # 在该行附近搜索GD&T类型关键字（前后200字符）
            line_pos = self.t.find(l)
            if line_pos < 0:
                line_pos = 0
            ctx_start = max(0, line_pos - 200)
            ctx_end = min(len(self.t), line_pos + 200)
            context = self.t[ctx_start:ctx_end]

            gdt_type = classify_gdt(context)
            self.r["GDT几何公差"].append({
                "公差值": m.group(1),
                "基准": m.group(2),
                "类型": gdt_type,
            })

    # ── 孔/螺纹/倒角/圆角 ────────────────────
    def _holes(self):
        seen = set()
        patterns = {
            "螺纹": r"\bM\d+(?:x[\d.]+)?(?:-\d+[Hh])?\b",
            "直径": r"[ΦØ]\s*[\d.]+",
            "倒角": r"[\d.]+\s*X45",
            "圆角R": r"\b\d+X\s+R\s*[\d.]+",
        }
        for typ, pat in patterns.items():
            for m in re.finditer(pat, self.t):
                v = re.sub(r"\s+", "", m.group(0))
                if v not in seen:
                    seen.add(v)
                    self.r["孔位螺纹"].append({"类型": typ, "标注": m.group(0).strip()})

    # ── 材料信息 ─────────────────────────────
    def _materials(self):
        found = OrderedDict()
        m = re.search(r"\b(PBT[ -]?GF\d+)\b", self.t)
        if m:
            found["基材"] = m.group(1).replace(" ", "-")
        rals = sorted(set(re.findall(r"\bRAL\s?\d{4}\b", self.t)))
        if rals:
            found["RAL颜色"] = ", ".join(rals)
        m = re.search(r"(UL94-?\w+)", self.t)
        if m:
            found["阻燃等级"] = m.group(1)
        m = re.search(r"(-?\d+\s*\.\.\.\s*\+?\d+\s*C)", self.t)
        if m:
            found["工作温度"] = re.sub(r"\s+", " ", m.group(1))
        m = re.search(r"VDA\s?\d+", self.t)
        if m:
            found["材料标识标准"] = m.group(0)
        for k, v in found.items():
            self.r["材料信息"].append({"项目": k, "内容": v})

    # ── 零件-接口矩阵 ────────────────────────
    def _part_matrix(self):
        seen = set()
        for m in re.finditer(r"(\d{8})\s*/\s*(\d+P)\s+([A-Z0-9]+)", self.t):
            key = m.group(1)
            if key in seen:
                continue
            seen.add(key)
            self.r["零件接口矩阵"].append({
                "Aptiv料号": m.group(1),
                "针数": m.group(2),
                "客户料号": m.group(3),
            })

    # ── 一般公差表 ───────────────────────────
    def _tolerance_chart(self):
        idx = self.t.find("GENERAL TOLERANCE")
        if idx < 0:
            idx = self.t.find("TOLERANCE UNLESS")
        if idx < 0:
            idx = self.t.find("DIMENSIONAL RANGE")
        if idx < 0:
            return
        block = self.t[idx:idx + 800]

        # ── 方案一：同一行内 "+/-0.15" 或 "±0.15" 写法（紧凑表格） ──
        tols = re.findall(r"(?:\+/-|±)\s*([\d.]+)", block)
        ranges = re.findall(r"(?:FROM\s+)?[>\s]*(\d+)\s*(?:TO|–|-)\s*(\d+)", block)

        # 同时抓取角度公差（±2° 格式，同行）
        ang_m = re.search(r"ANGULAR\s+TOLERANCE\s*[±+/-]+\s*([\d.]+)\s*°", block, re.IGNORECASE)
        angular_tol = float(ang_m.group(1)) if ang_m else None

        # ── 方案二：PyMuPDF 逐行提取，FROM/TO 数字与公差值各自成块 ──
        # 典型结构（每个数字单独一行）：
        #   FROM\nTO\n...ANGULAR TOLERANCE\nE1\n
        #   0\n20\n20\n30\n30\n70\n...\n300\n400        ← 9对 FROM/TO（18个数）
        #   0.15\n0.2\n0.3\n...\n1\n1.2                  ← 9个线性公差
        #   2                                            ← 角度公差
        if not ranges or not tols:
            lines = block.splitlines()
            # 找到 "E1" 或 CHART 标记后的数字序列起点
            start = 0
            for i, ln in enumerate(lines):
                if re.fullmatch(r"E\d+", ln.strip()):
                    start = i + 1
                    break
            else:
                # 没有 CHART 标记，从 "TO" 之后开始找数字
                for i, ln in enumerate(lines):
                    if ln.strip() == "TO":
                        start = i + 1
                        break

            nums = []
            for ln in lines[start:start + 40]:
                ln = ln.strip()
                if re.fullmatch(r"\d+(?:\.\d+)?", ln):
                    nums.append(ln)
                elif nums:
                    # 遇到非数字行（如 "A", "SECTION A-A"）说明数字块结束
                    break

            int_nums = [n for n in nums if "." not in n]
            if len(int_nums) >= 18:
                from_to = int_nums[:18]
                ranges = [(from_to[i], from_to[i + 1]) for i in range(0, 18, 2)]
                tail = nums[18:]
            else:
                tail = nums

            # tail 形如 [0.15, 0.2, 0.3, 0.4, 0.5, 0.6, 0.8, 1, 1.2, 2]
            #   前 N 个对应 ranges 数量，最后一个为角度公差
            if ranges and len(tail) >= len(ranges):
                n_lin = len(ranges)
                tols = tail[:n_lin]
                if len(tail) > n_lin and angular_tol is None:
                    try:
                        angular_tol = float(tail[n_lin])
                    except ValueError:
                        pass
            elif tail:
                tols = tail

        for i, (lo, hi) in enumerate(ranges):
            self.r["一般公差表"].append({
                "尺寸范围(mm)": f"{lo} – {hi}",
                "公差(±mm)": tols[i] if i < len(tols) else "",
            })

        # 角度公差单独存一条，供 Vision 层查用
        if angular_tol is not None:
            self.r["一般公差表"].append({
                "尺寸范围(mm)": "角度",
                "公差(±mm)": str(angular_tol),
            })


# ═════════════════════════════════════════════
# 3. Excel 输出层
# ═════════════════════════════════════════════

class ExcelExporter:
    C_HEAD = "1F3864"
    C_SUB = "2E75B6"
    C_ALT = "D9E2F3"
    C_KEY = "FFF2CC"
    C_W = "FFFFFF"

    def __init__(self, extractor: CADInfoExtractor, output_path: str):
        self.r = extractor.r
        self.meta = extractor.p.metadata
        self.full_text = extractor.t
        self.output_path = output_path
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    def _border(self):
        s = Side(style="thin", color="B0B0B0")
        return Border(left=s, right=s, top=s, bottom=s)

    def _head(self, cell):
        cell.font = Font(name="Arial", bold=True, color=self.C_W, size=10)
        cell.fill = PatternFill("solid", fgColor=self.C_HEAD)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = self._border()

    def _cell(self, cell, alt=False):
        cell.font = Font(name="Arial", size=9)
        if alt:
            cell.fill = PatternFill("solid", fgColor=self.C_ALT)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = self._border()

    def _widths(self, ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _table(self, ws, headers, rows, widths, start=1, empty_msg="（未识别）"):
        if not rows:
            ws.cell(row=start, column=1, value=empty_msg).font = Font(name="Arial", italic=True, color="999999")
            return
        for c, h in enumerate(headers, 1):
            self._head(ws.cell(row=start, column=c, value=h))
        ws.row_dimensions[start].height = 22
        for ri, row in enumerate(rows, start + 1):
            for ci, val in enumerate(row, 1):
                self._cell(ws.cell(row=ri, column=ci, value=val), alt=(ri % 2 == 0))
        if widths:
            self._widths(ws, widths)
        ws.freeze_panes = ws.cell(row=start + 1, column=1)

    def _overview(self):
        ws = self.wb.create_sheet("📋 图纸信息")
        ws.merge_cells("A1:D1")
        c = ws["A1"]
        c.value = "汽车CAD图纸识别报告"
        c.font = Font(name="Arial", bold=True, size=14, color=self.C_W)
        c.fill = PatternFill("solid", fgColor=self.C_HEAD)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        ws.merge_cells("A2:D2")
        ws["A2"].value = f"生成时间 {datetime.now():%Y-%m-%d %H:%M}   |   来源 {self.meta.get('文件名','')}"
        ws["A2"].font = Font(name="Arial", size=9, color="666666")
        ws["A2"].alignment = Alignment(horizontal="center")

        row = 4
        def section(title):
            nonlocal row
            ws.cell(row=row, column=1, value=title).font = Font(bold=True, color=self.C_SUB, name="Arial", size=11)
            row += 1
        def kv(k, v, hl=False):
            nonlocal row
            c1 = ws.cell(row=row, column=1, value=k)
            c1.font = Font(bold=True, name="Arial", size=9)
            if hl:
                c1.fill = PatternFill("solid", fgColor=self.C_KEY)
            c2 = ws.cell(row=row, column=2, value=str(v) if v else "（未识别）")
            c2.font = Font(name="Arial", size=9, color="C00000" if not v else "000000")
            c2.alignment = Alignment(wrap_text=True)
            row += 1

        section("【PDF 元数据】")
        for k, v in self.meta.items():
            kv(k, v)
        row += 1
        section("【标题栏识别结果】")
        for k, v in self.r["标题栏信息"].items():
            kv(k, v, hl=True)
        row += 1
        section("【识别统计】")
        for name, data in self.r.items():
            if name == "标题栏信息":
                continue
            c1 = ws.cell(row=row, column=1, value=name)
            c1.font = Font(name="Arial", size=9, bold=True)
            c2 = ws.cell(row=row, column=2, value=f"{len(data)} 条")
            c2.font = Font(name="Arial", size=9)
            c2.fill = PatternFill("solid", fgColor=self.C_ALT)
            row += 1
        self._widths(ws, [26, 52, 14, 14])

    def export(self):
        self._overview()

        ws = self.wb.create_sheet("📝 技术要求")
        self._table(ws, ["编号", "层级", "技术要求内容"],
                    [(d["编号"], d["层级"], d["内容"]) for d in self.r["技术要求"]],
                    [10, 8, 110], empty_msg="（未识别到NOTES）")

        ws = self.wb.create_sheet("🔄 修订历史")
        self._table(ws, ["版本", "日期", "变更说明", "ECN号"],
                    [(d["版本"], d["日期"], d["变更说明"], d["ECN"]) for d in self.r["修订历史"]],
                    [10, 12, 60, 12])

        ws = self.wb.create_sheet("📦 物料清单BOM")
        self._table(ws, ["#", "零件号", "颜色", "材料规格", "所属总成"],
                    [(i+1, d["零件号"], d["颜色"], d["材料规格"], d["所属总成"])
                     for i, d in enumerate(self.r["物料清单BOM"])],
                    [6, 16, 12, 22, 16])

        ws = self.wb.create_sheet("⭐ 关键特性")
        self._table(ws, ["#", "类型", "规格", "说明"],
                    [(i+1, d["类型"], d["规格"], d["说明"])
                     for i, d in enumerate(self.r["关键特性"])],
                    [6, 10, 18, 50], empty_msg="（未识别到SC/CI关键特性）")

        ws = self.wb.create_sheet("📋 偏差清单")
        self._table(ws, ["尺寸编号", "原规格", "偏差后规格"],
                    [(d["尺寸编号"], d["原规格"], d["偏差后"]) for d in self.r["偏差清单"]],
                    [12, 22, 28])

        ws = self.wb.create_sheet("📐 尺寸标注")
        self._table(ws, ["#", "数量倍数", "标注值", "偏差受控"],
                    [(i+1, d["数量倍数"], d["标注值"], d.get("偏差受控", ""))
                     for i, d in enumerate(self.r["尺寸标注"])],
                    [6, 12, 30, 12])

        ws = self.wb.create_sheet("⊕ GD&T几何公差")
        self._table(ws, ["#", "公差值", "基准", "类型"],
                    [(i+1, d["公差值"], d["基准"], d["类型"])
                     for i, d in enumerate(self.r["GDT几何公差"])],
                    [6, 12, 14, 34])

        ws = self.wb.create_sheet("🔩 孔位螺纹")
        self._table(ws, ["#", "类型", "标注"],
                    [(i+1, d["类型"], d["标注"])
                     for i, d in enumerate(self.r["孔位螺纹"])],
                    [6, 12, 24])

        ws = self.wb.create_sheet("🔬 材料信息")
        self._table(ws, ["项目", "内容"],
                    [(d["项目"], d["内容"]) for d in self.r["材料信息"]],
                    [18, 50])

        ws = self.wb.create_sheet("🔌 零件接口矩阵")
        self._table(ws, ["#", "Aptiv料号", "针数", "客户料号"],
                    [(i+1, d["Aptiv料号"], d["针数"], d["客户料号"])
                     for i, d in enumerate(self.r["零件接口矩阵"])],
                    [6, 16, 10, 20])

        ws = self.wb.create_sheet("📏 一般公差表")
        self._table(ws, ["尺寸范围(mm)", "公差(±mm)"],
                    [(d["尺寸范围(mm)"], d["公差(±mm)"]) for d in self.r["一般公差表"]],
                    [18, 14])

        ws = self.wb.create_sheet("📄 原始文本")
        ws["A1"] = "PDF原始抽取文本（供人工核对）"
        ws["A1"].font = Font(bold=True, color=self.C_SUB, name="Arial")
        for i, line in enumerate(self.full_text.splitlines(), 2):
            ws.cell(row=i, column=1, value=line).font = Font(name="Consolas", size=8)
        ws.column_dimensions["A"].width = 120

        self.wb.save(self.output_path)
        return self.output_path


# ═════════════════════════════════════════════
# 4. 主入口
# ═════════════════════════════════════════════

def process_cad_pdf(pdf_path: str, output_dir: str = ".") -> dict:
    pdf_path = str(pdf_path)
    stem = Path(pdf_path).stem
    output_path = str(Path(output_dir) / f"{stem}_识别结果.xlsx")

    print(f"[1/4] 解析PDF (PyMuPDF): {pdf_path}")
    parser = PDFParser(pdf_path).parse()
    print(f"      → {parser.page_count} 页，{len(parser.text)} 字符，{len(parser.lines)} 行")

    print("[2/4] 提取结构化信息...")
    extractor = CADInfoExtractor(parser).extract_all()
    for k, v in extractor.r.items():
        if isinstance(v, list):
            print(f"      {k}: {len(v)} 条")

    print("[3/4] 生成Excel报表...")
    ExcelExporter(extractor, output_path).export()

    print(f"[4/4] 完成 → {output_path}")
    summary = {k: (len(v) if isinstance(v, list) else v) for k, v in extractor.r.items()}
    return {"output": output_path, "summary": summary}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python cad_pdf_recognizer.py <图纸.pdf> [输出目录]")
        sys.exit(1)
    out = sys.argv[2] if len(sys.argv) > 2 else "."
    result = process_cad_pdf(sys.argv[1], out)
    print(json.dumps(result, ensure_ascii=False, indent=2))