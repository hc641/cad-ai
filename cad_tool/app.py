import os
import re
import csv
import json
import uuid
import threading
from io import StringIO
from pathlib import Path
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file, abort
from werkzeug.utils import secure_filename
import openpyxl

from cad_vision_recognizer import process_with_vision
from cad_smart_diff import SmartDiffEngine

app = Flask(__name__, template_folder='templates')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB

UPLOAD_DIR = Path("uploads")
OUTPUT_DIR = Path("outputs")
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# ─────────────────────────────────────────────
# 异步任务注册表（支持进度轮询）
# ─────────────────────────────────────────────
_tasks: dict[str, dict] = {}   # task_id → { status, progress, result, error }
_tasks_lock = threading.Lock()


def _new_task() -> str:
    tid = str(uuid.uuid4())
    with _tasks_lock:
        _tasks[tid] = {"status": "pending", "progress": 0, "result": None, "error": None}
    return tid


def _update_task(tid: str, **kwargs):
    with _tasks_lock:
        if tid in _tasks:
            _tasks[tid].update(kwargs)


def _run_recognition(tid: str, pdf_path: str, out_dir: str):
    """在子线程中运行识别，写入任务状态。"""
    try:
        _update_task(tid, status="running", progress=20)
        result = process_with_vision(
            pdf_path,
            output_dir=out_dir,
            model="qwen-vl-max",
            dpi=130,
            max_tiles=60,
            mock=os.environ.get("MOCK_VISION", "0") == "1",
        )
        _update_task(tid, status="done", progress=100, result={
            "success": True,
            "output": str(result["output"]),
            "pdf": pdf_path,
            "views": result.get("views", []),
            "elements": result.get("elements", 0),
        })
    except Exception as exc:
        _update_task(tid, status="error", error=str(exc))


# ═════════════════════════════════════════════
# 1. 页面路由
# ═════════════════════════════════════════════

@app.route('/')
def index():
    return render_template('compare.html')


# ═════════════════════════════════════════════
# 2. 上传 + 识别（支持同步 & 异步两种调用方式）
# ═════════════════════════════════════════════

@app.route('/api/upload', methods=['POST'])
def api_upload():
    """
    同步接口（原有行为）：上传 PDF → 阻塞识别 → 返回 Excel 路径。
    可加 ?async=1 切换为异步，立即返回 task_id，前端再轮询 /api/task/<id>。
    """
    if 'pdf' not in request.files:
        return jsonify({"error": "未收到文件"}), 400

    file = request.files['pdf']
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({"error": "仅支持 PDF 格式"}), 400

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = secure_filename(Path(file.filename).stem)
    pdf_path = UPLOAD_DIR / f"{stem}_{ts}.pdf"
    file.save(str(pdf_path))

    out_dir = OUTPUT_DIR / f"{stem}_{ts}"
    out_dir.mkdir(exist_ok=True)

    # ── 异步模式 ──
    if request.args.get("async") == "1":
        tid = _new_task()
        t = threading.Thread(target=_run_recognition, args=(tid, str(pdf_path), str(out_dir)), daemon=True)
        t.start()
        return jsonify({"task_id": tid}), 202

    # ── 同步模式（默认） ──
    try:
        result = process_with_vision(
            str(pdf_path),
            output_dir=str(out_dir),
            model="qwen-vl-max",
            dpi=130,
            max_tiles=60,
            mock=os.environ.get("MOCK_VISION", "0") == "1",
        )
        return jsonify({
            "success": True,
            "output": str(result["output"]),
            "pdf": str(pdf_path),
            "views": result.get("views", []),
            "elements": result.get("elements", 0),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/task/<task_id>', methods=['GET'])
def api_task_status(task_id: str):
    """轮询异步任务进度。"""
    with _tasks_lock:
        task = _tasks.get(task_id)
    if not task:
        return jsonify({"error": "任务不存在"}), 404
    return jsonify(task)


# ═════════════════════════════════════════════
# 3. 文件下载
# ═════════════════════════════════════════════

@app.route('/api/download', methods=['GET'])
def api_download():
    """下载识别结果 Excel。"""
    path = request.args.get('path', '').strip()
    if not path:
        return jsonify({"error": "缺少 path 参数"}), 400

    file_path = Path(path).resolve()
    # 安全检查：只允许下载 OUTPUT_DIR 内的文件
    try:
        file_path.relative_to(OUTPUT_DIR.resolve())
    except ValueError:
        abort(403)

    if not file_path.exists():
        return jsonify({"error": "文件不存在"}), 404

    return send_file(
        str(file_path),
        as_attachment=True,
        download_name=file_path.name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ═════════════════════════════════════════════
# 4. 比对引擎
# ═════════════════════════════════════════════

class DrawingComparator:
    """基于两份识别结果 Excel 进行结构化比对。"""

    SHEETS_TO_COMPARE = {
        "📋 检验报告表":   {"key_col": 0, "type": "dimension"},
        "⊕ GD&T关联(AI)": {"key_col": 2, "type": "gdt"},
        "📝 技术要求":     {"key_col": 2, "type": "tech"},
        "📦 物料清单BOM":  {"key_col": 1, "type": "bom"},
        "⭐ 关键特性":     {"key_col": 2, "type": "key"},
        "📋 偏差清单":     {"key_col": 0, "type": "deviation"},
        "📐 尺寸标注":     {"key_col": 2, "type": "raw_dim"},
        "🗺 视图清单(AI)": {"key_col": 1, "type": "view"},
    }

    def __init__(self, path_a: str, path_b: str):
        self.wb_a = openpyxl.load_workbook(path_a, data_only=True, read_only=True)
        self.wb_b = openpyxl.load_workbook(path_b, data_only=True, read_only=True)

    def compare_all(self) -> dict:
        diffs = []

        for sheet_name, cfg in self.SHEETS_TO_COMPARE.items():
            in_a = sheet_name in self.wb_a.sheetnames
            in_b = sheet_name in self.wb_b.sheetnames

            if in_a and in_b:
                diffs.extend(self._compare_sheet(
                    self.wb_a[sheet_name], self.wb_b[sheet_name],
                    cfg["key_col"], cfg["type"], sheet_name
                ))
            elif in_a and not in_b:
                # 整张表被删除
                for r in self._sheet_rows(self.wb_a[sheet_name], cfg["key_col"]):
                    diffs.append({
                        "type": "removed",
                        "content_a": self._row_to_str(r),
                        "content_b": "",
                        "category": sheet_name,
                        "position": str(r[0]) if r else "",
                    })
            elif not in_a and in_b:
                # 整张表新增
                for r in self._sheet_rows(self.wb_b[sheet_name], cfg["key_col"]):
                    diffs.append({
                        "type": "added",
                        "content_a": "",
                        "content_b": self._row_to_str(r),
                        "category": sheet_name,
                        "position": str(r[0]) if r else "",
                    })

        diffs.extend(self._compare_info())

        order = {"modified": 0, "removed": 1, "added": 2}
        diffs.sort(key=lambda x: (order.get(x["type"], 3), x.get("category", "")))

        summary = {
            "added":    sum(1 for d in diffs if d["type"] == "added"),
            "removed":  sum(1 for d in diffs if d["type"] == "removed"),
            "modified": sum(1 for d in diffs if d["type"] == "modified"),
            "total":    len(diffs),
        }
        return {"summary": summary, "diffs": diffs}

    # ── 工具方法 ──────────────────────────────

    def _sheet_rows(self, ws, key_col: int):
        """返回有效数据行列表（跳过空行）。"""
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r and len(r) > key_col and r[key_col] is not None:
                rows.append(r)
        return rows

    def _sheet_to_rows(self, ws, key_col: int) -> dict:
        """工作表 → {key: row}，自动处理重复键。"""
        result = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or len(r) <= key_col:
                continue
            raw_key = r[key_col]
            if raw_key is None:
                continue
            key = str(raw_key).strip()
            if not key:
                continue
            uniq, idx = key, 1
            while uniq in result:
                uniq = f"{key}#{idx}"; idx += 1
            result[uniq] = r
        return result

    def _compare_sheet(self, ws_a, ws_b, key_col: int, type_name: str, label: str) -> list:
        rows_a = self._sheet_to_rows(ws_a, key_col)
        rows_b = self._sheet_to_rows(ws_b, key_col)
        diffs = []

        keys_a, keys_b = set(rows_a), set(rows_b)

        for k in keys_b - keys_a:
            diffs.append({
                "type": "added",
                "content_a": "",
                "content_b": self._row_to_str(rows_b[k]),
                "category": label,
                "position": str(rows_b[k][0]) if rows_b[k] else "",
            })
        for k in keys_a - keys_b:
            diffs.append({
                "type": "removed",
                "content_a": self._row_to_str(rows_a[k]),
                "content_b": "",
                "category": label,
                "position": str(rows_a[k][0]) if rows_a[k] else "",
            })
        for k in keys_a & keys_b:
            if self._rows_differ(rows_a[k], rows_b[k]):
                diffs.append({
                    "type": "modified",
                    "content_a": self._row_to_str(rows_a[k]),
                    "content_b": self._row_to_str(rows_b[k]),
                    "category": label,
                    "position": str(rows_a[k][0]) if rows_a[k] else "",
                })
        return diffs

    def _rows_differ(self, a, b) -> bool:
        # 长度不同 → 肯定不同
        la, lb = len(a), len(b)
        length = max(la, lb)
        for i in range(length):
            va = str(a[i]).strip() if i < la and a[i] is not None else ""
            vb = str(b[i]).strip() if i < lb and b[i] is not None else ""
            if va != vb:
                return True
        return False

    def _row_to_str(self, row, max_len: int = 150) -> str:
        parts = [str(v) for v in row if v is not None and str(v).strip()]
        txt = " | ".join(parts)
        return txt if len(txt) <= max_len else txt[:max_len] + "…"

    def _compare_info(self) -> list:
        sn = "📋 图纸信息"
        if sn not in self.wb_a.sheetnames or sn not in self.wb_b.sheetnames:
            return []
        info_a = self._sheet_to_dict(self.wb_a[sn])
        info_b = self._sheet_to_dict(self.wb_b[sn])
        diffs = []
        for k in set(info_a) | set(info_b):
            va = info_a.get(k, "")
            vb = info_b.get(k, "")
            if va.strip() != vb.strip():
                diff_type = "modified" if va and vb else ("removed" if va else "added")
                diffs.append({
                    "type": diff_type,
                    "content_a": f"{k}: {va}" if va else "",
                    "content_b": f"{k}: {vb}" if vb else "",
                    "category": "图纸信息",
                    "position": "标题栏",
                })
        return diffs

    def _sheet_to_dict(self, ws) -> dict:
        d = {}
        for r in ws.iter_rows(values_only=True):
            if r and len(r) >= 2 and r[0] is not None and r[1] is not None:
                d[str(r[0]).strip()] = str(r[1]).strip()
        return d


# ═════════════════════════════════════════════
# 5. 比对接口
# ═════════════════════════════════════════════

@app.route('/api/compare/run', methods=['POST'])
def api_compare():
    data = request.get_json(force=True)
    path_a = (data.get("path_a") or "").strip()
    path_b = (data.get("path_b") or "").strip()

    if not path_a or not path_b:
        return jsonify({"error": "缺少 path_a 或 path_b"}), 400

    for p in (path_a, path_b):
        if not os.path.exists(p):
            return jsonify({"error": f"文件不存在：{p}"}), 400
        # 安全检查
        try:
            Path(p).resolve().relative_to(OUTPUT_DIR.resolve())
        except ValueError:
            return jsonify({"error": "路径非法"}), 403

    try:
        comp = DrawingComparator(path_a, path_b)
        result = comp.compare_all()
        return jsonify({"success": True, **result})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/compare/smart', methods=['POST'])
def api_smart_compare():
    """智能结构化比对（实体匹配 + 噪声过滤）"""
    data = request.get_json(force=True)
    path_a = (data.get("path_a") or "").strip()
    path_b = (data.get("path_b") or "").strip()

    if not path_a or not path_b:
        return jsonify({"error": "缺少 path_a 或 path_b"}), 400

    for p in (path_a, path_b):
        if not os.path.exists(p):
            return jsonify({"error": f"文件不存在：{p}"}), 400
        try:
            Path(p).resolve().relative_to(OUTPUT_DIR.resolve())
        except ValueError:
            return jsonify({"error": "路径非法"}), 403

    try:
        engine = SmartDiffEngine(path_a, path_b)
        result = engine.compare_all()
        return jsonify({"success": True, **result})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/compare/smart_csv', methods=['POST'])
def api_smart_compare_csv():
    """智能比对 CSV 导出"""
    data = request.get_json(force=True)
    path_a = (data.get("path_a") or "").strip()
    path_b = (data.get("path_b") or "").strip()

    if not path_a or not path_b:
        return jsonify({"error": "缺少 path_a 或 path_b"}), 400

    for p in (path_a, path_b):
        if not os.path.exists(p):
            return jsonify({"error": f"文件不存在：{p}"}), 400
        try:
            Path(p).resolve().relative_to(OUTPUT_DIR.resolve())
        except ValueError:
            return jsonify({"error": "路径非法"}), 403

    try:
        engine = SmartDiffEngine(path_a, path_b)
        result = engine.compare_all()
        csv_content = engine.to_csv(result)
        # 返回 CSV 文本内容，前端可自行保存
        return jsonify({"success": True, "csv": csv_content})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ═════════════════════════════════════════════
# 6. 历史记录接口
# ═════════════════════════════════════════════

@app.route('/api/history', methods=['GET'])
def api_history():
    """返回最近 50 条识别输出记录（按修改时间倒序）。"""
    records = []
    for p in sorted(OUTPUT_DIR.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True)[:50]:
        if not p.is_dir():
            continue
        xlsx_files = list(p.glob("*.xlsx"))
        if not xlsx_files:
            continue
        first = xlsx_files[0]
        records.append({
            "name": p.name,
            "path": str(first),
            "size": first.stat().st_size,
            "mtime": datetime.fromtimestamp(first.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
        })
    return jsonify({"records": records})


# ═════════════════════════════════════════════
# 7. 启动
# ═════════════════════════════════════════════

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)