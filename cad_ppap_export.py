"""
CAD识别系统 - PPAP 标准格式导出

对齐官方 PPAP 表格格式（参考 35957399_2026-1-20_PPAP.xls）：
  - 头部：Part Number / ECL / Date / Part Name / Tool No. / Cavities /
           Material / Material Spec / Report Requested By 等
  - 双行表头：Dim No. / Ref Only / Drawing Dimension / Lower / Upper /
              1A~4A / Discrep Col# / Fix Tool / Will Change Dwg. To
  - 数据行：SEE GD&T 内嵌，倍数展开行 dim_no 为空
  - 页尾：Inspection Source / Inspected by / Inspector Supervisor /
           Approved by 签名区

用法：
  python cad_ppap_export.py 识别结果.xlsx --cavities 4
"""

import re
import argparse
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 颜色常量 ─────────────────────────────────
C_HEAD  = "1F3864"   # 深蓝表头
C_HDR2  = "D6DCE4"   # 浅灰副表头
C_INFO  = "BDD7EE"   # 信息区浅蓝
C_FOOT  = "EDEDED"   # 页尾浅灰
C_W     = "FFFFFF"
C_BLACK = "000000"


def _thin_border():
    s = Side(style="thin", color=C_BLACK)
    return Border(left=s, right=s, top=s, bottom=s)


def _medium_border():
    m = Side(style="medium", color=C_BLACK)
    return Border(left=m, right=m, top=m, bottom=m)


def _cell(ws, row, col, value="", bold=False, size=10, color=C_BLACK,
          bg=None, halign="left", valign="center", wrap=False, border=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=halign, vertical=valign,
                             wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if border:
        c.border = border
    return c


def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2,   end_column=c2)


# ── 从检验报告表读取数据 ──────────────────────
def _read_source(wb):
    """
    读取检验报告表，兼容新旧两种格式：

    旧格式（dim_no 始终有值，展开行用 '7-1','7-2'...）
    新格式（展开行 dim_no 为 None）

    SEE GD&T 行内嵌在数据流中，不分离到末尾。

    数据清洗（对齐PDF图纸）：
      - 跳过粗糙度行、旧式 GD&T: 行
      - 跳过角度标注行（含 ° 或 " 符号，如 6X 1°±1'）——角度不属于线性尺寸检验项
      - 跳过 lower==upper 的异常行（公差解析失败，如 6X R 0.4±0. → Lower=Upper=0.4）
      - 按 (drawing, lower, upper) 跨视图去重（同尺寸在多个视图各识别一次时只保留一组）
    """
    sn = "📋 检验报告表"
    if sn not in wb.sheetnames:
        return []

    def _normalize_drawing(draw, raw):
        """
        将 Drawing Dimension 规范化为官方 PPAP 模式：
          - 去掉倍数前缀 2X / 4X / 6X / 12X 等
          - 去掉 ± 公差部分（公差已在 Lower/Upper 列体现）
          - 直径 Ø/Φ/⌀/D → 'D' 前缀（如 Ø4.5 → D4.5）
          - 半径 R → 'R' 前缀（如 6X R 0.3 → R0.3）
          - 粗糙度 Ra/Rz → 保留前缀
          - 纯线性尺寸 → 纯数字（如 2X 1.15 → 1.15）
        """
        if draw == "SEE GD&T":
            return draw
        s   = str(draw).strip()
        src = str(raw) if raw else s

        # 去掉倍数前缀（如 2X / 4X / 12X）
        s = re.sub(r"^\s*\d+\s*[Xx×]\s*", "", s)
        # 去掉 ± 及之后的公差
        s = re.split(r"[±]", s)[0].strip()

        is_dia    = bool(re.search(r"[ØΦ⌀]", src) or re.match(r"^D[\d.]", s))
        is_radius = bool(re.match(r"^\s*R\b", s) or re.search(r"\bR\s*[\d.]", src))
        is_rough  = bool(re.match(r"(?i)^(ra|rz|rmax)\b", s))

        num_m = re.search(r"\d+\.?\d*", s)
        num   = num_m.group(0) if num_m else s

        if is_rough:
            pre = re.match(r"(?i)^(ra|rz|rmax)", s)
            return (pre.group(0) + num) if (pre and num_m) else s
        if is_dia:
            return ("D" + num) if num_m else s
        if is_radius:
            return ("R" + num) if num_m else s
        return num if num_m else s

    # ── 第一遍：收集原始行，做清洗 + 跨组去重 ──
    raw_items = []
    seen_content = set()   # (drawing, lower, upper) 已出现的内容，用于去重

    for r in wb[sn].iter_rows(min_row=2, values_only=True):
        if len(r) < 5:
            continue

        dim_no_raw = r[0]
        ref        = "Ref" if str(r[1] or "").strip() in ("✓", "Ref") else ""
        draw_raw   = str(r[2] or "").strip()
        lower      = r[3]
        upper      = r[4]
        raw_orig   = str(r[6] or "").strip() if len(r) > 6 else ""

        # 跳过粗糙度行
        if draw_raw.lower().startswith("粗糙度:"):
            continue
        # 跳过旧式 GD&T: 行
        if draw_raw.lower().startswith("gd&t:"):
            continue

        is_gdt_row = (draw_raw == "SEE GD&T")

        if not is_gdt_row:
            # 跳过角度标注行：drawing 或 raw 含度/分/秒符号
            combined = draw_raw + " " + raw_orig
            if "°" in combined or '"' in draw_raw or "′" in combined or "″" in combined:
                # 但排除 X45°（倒角，属正常尺寸）—— 仅当含独立角度公差时才跳
                if re.search(r"\d\s*°", combined) and "X45" not in combined.upper():
                    continue

            # 过滤：lower/upper 均空 → 无验收限
            lower_empty = lower is None or str(lower).strip() == ""
            upper_empty = upper is None or str(upper).strip() == ""
            if lower_empty and upper_empty:
                continue

            # 过滤：lower == upper（公差解析失败，如 ±0. 截断）
            try:
                if not lower_empty and not upper_empty and float(lower) == float(upper):
                    continue
            except (ValueError, TypeError):
                pass

            # 跨视图去重：仅对"有编号的行"去重
            # （dim_no 为 None 的是倍数展开行，必须保留，否则丢失展开实例）
            dim_no_present = (dim_no_raw is not None
                              and str(dim_no_raw).strip() != "")
            norm_draw_val = _normalize_drawing(draw_raw, raw_orig)
            if dim_no_present:
                content_key = (norm_draw_val, str(lower), str(upper))
                if content_key in seen_content:
                    continue
                seen_content.add(content_key)
        else:
            norm_draw_val = draw_raw   # SEE GD&T 原样

        raw_items.append({
            "dim_no_raw": dim_no_raw,
            "ref":        ref,
            "drawing":    norm_draw_val,
            "lower":      lower,
            "upper":      upper,
        })

    # ── 第二遍：分配连续编号，处理倍数展开行 ──
    dims = []
    seq = 0
    last_seq_key = None

    for it in raw_items:
        dim_no_str    = str(it["dim_no_raw"]).strip() if it["dim_no_raw"] is not None else ""
        dim_no_is_set = dim_no_str != ""

        if dim_no_is_set:
            base      = re.sub(r"-\d+$", "", dim_no_str)
            group_key = (base, it["drawing"])
            if group_key != last_seq_key:
                seq += 1
                last_seq_key = group_key
                assign_no = f"{seq}.0"
            else:
                assign_no = None
        else:
            assign_no = None

        dims.append({
            "dim_no":  assign_no,
            "ref":     it["ref"],
            "drawing": it["drawing"],
            "lower":   it["lower"],
            "upper":   it["upper"],
        })

    return dims


# ── 从图纸信息 / 材料信息 / 修订历史读取头部信息 ──
def _read_info(wb):
    info = {}

    # 图纸信息
    sn = "📋 图纸信息"
    if sn in wb.sheetnames:
        for r in wb[sn].iter_rows(values_only=True):
            if not (r[0] and r[1]):
                continue
            k, v = str(r[0]), str(r[1])
            if re.fullmatch(r"\d+\s*条", v):
                continue
            if   "图号"     in k: info["part_number"] = v
            elif "图名"     in k: info["part_name"]   = v
            elif "最新版本" in k: info["rev"]          = v

    # 材料：优先从 🔬材料信息 sheet 取基材
    sn_mat = "🔬 材料信息"
    if sn_mat in wb.sheetnames:
        for r in wb[sn_mat].iter_rows(values_only=True):
            if r[0] and r[1] and "基材" in str(r[0]):
                info.setdefault("material", str(r[1]))
                break
    # 回退：图纸信息里的材料字段
    if "material" not in info and sn in wb.sheetnames:
        for r in wb[sn].iter_rows(values_only=True):
            if r[0] and r[1] and "材料" in str(r[0]):
                v = str(r[1])
                if not re.fullmatch(r"\d+\s*条", v):
                    info["material"] = v
                    break

    # ECL：从修订历史最后一个 R 版本提取数字
    sn2 = "🔄 修订历史"
    if sn2 in wb.sheetnames:
        valid = [r for r in wb[sn2].iter_rows(min_row=2, values_only=True)
                 if r[0] and str(r[0]).strip().startswith("R")]
        if valid:
            m = re.search(r"\d+", str(valid[-1][0]))
            if m:
                info["ecl"] = m.group(0).zfill(2)

    return info


# ── 主构建函数 ────────────────────────────────
def _build(wb, cavities):
    dims = _read_source(wb)
    info = _read_info(wb)

    if "PPAP" in wb.sheetnames:
        del wb["PPAP"]
    ws   = wb.create_sheet("PPAP", 0)
    thin = _thin_border()

    # 总列数：Dim No.(1) + Ref(2) + Drawing(3) + Lower(4) + Upper(5)
    #         + cavities列 + Discrep(DC) + Fix Tool(DC+1) + Will Change(DC+2)
    CS = 6                   # cavity 起始列
    DC = CS + cavities       # Discrep 列
    TC = DC + 2              # 最后一列（Will Change Dwg. To）

    part_number = info.get("part_number", "")
    ecl         = info.get("ecl", "")
    date_str    = datetime.now().strftime("%d%b%y").upper()
    part_name   = info.get("part_name", "")
    material    = info.get("material", "")
    rev         = info.get("rev", "")

    # ══════════════════════════════════════════
    # 行1  大标题
    # ══════════════════════════════════════════
    _merge(ws, 1, 1, 1, TC)
    _cell(ws, 1, 1, "PART INSPECTION REPORT",
          bold=True, size=14, color=C_W, bg=C_HEAD,
          halign="center", border=thin)
    ws.row_dimensions[1].height = 28

    # ══════════════════════════════════════════
    # 行3  Part Certification
    # ══════════════════════════════════════════
    _merge(ws, 3, 1, 3, TC)
    _cell(ws, 3, 1, "Part Certification",
          bold=True, size=10, bg=C_HDR2, border=thin)
    ws.row_dimensions[3].height = 16

    # ══════════════════════════════════════════
    # 行4-5  Part Number / ECL / Date / Part Name
    # ══════════════════════════════════════════
    def info_label(row, col, span_end, label):
        _merge(ws, row, col, row, span_end)
        _cell(ws, row, col, label, bold=True, size=9,
              bg=C_HDR2, border=thin)

    def info_val(row, col, span_end, value):
        _merge(ws, row, col, row, span_end)
        _cell(ws, row, col, value, size=10, border=thin)

    # 行4：标签行
    info_label(4, 1, 6, "Part Number")
    info_label(4, 7, 7, "ECL")
    info_label(4, 8, 8, "Date")
    info_label(4, 9, TC, "Part Name")
    # 行5：值行
    info_val(5, 1, 6, part_number)
    info_val(5, 7, 7, ecl)
    info_val(5, 8, 8, date_str)
    info_val(5, 9, TC, part_name)
    for row in (4, 5):
        ws.row_dimensions[row].height = 16

    # ══════════════════════════════════════════
    # 行6-7  Tool No. / Cavities / Cav.# / Material
    # ══════════════════════════════════════════
    info_label(6, 1, 6, "Die Master Mold No. / Tool No.")
    info_label(6, 7, 8, "No. of Tool Cavities")
    info_label(6, 9, TC, "Cav. #")

    info_val(7, 1, 6, "")          # 用户填写
    info_val(7, 7, 8, str(cavities))
    cav_labels = "-".join([f"{i+1}A" for i in range(cavities)])
    info_val(7, 9, TC, cav_labels)
    for row in (6, 7):
        ws.row_dimensions[row].height = 16

    # ══════════════════════════════════════════
    # 行8-9  Material / Material Spec / Regrind
    # ══════════════════════════════════════════
    info_label(8, 1, 4, "Material")
    info_label(8, 5, 8, "Material Spec")
    info_label(8, 9, 11, "Max % Regrind")
    info_label(8, 12, TC, "Actual % Regrind")

    info_val(9, 1, 4, material)
    info_val(9, 5, 8, "")          # 用户填写
    info_val(9, 9, 11, "N/A")
    info_val(9, 12, TC, "N/A")
    for row in (8, 9):
        ws.row_dimensions[row].height = 16

    # ══════════════════════════════════════════
    # 行10-11  Report Requested By / Rev
    # ══════════════════════════════════════════
    info_label(10, 1, 6, "Report Requested By")
    info_label(10, 7, TC, "Drawing Rev Level")
    info_val(11, 1, 6, "")         # 用户填写
    info_val(11, 7, TC, rev)
    for row in (10, 11):
        ws.row_dimensions[row].height = 16

    # ══════════════════════════════════════════
    # 行12-13  Additional Eng. Changes
    # ══════════════════════════════════════════
    info_label(12, 1, TC, "Drawing is being changed by Product Engineering per ECN No.")
    info_label(13, 1, 4, "Additional Eng. Changes")
    info_val(13, 5, TC, "N/A")
    for row in (12, 13):
        ws.row_dimensions[row].height = 15

    # ══════════════════════════════════════════
    # 行14 空行分隔
    # ══════════════════════════════════════════
    ws.row_dimensions[14].height = 6

    # ══════════════════════════════════════════
    # 行15-16  双行表头
    # ══════════════════════════════════════════
    HR1, HR2 = 15, 16

    def mh(r1, c1, r2, c2, val, wrap=False):
        if r1 != r2 or c1 != c2:
            _merge(ws, r1, c1, r2, c2)
        c = ws.cell(r1, c1, val)
        c.font      = Font(name="Arial", bold=True, size=9, color=C_W)
        c.fill      = PatternFill("solid", fgColor=C_HEAD)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=wrap)
        c.border    = thin

    mh(HR1, 1,  HR2, 1,  "Dim\nNo.",   wrap=True)
    mh(HR1, 2,  HR2, 2,  "Ref\nOnly",  wrap=True)
    mh(HR1, 3,  HR2, 3,  "Drawing\nDimension", wrap=True)
    mh(HR1, 4,  HR1, 5,  "Acceptance")
    mh(HR2, 4,  HR2, 4,  "Lower")
    mh(HR2, 5,  HR2, 5,  "Upper")
    mh(HR1, CS, HR1, CS + cavities - 1, "PARTS / TOOL CAVITIES CHECKED")
    for k in range(cavities):
        mh(HR2, CS + k, HR2, CS + k, f"{k+1}A")
    mh(HR1, DC,   HR2, DC,   "Discrep\nCol #",    wrap=True)
    mh(HR1, DC+1, HR2, DC+1, "Fix\nTool",         wrap=True)
    mh(HR1, DC+2, HR2, DC+2, "Will Change\nDwg. To", wrap=True)

    ws.row_dimensions[HR1].height = 20
    ws.row_dimensions[HR2].height = 16

    # ══════════════════════════════════════════
    # 数据行
    # ══════════════════════════════════════════
    r = HR2 + 1
    sm = Font(name="Arial", size=9)

    def style_data_row(row_num, is_gdt=False):
        for cc in range(1, TC + 1):
            c = ws.cell(row_num, cc)
            c.font      = sm
            c.border    = thin
            c.alignment = Alignment(horizontal="center", vertical="center")
        if is_gdt:
            # GD&T 行轻微底色区分
            for cc in range(1, TC + 1):
                ws.cell(row_num, cc).fill = PatternFill(
                    "solid", fgColor="EBF3FB")

    seq_count = 0
    for item in dims:
        if item["dim_no"] is not None:
            seq_count += 1
        is_gdt = (item["drawing"] == "SEE GD&T")
        ws.cell(r, 1, item["dim_no"])
        ws.cell(r, 2, item["ref"])
        ws.cell(r, 3, item["drawing"])
        ws.cell(r, 4, item["lower"])
        ws.cell(r, 5, item["upper"])
        style_data_row(r, is_gdt=is_gdt)
        r += 1

    # ══════════════════════════════════════════
    # 页尾签名区
    # ══════════════════════════════════════════
    r += 1   # 空一行

    def footer_label(row, col, cend, text):
        _merge(ws, row, col, row, cend)
        _cell(ws, row, col, text, bold=True, size=9,
              bg=C_FOOT, border=thin)
        ws.row_dimensions[row].height = 15

    def footer_val(row, col, cend):
        _merge(ws, row, col, row, cend)
        _cell(ws, row, col, "", size=10, border=thin)
        ws.row_dimensions[row].height = 18

    footer_label(r,   1, TC, "Inspection Source Company Name")
    footer_val(r+1,   1, TC)

    footer_label(r+2, 1, 7,  "Inspected by")
    footer_label(r+2, 8, 11, "Title")
    footer_label(r+2, 12, TC, "Inspection Report Date")
    footer_val(r+3,   1, 7)
    footer_val(r+3,   8, 11)
    _merge(ws, r+3, 12, r+3, TC)
    _cell(ws, r+3, 12, date_str, size=10, border=thin)
    ws.row_dimensions[r+3].height = 18

    footer_label(r+4, 1, 7,  "Inspector Supervisor")
    footer_label(r+4, 8, TC, "Title")
    footer_val(r+5,   1, 7)
    footer_val(r+5,   8, TC)

    footer_label(r+6, 1, 7,  "Approved by")
    footer_label(r+6, 8, 11, "Title")
    footer_label(r+6, 12, TC, "Date")
    footer_val(r+7,   1, 7)
    footer_val(r+7,   8, 11)
    footer_val(r+7,   12, TC)

    # ══════════════════════════════════════════
    # 列宽
    # ══════════════════════════════════════════
    col_widths = {1: 8, 2: 7, 3: 20, 4: 10, 5: 10}
    for k in range(cavities):
        col_widths[CS + k] = 10
    col_widths[DC]   = 9
    col_widths[DC+1] = 10
    col_widths[DC+2] = 18
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = ws.cell(HR2 + 1, 1)
    return seq_count


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--cavities", type=int, default=4)
    args = ap.parse_args()
    print(f"[1/3] 读取: {args.xlsx}")
    wb = openpyxl.load_workbook(args.xlsx)
    print(f"[2/3] 生成PPAP (型腔数={args.cavities})...")
    n = _build(wb, args.cavities)
    print(f"      → 共 {n} 个尺寸编号")
    wb.save(args.xlsx)
    print(f"[3/3] 完成 → {args.xlsx}")


if __name__ == "__main__":
    main()